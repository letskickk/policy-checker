"""
개혁신당 공약 멘토링 API. 공약 텍스트를 받아 GPT 기반 부합 점검 결과를 반환한다.
접근제어: 회원가입→관리자 승인→쿼터/레이트리밋 적용.
"""
import json
import locale
import logging
import os
import re
import sys
import uuid
from datetime import datetime
from pathlib import Path

RANKING_SCORE_START_DATE = datetime(2026, 4, 13).date()
from typing import Literal, Optional

from fastapi import FastAPI, HTTPException, Query, Request, UploadFile, File
from fastapi.responses import FileResponse, RedirectResponse, StreamingResponse, Response
from pydantic import BaseModel, Field

from backend.config import (
    ADMIN_EMAILS,
    ROOT_DIR,
    PDF_DIR,
    INDEX_CACHE_DIR,
    OPENAI_MODEL,
    CHAT_MODEL,
    DEBUG_ENDPOINTS_ENABLED,
    PDF_S3_URI,
    USE_OPENAI_VECTOR_STORE,
    SKIP_PDF_SCAN_ON_STARTUP,
    OPENAI_VECTOR_STORE_ID,
    DATA_GO_KR_API_KEY,
    _nfc,
    POLICY_DRAFTER_TEST_EMAILS,
)
from backend.auth import (
    STATUS_APPROVED,
    STATUS_PENDING,
    ROLE_ADMIN,
    create_session_token,
    verify_session_token,
    signup as auth_signup,
    login as auth_login,
    verify_email_token,
    resend_verification_email,
    get_user,
    list_users_pending,
    list_users_all,
    set_user_status,
)
from backend.usage_logger import log_usage
from backend.quota_rate import check_rate_limit_ip, check_rate_limit_user
# 무거운 import는 지연 로딩으로 변경
# from backend.database import init_db
# from backend.pdf_loader import (
#     HAS_PDFPLUMBER,
#     _iter_doc_files,
#     load_platform_context,
#     load_pledges_context,
#     get_context_summary,
# )
# from backend.index_builder import build_all_indexes

# 로깅 설정
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

PUBLIC_NOMINATION_NOTE = "공천 확정"


def _normalize_header_label(value) -> str:
    text = str(value or "").strip().lower()
    return re.sub(r"[\s_/()\-]+", "", text)


def _applicant_cell_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _excel_truthy(value) -> int:
    normalized = _normalize_header_label(value)
    return int(normalized in {"y", "yes", "true", "1", "o", "v", "예", "완료", "제출", "있음"})


def _extract_applicants_from_workbook(content: bytes, filename: str) -> tuple[list[dict], int]:
    if not filename or not filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="지원서 엑셀은 .xlsx 파일만 업로드 가능합니다.")

    try:
        import io
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="서버에 openpyxl이 설치되지 않았습니다. pip install openpyxl")

    try:
        workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"엑셀 파일을 열 수 없습니다: {str(exc)[:100]}")

    try:
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            raise HTTPException(status_code=400, detail="엑셀 첫 줄에서 헤더를 찾을 수 없습니다.")

        normalized_headers = [_normalize_header_label(cell) for cell in header_row]
        aliases = {
            "name": {"이름", "성명", "후보자명", "지원자명"},
            "phone": {"휴대폰", "휴대전화", "전화번호", "연락처", "핸드폰"},
            "email": {"이메일", "email", "메일"},
            "region_province": {"시도", "지역", "광역시도", "시도명"},
            "district_info": {"선거구", "선거구정보", "지역구", "출마지역", "선거구명"},
            "election_position": {"출마직", "출마직위", "선거직", "직책", "직위"},
            "doc_submitted": {"서류제출", "서류", "제출여부", "서류제출여부"},
            "interview_done": {"면접", "면접완료", "면접여부"},
            "status_note": {"상태", "상태메모", "비고", "메모", "진행상태"},
        }

        column_map: dict[str, int] = {}
        for field, candidates in aliases.items():
            for idx, header in enumerate(normalized_headers):
                if header in {_normalize_header_label(item) for item in candidates}:
                    column_map[field] = idx
                    break

        if "name" not in column_map:
            if len(header_row) < 9:
                raise HTTPException(status_code=400, detail="지원자 이름 컬럼을 찾을 수 없습니다. 업로드 양식을 확인해주세요.")
            column_map = {
                "name": 1,
                "phone": 4,
                "email": 5,
                "region_province": 6,
                "district_info": 7,
                "election_position": 8,
                "doc_submitted": 9,
                "interview_done": 10,
                "status_note": 11,
            }

        def cell(row, key: str) -> str:
            idx = column_map.get(key)
            if idx is None or idx >= len(row):
                return ""
            return _applicant_cell_str(row[idx])

        applicants: list[dict] = []
        skipped_rows = 0
        for row in rows_iter:
            if not row or not any(item not in (None, "") for item in row):
                continue
            name = cell(row, "name")
            if not name:
                skipped_rows += 1
                continue
            applicants.append(
                {
                    "name": name,
                    "phone": cell(row, "phone"),
                    "email": cell(row, "email"),
                    "region_province": cell(row, "region_province"),
                    "district_info": cell(row, "district_info"),
                    "election_position": cell(row, "election_position"),
                    "doc_submitted": _excel_truthy(cell(row, "doc_submitted")),
                    "interview_done": _excel_truthy(cell(row, "interview_done")),
                    "status_note": cell(row, "status_note"),
                }
            )

        if not applicants:
            raise HTTPException(status_code=400, detail="업로드된 엑셀에서 저장할 지원자 행을 찾지 못했습니다.")

        return applicants, skipped_rows
    finally:
        workbook.close()


def _sql_normalized_phone_expr(expr: str) -> str:
    return f"replace(replace(replace(replace(trim(coalesce({expr}, '')), '-', ''), ' ', ''), '(', ''), ')', '')"


def _sql_normalized_name_expr(expr: str) -> str:
    return f"lower(replace(trim(coalesce({expr}, '')), ' ', ''))"


def _sql_public_nomination_condition(user_alias: str = "u") -> str:
    return f"""
        (
            {user_alias}.applicant_match_id IS NOT NULL
            OR TRIM(COALESCE(pa.status_note, '')) = ''
            OR EXISTS (
                SELECT 1
                FROM party_applicants pa_public
                WHERE TRIM(COALESCE(pa_public.status_note, '')) = '{PUBLIC_NOMINATION_NOTE}'
                  AND (
                      (
                          lower(trim(coalesce(pa_public.email, ''))) <> ''
                          AND lower(trim(coalesce(pa_public.email, ''))) = lower(trim(coalesce({user_alias}.email, '')))
                          AND {_sql_normalized_name_expr('pa_public.name')} = {_sql_normalized_name_expr(f'{user_alias}.name')}
                      )
                      OR (
                          {_sql_normalized_phone_expr('pa_public.phone')} <> ''
                          AND {_sql_normalized_phone_expr('pa_public.phone')} = {_sql_normalized_phone_expr(f'{user_alias}.phone')}
                          AND {_sql_normalized_name_expr('pa_public.name')} = {_sql_normalized_name_expr(f'{user_alias}.name')}
                      )
                  )
            )
        )
    """

app = FastAPI(
    title="개혁신당 공약 멘토링",
    description="출마자 공약의 중앙당 정강정책·공약과의 적합도 점검 API",
    version="1.3.0",
)

# 서버 시작 시 즉시 출력
print("=" * 60, flush=True)
print("FastAPI 앱 생성 완료", flush=True)
print("서버가 시작됩니다...", flush=True)
print("=" * 60, flush=True)

# 전역 인덱스 (서버 시작 시 초기화). USE_OPENAI_VECTOR_STORE=1이면 _vector_store_id 사용.
_indexes = None
_vector_store_id = None
_regional_vector_store_id = None
_winners2022_vector_store_id = None


def _startup_self_check() -> int:
    """
    서버 시작 시 강제 진단. 조건 불만족 시 RuntimeError.
    SKIP_PDF_SCAN_ON_STARTUP=1 + USE_OPENAI_VECTOR_STORE + OPENAI_VECTOR_STORE_ID 설정 시 PDF 스캔 생략.
    Returns: 공약 폴더 PDF 개수 (0이면 그대로 raise, skip 시 0 반환)
    """
    if USE_OPENAI_VECTOR_STORE and SKIP_PDF_SCAN_ON_STARTUP and OPENAI_VECTOR_STORE_ID:
        logger.info("[SELF-CHECK] SKIP_PDF_SCAN_ON_STARTUP=1 → PDF 스캔 생략")
        return 0

    # locale 확인: UTF-8 아님 → fail-fast
    enc = locale.getpreferredencoding()
    try:
        lc = locale.setlocale(locale.LC_ALL, None)
    except Exception:
        lc = "unknown"
    logger.info(f"[LOCALE] encoding={enc} LC_ALL={lc}")
    # Linux/컨테이너에서만 UTF-8 강제 (한글 rglob용). Windows(cp949)는 통과.
    if sys.platform != "win32" and enc.upper() not in ("UTF-8", "UTF8"):
        raise RuntimeError(
            f"UTF-8 locale이 필요합니다. 현재 encoding={enc}. "
            "Dockerfile에 ENV LANG=C.UTF-8 LC_ALL=C.UTF-8 또는 export LC_ALL=C.UTF-8 을 설정하세요."
        )

    # 경로
    cwd = os.getcwd()
    try:
        backend_file = Path(__file__).resolve()
        base_dir = backend_file.parent.parent
    except Exception:
        base_dir = Path(cwd)
    logger.info(f"[SELF-CHECK] cwd={cwd!r}, __file__ base={base_dir!s}")

    pdf_dir = Path(PDF_DIR).resolve()
    pdf_dir_exists = pdf_dir.exists()
    logger.info(f"[SELF-CHECK] PDF_DIR={pdf_dir!s}, exists={pdf_dir_exists}")

    folders = [
        ("정강정책", pdf_dir / _nfc("정강정책")),
        ("공약", pdf_dir / _nfc("공약")),
        ("지역별 공약", pdf_dir / _nfc("지역별 공약")),
    ]
    pledge_pdf_count = 0
    for name, dir_path in folders:
        exists = dir_path.exists()
        try:
            raw_entries = list(dir_path.iterdir())[:5] if exists else []
            logger.info(f"[SCAN RAW] {name} iterdir sample={[str(p) for p in raw_entries]}")
            pdf_list = list(_iter_doc_files(dir_path)) if exists else []
            logger.info(f"[SCAN DOC] {name} pdf+txt count={len(pdf_list)}")
        except Exception as e:
            logger.warning(f"[SELF-CHECK] {name} rglob failed: {e}")
            pdf_list = []
        count = len(pdf_list)
        samples = [p.name for p in sorted(pdf_list)[:5]]
        if name == "공약":
            pledge_pdf_count = count
        has_sin_gu = any("신구연금" in p.name for p in pdf_list)
        logger.info(f"[SELF-CHECK] {name} exists={exists} pdf_count={count} sample={samples!r} 신구연금포함={has_sin_gu}")

    if not HAS_PDFPLUMBER:
        raise RuntimeError("HAS_PDFPLUMBER is False. pdfplumber is required. Install: pip install pdfplumber pdfminer.six")
    logger.info("[SELF-CHECK] HAS_PDFPLUMBER=True")

    cache_dir = Path(INDEX_CACHE_DIR).resolve()
    cache_dir.mkdir(parents=True, exist_ok=True)
    cache_exists = cache_dir.exists()
    writable = False
    try:
        touch = cache_dir / ".write_test"
        touch.write_text("ok")
        touch.unlink(missing_ok=True)
        writable = True
    except Exception as e:
        logger.error(f"[SELF-CHECK] INDEX_CACHE_DIR not writable: {cache_dir} - {e}")
    logger.info(f"[SELF-CHECK] INDEX_CACHE_DIR={cache_dir!s} exists={cache_exists} writable={writable}")
    if not writable:
        raise RuntimeError(f"INDEX_CACHE_DIR is not writable: {cache_dir}")

    if pledge_pdf_count == 0 and not PDF_S3_URI:
        raise RuntimeError(
            "공약 폴더 PDF 개수가 0입니다. AWS에 PDF를 배포했는지 확인하세요. "
            "또는 PDF_S3_URI를 설정해 S3에서 내려받도록 하세요."
        )
    logger.info(f"[SELF-CHECK] 공약 pdf count={pledge_pdf_count} (>0 or PDF_S3_URI set)")

    if pledge_pdf_count == 0 and PDF_S3_URI:
        try:
            import subprocess
            pdf_dir.mkdir(parents=True, exist_ok=True)
            pdf_pledge = pdf_dir / _nfc("공약")
            pdf_pledge.mkdir(parents=True, exist_ok=True)
            subprocess.run(
                ["aws", "s3", "sync", PDF_S3_URI.rstrip("/") + "/", str(pdf_pledge)],
                check=True,
                timeout=300,
                capture_output=True,
            )
            pledge_pdf_count = len(list(_iter_doc_files(pdf_pledge)))
            logger.info(f"[SELF-CHECK] S3 sync done, 공약 pdf count={pledge_pdf_count}")
        except Exception as e:
            raise RuntimeError(f"PDF_S3_URI sync failed: {e}") from e
        if pledge_pdf_count == 0:
            raise RuntimeError("S3 sync 후에도 공약 폴더 PDF가 0건입니다.")

    return pledge_pdf_count


_startup_done = False
_db_ready = False


def _ensure_db_ready():
    """후보/관리자 API용 경량 초기화: DB만 보장. 최초 1회 전원 재검증."""
    global _db_ready
    if _db_ready:
        return
    from backend.database import init_db, get_connection
    init_db()
    _db_ready = True

    # 서버 시작 시 전원 재검증 (party_applicants 데이터가 있을 때만)
    try:
        conn = get_connection()
        has_data = conn.execute("SELECT 1 FROM party_applicants LIMIT 1").fetchone()
        conn.close()
        if has_data:
            from backend.applicant_verify import reverify_all_users
            n = reverify_all_users()
            logger.info("서버 시작 시 전원 재검증 완료: %d명", n)
    except Exception:
        logger.debug("서버 시작 시 재검증 스킵 (party_applicants 미생성)")

def _ensure_startup():
    """지연 초기화: 첫 요청 시 한 번만 실행."""
    global _indexes, _vector_store_id, _regional_vector_store_id, _winners2022_vector_store_id, _startup_done, _db_ready
    
    if _startup_done:
        return
    
    import traceback
    
    # 지연 import
    from backend.database import init_db
    from backend.index_builder import build_all_indexes
    from backend.vector_index import VectorIndex
    from backend.config import EMBEDDING_DIMENSION, OPENAI_REGIONAL_VECTOR_STORE_ID
    
    print("=" * 60, flush=True)
    print("서버 초기화 시작...", flush=True)
    print("=" * 60, flush=True)
    
    try:
        print("[1/2] DB 초기화...", flush=True)
        init_db()
        _db_ready = True
        from backend.database import rebuild_hub_fts
        rebuild_hub_fts()
        print("[1/2] DB 초기화 완료", flush=True)
        
        print("[2/2] 인덱스/Vector Store 준비...", flush=True)
        if USE_OPENAI_VECTOR_STORE:
            from backend.rag_registry import get_vector_store_ids
            
            policy_id, regional_id, winners2022_id = get_vector_store_ids()
            if not policy_id and OPENAI_VECTOR_STORE_ID:
                policy_id = OPENAI_VECTOR_STORE_ID
                regional_id = OPENAI_REGIONAL_VECTOR_STORE_ID
            
            if not policy_id:
                print("[경고] Vector Store ID 없음. 일부 기능이 작동하지 않을 수 있습니다.", flush=True)
            else:
                _vector_store_id = policy_id
                _regional_vector_store_id = regional_id
                _winners2022_vector_store_id = winners2022_id or None
        else:
            print("[인덱스] 빌드 중 (시간이 걸릴 수 있습니다)...", flush=True)
            _indexes = build_all_indexes(force_rebuild=False)
            if "platform" not in _indexes:
                _indexes["platform"] = VectorIndex(dimension=EMBEDDING_DIMENSION, use_cosine=True)
            if "pledge" not in _indexes:
                _indexes["pledge"] = VectorIndex(dimension=EMBEDDING_DIMENSION, use_cosine=True)
            if "regional" not in _indexes:
                _indexes["regional"] = VectorIndex(dimension=EMBEDDING_DIMENSION, use_cosine=True)
            # FAISS 모드에서도 winners2022 벡터 스토어 ID 로드 (공약 유사도 검색용)
            from backend.config import OPENAI_WINNERS2022_VECTOR_STORE_ID
            from backend.rag_registry import get_vector_store_ids
            _w2022_env = OPENAI_WINNERS2022_VECTOR_STORE_ID.strip()
            if not _w2022_env:
                _, _, _w2022_env = get_vector_store_ids()
            if _w2022_env:
                _winners2022_vector_store_id = _w2022_env
                print(f"[winners2022] 벡터 스토어 로드: {_w2022_env}", flush=True)
        
        _startup_done = True
        print("=" * 60, flush=True)
        print("서버 초기화 완료!", flush=True)
        print("=" * 60, flush=True)
    except Exception as e:
        print(f"초기화 실패: {e}", flush=True)
        traceback.print_exc()
        raise


# startup_event 제거됨 - 서버가 즉시 시작되도록 함
# 필요한 초기화는 _ensure_startup()에서 lazy loading으로 처리

STATIC_DIR = ROOT_DIR / "static"
AUTH_COOKIE = "policy_auth"
AUTH_COOKIE_MAX_AGE = 7 * 24 * 3600


def _client_ip(request: Request) -> str:
    return request.client.host if request.client else (request.headers.get("x-forwarded-for", "").split(",")[0].strip() or "0.0.0.0")


def get_current_user(request: Request) -> Optional[dict]:
    token = request.cookies.get(AUTH_COOKIE)
    return verify_session_token(token) if token else None


def require_user(request: Request) -> dict:
    user = get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="로그인이 필요합니다.")
    return user


def require_admin(request: Request) -> dict:
    user = require_user(request)
    # role 컬럼과 ADMIN_EMAILS 둘 중 하나라도 관리자 조건이면 허용
    if user["role"] == ROLE_ADMIN or user["email"] in ADMIN_EMAILS:
        return user
    raise HTTPException(status_code=403, detail="관리자만 접근 가능합니다.")


def require_approved(request: Request) -> dict:
    user = require_user(request)
    # ADMIN_EMAILS/관리자는 항상 승인된 것으로 처리
    if user["email"] in ADMIN_EMAILS or user["role"] == ROLE_ADMIN:
        return user
    if user["status"] != STATUS_APPROVED:
        log_usage(
            user_id=user["id"],
            ip=_client_ip(request),
            endpoint=request.url.path,
            action="blocked_unapproved",
            input_chars=0,
            output_chars=0,
            model="",
            token_in=None,
            token_out=None,
            cost_estimate=None,
            status_code=403,
            latency_ms=0,
            error_message="승인되지 않은 사용자",
        )
        raise HTTPException(status_code=403, detail="승인되지 않은 사용자입니다. 관리자 승인 후 이용 가능합니다.")
    return user


class PledgeCheckRequest(BaseModel):
    pledge: str = Field(..., max_length=8000, description="점검할 출마자 공약 텍스트")


class PledgeCheckResponse(BaseModel):
    result: str = Field(..., description="부합 점검 결과 (판정, 근거, 체크리스트 등)")


def _serve_html(filename: str):
    path = STATIC_DIR / filename
    if path.exists():
        return FileResponse(
            path,
            media_type="text/html; charset=utf-8",
            headers={"Cache-Control": "no-store"},
        )
    return None


@app.api_route("/og.svg", methods=["GET", "HEAD"])
def og_image():
    """Open Graph thumbnail image (for KakaoTalk link preview)."""
    path = STATIC_DIR / "og.svg"
    if path.exists():
        return FileResponse(path, media_type="image/svg+xml; charset=utf-8")
    raise HTTPException(status_code=404, detail="og.svg not found")


@app.api_route("/og.png", methods=["GET", "HEAD"])
def og_image_png():
    """Open Graph thumbnail image (PNG for KakaoTalk link preview)."""
    path = STATIC_DIR / "og.png"
    if path.exists():
        return FileResponse(path, media_type="image/png")
    raise HTTPException(status_code=404, detail="og.png not found")


@app.api_route("/og-coach.jpg", methods=["GET", "HEAD"])
def og_image_coach():
    """OG image for 공약코치 page."""
    path = STATIC_DIR / "og-coach.jpg"
    if path.exists():
        return FileResponse(path, media_type="image/jpeg")
    raise HTTPException(status_code=404, detail="og-coach.jpg not found")


@app.get("/static/{path:path}")
def serve_static(path: str):
    """정적 파일 (JS, CSS 등) 제공."""
    safe = Path(path)
    if safe.is_absolute() or ".." in safe.parts:
        raise HTTPException(status_code=404, detail="Not found")
    file_path = (STATIC_DIR / path).resolve()
    if not file_path.is_file() or not str(file_path).startswith(str(STATIC_DIR.resolve())):
        raise HTTPException(status_code=404, detail="Not found")
    suffix = file_path.suffix.lower()
    media = "application/javascript; charset=utf-8" if suffix == ".js" else "text/css; charset=utf-8" if suffix == ".css" else None
    return FileResponse(file_path, media_type=media)


@app.api_route("/", methods=["GET", "HEAD"])
def index():
    """메인 페이지: 서비스 소개 및 공약 점검 진입."""
    res = _serve_html("index.html")
    if res is not None:
        return res
    return {"service": "개혁신당 공약 멘토링", "endpoint": "POST /check"}


def _login_redirect(path: str, query: str = ""):
    from urllib.parse import quote
    full = path + ("?" + query if query else "")
    return RedirectResponse(url=f"/login?next={quote(full)}", status_code=302)


@app.api_route("/test-check", methods=["GET", "HEAD"])
def test_check_page():
    """UI 테스트 전용 페이지. GPT API 호출 없이 샘플 데이터로 렌더링."""
    res = _serve_html("test-check.html")
    if res is not None:
        return res
    raise HTTPException(status_code=404, detail="test-check.html not found")


import re as _re
_OG_BOT_RE = _re.compile(r"kakaotalk|facebookexternalhit|facebot|twitterbot|slackbot|linkedinbot|discordbot|telegrambot|whatsapp", _re.IGNORECASE)

@app.api_route("/pledge", methods=["GET", "HEAD"])
def pledge_page(request: Request):
    """공약 입력·점검 폼 페이지. (승인 사용자 전용)"""
    # OG 크롤러는 로그인 없이 HTML(OG 태그) 접근 허용
    ua = request.headers.get("user-agent", "")
    if _OG_BOT_RE.search(ua):
        res = _serve_html("pledge.html")
        if res is not None:
            return res
    user = get_current_user(request)
    # 비로그인·미승인도 페이지 접근 허용 — JS 게이트가 탭별로 처리
    if user and user["status"] != STATUS_APPROVED and user["email"] not in ADMIN_EMAILS and user["role"] != ROLE_ADMIN:
        return RedirectResponse(url="/pending", status_code=302)
    res = _serve_html("pledge.html")
    if res is not None:
        return res
    raise HTTPException(status_code=404, detail="pledge.html not found")


@app.api_route("/tools", methods=["GET", "HEAD"])
def tools_page(request: Request):
    """후보자 AI 정책 도구 페이지. (승인 사용자 전용)"""
    user = get_current_user(request)
    if not user:
        return _login_redirect(request.url.path)
    if (
        user["status"] != STATUS_APPROVED
        and user["email"] not in ADMIN_EMAILS
        and user["role"] != ROLE_ADMIN
    ):
        return RedirectResponse(url="/pending", status_code=302)
    res = _serve_html("tools.html")
    if res is not None:
        return res
    raise HTTPException(status_code=404, detail="tools.html not found")


@app.api_route("/signup", methods=["GET", "HEAD"])
def signup_page():
    res = _serve_html("signup.html")
    if res:
        return res
    raise HTTPException(status_code=404, detail="signup.html not found")


@app.api_route("/login", methods=["GET", "HEAD"])
def login_page():
    res = _serve_html("login.html")
    if res:
        return res
    raise HTTPException(status_code=404, detail="login.html not found")


@app.api_route("/pending", methods=["GET", "HEAD"])
def pending_page():
    res = _serve_html("pending.html")
    if res:
        return res
    raise HTTPException(status_code=404, detail="pending.html not found")


@app.api_route("/dashboard", methods=["GET", "HEAD"])
def dashboard_page(request: Request):
    user = get_current_user(request)
    if not user:
        return _login_redirect(request.url.path)
    res = _serve_html("dashboard.html")
    if res:
        return res
    raise HTTPException(status_code=404, detail="dashboard.html not found")


@app.api_route("/admin", methods=["GET", "HEAD"])
def admin_page(request: Request):
    user = get_current_user(request)
    if not user:
        return _login_redirect(request.url.path)
    if user["role"] != ROLE_ADMIN and user["email"] not in ADMIN_EMAILS:
        raise HTTPException(status_code=403, detail="관리자만 접근 가능합니다.")
    res = _serve_html("admin/index.html")
    if res:
        return res
    raise HTTPException(status_code=404, detail="admin/index.html not found")


@app.api_route("/admin/users", methods=["GET", "HEAD"])
def admin_users_page(request: Request):
    user = get_current_user(request)
    if not user:
        return _login_redirect(request.url.path)
    if user["role"] != ROLE_ADMIN and user["email"] not in ADMIN_EMAILS:
        raise HTTPException(status_code=403, detail="관리자만 접근 가능합니다.")
    res = _serve_html("admin/users.html")
    if res:
        return res
    raise HTTPException(status_code=404, detail="admin/users.html not found")


@app.api_route("/admin/candidates", methods=["GET", "HEAD"])
def admin_candidates_page(request: Request):
    """관리자 전용: 출마자·공약 등록 페이지."""
    _ = require_admin(request)
    res = _serve_html("admin/candidates.html")
    if res is not None:
        return res
    raise HTTPException(status_code=404, detail="admin/candidates.html not found")


@app.api_route("/admin/usage", methods=["GET", "HEAD"])
def admin_usage_page(request: Request):
    user = get_current_user(request)
    if not user:
        return _login_redirect(request.url.path)
    if user["role"] != ROLE_ADMIN and user["email"] not in ADMIN_EMAILS:
        raise HTTPException(status_code=403, detail="관리자만 접근 가능합니다.")
    res = _serve_html("admin/usage.html")
    if res:
        return res
    raise HTTPException(status_code=404, detail="admin/usage.html not found")


@app.api_route("/admin/ops", methods=["GET", "HEAD"])
def admin_ops_page(request: Request):
    """관리자 전용: 운영 상태 (벡터스토어·PDF 디렉터리) 페이지."""
    _ = require_admin(request)
    res = _serve_html("admin/ops.html")
    if res is not None:
        return res
    raise HTTPException(status_code=404, detail="admin/ops.html not found")


class SignupBody(BaseModel):
    name: str = Field(..., description="이름")
    phone: str = Field(..., description="전화번호")
    email: str = Field(..., description="이메일")
    password: str = Field(..., description="비밀번호")
    election_position: str = Field(default="", description="출마 유형: metro_mayor|regional_council|local_mayor|local_council")
    region_code: str = Field(default="", description="행정구역 코드")
    region_name: str = Field(default="", description="행정구역명")
    district_code: str = Field(default="", description="선거구 코드")
    district_name: str = Field(default="", description="선거구명")


def _data_gokr_gusigun(sd_name: Optional[str] = None, page_no: int = 1, num_of_rows: int = 500) -> list:
    """공공데이터 getCommonGusigunCodeList. sgId=20220601(제8회 지방선거). sd_name 있으면 해당 시도만."""
    if not DATA_GO_KR_API_KEY:
        return []
    from urllib.parse import urlencode
    from urllib.request import Request, urlopen
    from urllib.error import HTTPError
    base = "https://apis.data.go.kr/9760000/CommonCodeService/getCommonGusigunCodeList"
    params = {"ServiceKey": DATA_GO_KR_API_KEY, "sgId": "20220601", "pageNo": page_no, "numOfRows": num_of_rows, "resultType": "json"}
    if sd_name:
        params["sdName"] = sd_name
    url = f"{base}?{urlencode(params)}"
    req = Request(url, headers={"Accept": "application/json", "User-Agent": "Mozilla/5.0", "Referer": "https://www.data.go.kr/"})
    try:
        with urlopen(req, timeout=15) as res:
            data = json.loads(res.read().decode("utf-8"))
    except (HTTPError, OSError, ValueError) as e:
        logger.warning("공공데이터 구시군 API 오류: %s", e)
        return []
    body = data.get("response", {}).get("body", {}) or data.get("body", {})
    items = body.get("items") or body.get("item")
    if items is None:
        return []
    if isinstance(items, dict):
        items = items.get("item")
    return items if isinstance(items, list) else [items]


def _data_gokr_sgg_list(sg_typecode: int, page_no: int = 1, num_of_rows: int = 100) -> list:
    """공공데이터 getCommonSggCodeList. sgId=20220601, sgTypecode 4=광역의원 6=기초의원.
    NOTE: 이 API는 numOfRows 최대 100개 제한."""
    if not DATA_GO_KR_API_KEY:
        return []
    from urllib.parse import urlencode
    from urllib.request import Request, urlopen
    from urllib.error import HTTPError
    base = "https://apis.data.go.kr/9760000/CommonCodeService/getCommonSggCodeList"
    params = {
        "ServiceKey": DATA_GO_KR_API_KEY,
        "sgId": "20220601",
        "sgTypecode": sg_typecode,
        "pageNo": page_no,
        "numOfRows": num_of_rows,
        "resultType": "json",
    }
    url = f"{base}?{urlencode(params)}"
    req = Request(url, headers={"Accept": "application/json", "User-Agent": "Mozilla/5.0", "Referer": "https://www.data.go.kr/"})
    try:
        with urlopen(req, timeout=15) as res:
            data = json.loads(res.read().decode("utf-8"))
    except (HTTPError, OSError, ValueError) as e:
        logger.warning("공공데이터 getCommonSggCodeList 오류: %s", e)
        return []
    body = data.get("response", {}).get("body", {}) or data.get("body", {})
    items = body.get("items") or body.get("item")
    if items is None:
        return []
    if isinstance(items, dict):
        items = items.get("item")
    return items if isinstance(items, list) else [items]


def _extract_sub_from_sgg(sgg_name: str, wiw_name: str) -> str:
    """sggName에서 구시군명 제거 후 세부선거구명만 반환 (가선거구, 제1선거구 등)."""
    sgg = (sgg_name or "").strip()
    wiw = (wiw_name or "").strip()
    if not sgg:
        return "단독"
    if not wiw or wiw == sgg:
        return "단독"
    wiw_compact = "".join(wiw.split())
    sgg_compact = "".join(sgg.split())
    if wiw_compact and sgg_compact.startswith(wiw_compact):
        sub = sgg_compact[len(wiw_compact) :].strip()
        return sub if sub else "단독"
    if wiw in sgg:
        sub = sgg.replace(wiw, "", 1).strip()
        return sub if sub else "단독"
    # 다구 도시: wiwName("성남시 분당구")과 sggName("성남시아선거구") 접두사가 다를 때
    # 시/군 단위까지만 제거하여 세부선거구 추출
    for suffix in ("시", "군"):
        idx = wiw_compact.find(suffix)
        if idx >= 0:
            city = wiw_compact[: idx + 1]
            if sgg_compact.startswith(city):
                sub = sgg_compact[len(city) :].strip()
                if sub:
                    return sub
    return sgg




@app.get("/favicon.ico", include_in_schema=False)
def favicon():
    icon_path = ROOT_DIR / "static" / "favicon.ico"
    if icon_path.exists():
        return FileResponse(icon_path)
    raise HTTPException(status_code=404, detail="favicon not found")

@app.get("/api/signup/regions")
def api_signup_regions():
    """회원가입·어드민용 시/도 목록. region_map.json 전체를 반환하여 전 시도가 항상 노출되도록 함."""
    import json as _json
    path = ROOT_DIR / "data" / "region_map.json"
    if path.exists():
        data = _json.loads(path.read_text(encoding="utf-8"))
        regions = data.get("regions", [])
        if regions:
            out = [{"region_code": str(r.get("region_code", "")).strip(), "region_name": str(r.get("region_name", "")).strip()} for r in regions if r.get("region_code")]
            if out:
                return out
    return [{"region_code": k, "region_name": v} for k, v in REGION_NAME_MAP.items()]


@app.get("/api/signup/districts")
def api_signup_districts(
    region_code: str = Query(..., description="행정구역 코드"),
    election_position: str = Query(default="", description="metro_mayor|regional_council|local_mayor|local_council"),
):
    """회원가입용 선거구(시군구) 목록. 광역 단체장이면 빈 배열. 그 외는 공공데이터 getCommonGusigunCodeList 기반."""
    if (election_position or "").strip().lower() == "metro_mayor":
        return []
    code = (region_code or "").strip()
    if not code:
        return []
    region_name = REGION_NAME_MAP.get(code, "")
    if not region_name:
        import json as _json
        path = ROOT_DIR / "data" / "region_map.json"
        if path.exists():
            for r in _json.loads(path.read_text(encoding="utf-8")).get("regions", []):
                if str(r.get("region_code", "")) == code:
                    region_name = str(r.get("region_name", ""))
                    break
    if not region_name:
        return []
    if DATA_GO_KR_API_KEY:
        items = _data_gokr_gusigun(sd_name=region_name, page_no=1, num_of_rows=1000)
        if items:
            result = []
            seen_wiw = set()
            for it in items:
                wiw = (it.get("wiwName") or it.get("WIW_NAME") or "").strip()
                wiw_norm = "".join(wiw.split()) or wiw
                if wiw_norm and wiw_norm not in seen_wiw:
                    seen_wiw.add(wiw_norm)
                    result.append({"district_code": f"{code}:{wiw_norm}", "district_name": wiw, "region_code": code})
            if result:
                return result
    path = ROOT_DIR / "data" / "district_map.json"
    if not path.exists():
        return []
    import json as _json
    try:
        data = _json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        logger.warning("district_map.json load failed: %s", e)
        return []
    raw_groups = data.get("data", [])
    region_aliases = {region_name}
    region_map_path = ROOT_DIR / "data" / "region_map.json"
    if region_map_path.exists():
        rm = _json.loads(region_map_path.read_text(encoding="utf-8"))
        for r in rm.get("regions", []):
            if str(r.get("region_code", "")) == code:
                region_aliases.add(str(r.get("region_name", "")).strip())
                region_aliases.update(str(a).strip() for a in r.get("aliases", []))
                break
    for row in raw_groups or []:
        if not isinstance(row, dict) or len(row) != 1:
            continue
        rname, district_names = next(iter(row.items()))
        if str(rname).strip() not in region_aliases:
            continue
        names = list(district_names) if district_names else [rname]
        out = []
        for d in names:
            name = str(d).strip()
            if not name:
                continue
            district_norm = "".join(c for c in name if c not in (" ", "\t"))
            out.append({"district_code": f"{code}:{district_norm}", "district_name": name, "region_code": code})
        return out


# 기초의원 세부선거구: 제N선거구 → 가나다라 (가선거구, 나선거구, ...) 변환용
_SGG_NUMBER_TO_GANADARA = (
    "가", "나", "다", "라", "마", "바", "사", "아", "자", "차", "카", "타", "파", "하",
    "거", "너", "더", "러", "머", "버", "서", "어", "저", "처", "커", "터", "퍼", "허",
)


_local_council_cache: dict[str, list] = {}


@app.get("/api/signup/district-sub")
def api_signup_district_sub(
    district_code: str = Query(..., description="시군구 코드 (예: 11:강북구)"),
    election_position: str = Query(default="", description="local_council이면 기초의원 → 가나다라 표기"),
):
    """시군구 선택 후 세부선거구 목록. 기초의원이면 공공 API getCommonSggCodeList(sgTypecode=6)로 가·나·다 등 전부 조회, 없으면 JSON 폴백."""
    key = (district_code or "").strip()
    if not key:
        return [{"sub_code": "단독", "sub_name": "단독"}]
    import re as _re
    default = [{"sub_code": "단독", "sub_name": "단독"}]
    is_local_council = (election_position or "").strip().lower() in ("local_council", "기초의원", "기초 의원")

    # 기초의원 + API 키 있으면 공공 API로 해당 구 세부선거구 전부 조회 (가~까 등)
    # 결과를 메모리 캐시하여 반복 호출 시 즉시 반환
    if is_local_council and DATA_GO_KR_API_KEY and ":" in key:
        cached = _local_council_cache.get(key)
        if cached is not None:
            return cached if cached else default
        parts = key.split(":", 1)
        region_code = (parts[0] or "").strip()
        wiw_norm = "".join((parts[1] or "").split()) or (parts[1] or "").strip()
        sd_name = REGION_NAME_MAP.get(region_code, "")
        if region_code and wiw_norm and sd_name:
            seen = set()
            page = 1
            while True:
                items = _data_gokr_sgg_list(6, page_no=page, num_of_rows=100)
                if not items:
                    break
                for it in items:
                    sd = (it.get("sdName") or it.get("SD_NAME") or "").strip()
                    wiw = (it.get("wiwName") or it.get("WIW_NAME") or "").strip()
                    wiw_c = "".join(wiw.split()) or wiw
                    if sd != sd_name and "".join(sd.split()) != "".join(sd_name.split()):
                        continue
                    if wiw_c != wiw_norm and wiw != wiw_norm:
                        # 다구(多區) 도시 지원: "성남시분당구"가 "성남시"로 시작하면 매칭
                        if not wiw_norm.startswith(wiw_c):
                            continue
                    sub = _extract_sub_from_sgg(
                        it.get("sggName") or it.get("SGG_NAME") or "",
                        wiw,
                    )
                    if sub and sub != "단독" and sub not in seen:
                        seen.add(sub)
                if len(items) < 100:
                    break
                page += 1
            if seen:
                sorted_subs = sorted(seen)
                out = []
                for sub in sorted_subs:
                    if sub.startswith("제") and _re.match(r"제\d+선거구", sub):
                        m = _re.match(r"제(\d+)선거구", sub)
                        if m:
                            idx = int(m.group(1))
                            if 1 <= idx <= len(_SGG_NUMBER_TO_GANADARA):
                                sub_name = _SGG_NUMBER_TO_GANADARA[idx - 1] + "선거구"
                                out.append({"sub_code": sub_name, "sub_name": sub_name})
                                continue
                    out.append({"sub_code": sub, "sub_name": sub})
                if out:
                    _local_council_cache[key] = out
                    return out
            _local_council_cache[key] = []

    path = ROOT_DIR / "data" / "district_sub_map.json"
    if not path.exists():
        return default
    import json as _json
    try:
        data = _json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        logger.warning("district_sub_map.json load failed: %s", e)
        return default
    subs = data.get("subs") or {}
    names = subs.get(key)
    if not names or not isinstance(names, list):
        return default
    is_regional_council = (election_position or "").strip().lower() in ("regional_council", "광역의원", "시도의원")
    all_items = [str(s).strip() for s in names if s and str(s).strip()]
    if is_local_council:
        ganadara = [s for s in all_items if not s.startswith("제") and s != "단독" and s.endswith("선거구")]
        if ganadara:
            return [{"sub_code": s, "sub_name": s} for s in sorted(ganadara)]
        return default  # 기초의원인데 가나다 선거구가 없으면 단독 반환 (광역의원 선거구 노출 방지)
    elif is_regional_council:
        jenu = [s for s in all_items if _re.match(r"제\d+선거구", s)]
        if jenu:
            return [{"sub_code": s, "sub_name": s} for s in sorted(jenu, key=lambda x: int(_re.search(r"\d+", x).group()) if _re.search(r"\d+", x) else 0)]
    filtered = [s for s in all_items if s != "단독"]
    if not filtered:
        return default
    return [{"sub_code": s, "sub_name": s} for s in filtered]


@app.post("/api/auth/signup")
def api_signup(body: SignupBody):
    ok, msg = auth_signup(
        body.email,
        body.password,
        name=body.name,
        phone=body.phone,
        election_position=body.election_position or "",
        region_code=body.region_code or "",
        region_name=body.region_name or "",
        district_code=body.district_code or "",
        district_name=body.district_name or "",
    )
    if not ok:
        raise HTTPException(status_code=400, detail=msg)
    return {"message": msg}


class LoginBody(BaseModel):
    email: str = Field(..., description="이메일")
    password: str = Field(..., description="비밀번호")
    next: str = Field(default="", description="로그인 후 이동할 경로")


@app.post("/api/auth/login")
def api_login(body: LoginBody, request: Request):
    user = auth_login(body.email, body.password)
    if user is None:
        raise HTTPException(status_code=401, detail="이메일 또는 비밀번호가 올바르지 않습니다.")
    if isinstance(user, dict) and user.get("error") == "email_not_verified":
        raise HTTPException(status_code=401, detail="이메일 인증이 필요합니다. 아래 '인증 메일 다시 받기'를 이용하세요.")
    token = create_session_token(user)
    next_path = (body.next or "").strip()
    if user["status"] != STATUS_APPROVED:
        redirect_url = "/pending"
    elif next_path and next_path.startswith("/") and next_path not in ("/login", "/signup"):
        redirect_url = next_path
    elif user["email"] in ADMIN_EMAILS or user["role"] == ROLE_ADMIN:
        redirect_url = "/admin"
    else:
        redirect_url = "/"
    from fastapi.responses import JSONResponse
    resp = JSONResponse({"redirect": redirect_url})
    resp.set_cookie(AUTH_COOKIE, token, max_age=AUTH_COOKIE_MAX_AGE, httponly=True, samesite="lax")
    return resp


class ResendVerificationBody(BaseModel):
    email: str = Field(..., description="이메일")


@app.post("/api/auth/resend-verification")
def api_resend_verification(body: ResendVerificationBody):
    ok, msg = resend_verification_email(body.email)
    if not ok:
        raise HTTPException(status_code=400, detail=msg)
    return {"message": msg}


@app.api_route("/verify-email", methods=["GET", "HEAD"])
def verify_email_page(request: Request, token: str = Query(default="", alias="token")):
    """이메일 인증 링크 처리. token 검증 후 로그인 페이지로 리다이렉트."""
    from urllib.parse import quote
    if request.method == "HEAD":
        return RedirectResponse(url="/login", status_code=302)
    ok, msg = verify_email_token(token)
    if ok:
        return RedirectResponse(url="/login?verified=1", status_code=302)
    return RedirectResponse(url=f"/login?verified=0&msg={quote(msg)}", status_code=302)


@app.post("/api/auth/logout")
def api_logout():
    resp = RedirectResponse(url="/", status_code=302)
    resp.delete_cookie(AUTH_COOKIE)
    return resp


@app.get("/api/auth/me")
def api_me(request: Request):
    user = get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="로그인 필요")
    full = get_user(user["id"])
    out = {"id": user["id"], "email": user["email"], "status": user["status"], "role": user["role"]}
    out["is_admin"] = user["role"] == ROLE_ADMIN or user["email"] in ADMIN_EMAILS
    if full:
        out["name"] = full.get("name") or ""
        out["election_position"] = full.get("election_position") or ""
        out["region_code"] = full.get("region_code") or ""
        out["region_name"] = full.get("region_name") or ""
        out["district_code"] = full.get("district_code") or ""
        out["district_name"] = full.get("district_name") or ""
    return out


@app.get("/api/admin/users/pending")
def api_admin_users_pending(request: Request):
    user = require_admin(request)
    return {"users": list_users_pending()}


@app.get("/api/admin/users")
def api_admin_users_all(request: Request):
    user = require_admin(request)
    return {"users": list_users_all()}


class ApproveBody(BaseModel):
    user_id: int = Field(..., description="사용자 ID")
    status: str = Field(..., description="APPROVED | REJECTED | SUSPENDED")
    note: str = Field(default="", description="결정 사유")


@app.post("/api/admin/users/approve")
def api_admin_approve(body: ApproveBody, request: Request):
    user = require_admin(request)
    if body.status not in ("APPROVED", "REJECTED", "SUSPENDED"):
        raise HTTPException(status_code=400, detail="status must be APPROVED, REJECTED, or SUSPENDED")
    # 메일 발송용 대상 사용자 정보 미리 조회
    target_user = get_user(body.user_id)
    ok = set_user_status(body.user_id, body.status, user["id"], body.note)
    if not ok:
        raise HTTPException(status_code=400, detail="처리 실패")
    # 승인/거절/정지 결과를 당사자에게 메일로 통보 (실패해도 처리 결과에 영향 없음)
    if target_user:
        try:
            from backend.email_sender import send_approval_status_email
            send_approval_status_email(
                to_email=target_user["email"],
                status=body.status,
                name=target_user.get("name", ""),
            )
        except Exception:
            logger.exception("승인 알림 메일 발송 중 오류 (무시)")
    return {"message": "처리 완료"}


class UpdateUserProfileBody(BaseModel):
    user_id: int = Field(..., description="사용자 ID")
    election_position: str = Field(default="", description="출마 유형")
    region_code: str = Field(default="", description="행정구역 코드")
    region_name: str = Field(default="", description="행정구역명")
    district_code: str = Field(default="", description="선거구 코드")
    district_name: str = Field(default="", description="선거구명")


@app.post("/api/admin/users/update-profile")
def api_admin_update_user_profile(body: UpdateUserProfileBody, request: Request):
    """관리자가 회원의 출마지역/선거유형을 변경."""
    _ = require_admin(request)
    _ensure_db_ready()
    from backend.database import get_connection

    conn = get_connection()
    try:
        row = conn.execute("SELECT id FROM users WHERE id = ?", (body.user_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="사용자를 찾을 수 없습니다.")

        ep = (body.election_position or "").strip()
        rc = (body.region_code or "").strip()
        rn = (body.region_name or "").strip()
        dc = (body.district_code or "").strip()
        dn = (body.district_name or "").strip()

        conn.execute(
            """
            UPDATE users
            SET election_position = ?,
                region_code = ?,
                region_name = ?,
                district_code = ?,
                district_name = ?,
                updated_at = datetime('now')
            WHERE id = ?
            """,
            (ep, rc, rn, dc, dn, body.user_id),
        )

        # candidates 테이블도 동기화 (랭킹·지도 등에서 참조)
        _ep_to_election_type = {
            "metro_mayor": "metro_mayor", "local_mayor": "local_mayor",
            "regional_council": "regional_council", "local_council": "local_council",
        }
        et = _ep_to_election_type.get(ep, ep)
        el = "metro" if ep == "metro_mayor" else "local"
        # district_code에서 세부선거구(:가선거구 등) 포함된 경우 district_name 생성
        # dc 예: "41:성남시분당구:아선거구" → district_name: "성남시분당구 아선거구"
        parts = dc.split(":") if dc else []
        cand_district_name = dn
        cand_district_code = parts[0] + ":" + parts[1] if len(parts) >= 2 else dc
        conn.execute(
            """
            UPDATE candidates
            SET district_name = ?,
                district_code = ?,
                region_code = ?,
                election_type = ?,
                election_level = ?,
                updated_at = datetime('now')
            WHERE user_id = ?
            """,
            (cand_district_name, cand_district_code, rc, et, el, body.user_id),
        )

        conn.commit()
        return {"ok": True, "message": "수정 완료"}
    except HTTPException:
        conn.rollback()
        raise
    except Exception:
        conn.rollback()
        raise HTTPException(status_code=500, detail="수정 중 오류가 발생했습니다.")
    finally:
        conn.close()


class DeleteUserBody(BaseModel):
    user_id: int = Field(..., description="사용자 ID")


class ResetQuotaBody(BaseModel):
    email: str = Field(..., description="이메일")


class ResetPasswordBody(BaseModel):
    email: str = Field(..., description="이메일")
    password: str = Field(..., description="새 비밀번호")


@app.post("/api/admin/users/reset-password")
def api_admin_reset_password(body: ResetPasswordBody, request: Request):
    """관리자: 특정 사용자의 비밀번호 재설정."""
    require_admin(request)
    from backend.database import get_connection
    from passlib.hash import pbkdf2_sha256
    conn = get_connection()
    try:
        row = conn.execute("SELECT id FROM users WHERE email = ?", (body.email.strip().lower(),)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="사용자를 찾을 수 없습니다.")
        hashed = pbkdf2_sha256.using(rounds=100000).hash(body.password)
        conn.execute("UPDATE users SET password_hash = ? WHERE id = ?", (hashed, row["id"]))
        conn.commit()
        return {"message": f"비밀번호 재설정 완료", "email": body.email}
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@app.post("/api/admin/users/reset-quota")
def api_admin_reset_quota(body: ResetQuotaBody, request: Request):
    """관리자: 특정 사용자의 오늘 쿼터(usage_logs 오늘 행) 삭제."""
    require_admin(request)
    from backend.database import get_connection
    import time
    conn = get_connection()
    try:
        row = conn.execute("SELECT id FROM users WHERE email = ?", (body.email.strip().lower(),)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="사용자를 찾을 수 없습니다.")
        today = time.strftime("%Y-%m-%d")
        deleted = conn.execute(
            "DELETE FROM usage_logs WHERE user_id = ? AND date(created_at) = ?",
            (row["id"], today),
        ).rowcount
        conn.commit()
        return {"message": f"오늘 사용 기록 {deleted}건 삭제 완료", "email": body.email}
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@app.post("/api/admin/users/delete")
def api_admin_delete_user(body: DeleteUserBody, request: Request):
    user = require_admin(request)
    if body.user_id == user["id"]:
        raise HTTPException(status_code=400, detail="자기 자신은 삭제할 수 없습니다.")

    from backend.database import get_connection
    conn = get_connection()
    try:
        # 사용자 존재 확인
        cur = conn.execute("SELECT id, role FROM users WHERE id = ?", (body.user_id,))
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="사용자를 찾을 수 없습니다.")
        if row["role"] == ROLE_ADMIN:
            raise HTTPException(status_code=400, detail="관리자 계정은 삭제할 수 없습니다.")

        # 연관 데이터 먼저 삭제 (FK cascade 없음)
        conn.execute("DELETE FROM approval_requests WHERE user_id = ? OR decided_by = ?", (body.user_id, body.user_id))
        conn.execute("DELETE FROM usage_logs WHERE user_id = ?", (body.user_id,))
        conn.execute("DELETE FROM analysis_cache WHERE user_id = ?", (body.user_id,))
        conn.execute("DELETE FROM users WHERE id = ?", (body.user_id,))
        conn.commit()
        return {"message": "삭제 완료"}
    except HTTPException:
        conn.rollback()
        raise
    except Exception:
        conn.rollback()
        raise HTTPException(status_code=500, detail="삭제 중 오류가 발생했습니다.")
    finally:
        conn.close()


@app.post("/api/admin/applicants/upload")
async def api_admin_applicants_upload(request: Request, file: UploadFile = File(...)):
    """관리자가 지원서 엑셀을 업로드하면 party_applicants 테이블에 저장 후 기존 사용자 재검증."""
    _ = require_admin(request)
    _ensure_db_ready()

    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="엑셀 파일(.xlsx)만 업로드 가능합니다.")

    try:
        import io
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="서버에 openpyxl이 설치되지 않았습니다. pip install openpyxl")

    from backend.database import get_connection
    from backend.applicant_verify import reverify_all_users

    content = await file.read()
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="지원서 엑셀은 .xlsx 파일만 업로드 가능합니다.")

    applicants, skipped_rows = _extract_applicants_from_workbook(content, file.filename)
    conn = get_connection()
    try:
        conn.execute("DELETE FROM party_applicants")
        inserted = 0
        for applicant in applicants:
            conn.execute(
                """INSERT INTO party_applicants (name, phone, email, region_province, district_info, election_position, doc_submitted, interview_done, status_note)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    applicant["name"],
                    applicant["phone"],
                    applicant["email"],
                    applicant["region_province"],
                    applicant["district_info"],
                    applicant["election_position"],
                    applicant["doc_submitted"],
                    applicant["interview_done"],
                    applicant["status_note"],
                ),
            )
            inserted += 1
        conn.commit()
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        logger.exception("Applicant upload DB save failed")
        raise HTTPException(status_code=500, detail=f"DB 저장 실패: {str(e)[:200]}")
    finally:
        conn.close()

    try:
        reverified = reverify_all_users()
    except Exception:
        logger.exception("Applicant reverify failed after upload")
        reverified = 0
    skipped_suffix = f", 빈 이름 행 {skipped_rows}건 건너뜀" if skipped_rows else ""
    return {"message": f"지원서 {inserted}건 저장, 기존 사용자 {reverified}명 재검증 완료{skipped_suffix}"}
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"엑셀 파일을 열 수 없습니다: {str(e)[:100]}")

    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # 헤더 제외
    wb.close()

    conn = get_connection()
    try:
        conn.execute("DELETE FROM party_applicants")  # 전체 교체
        inserted = 0
        for row in rows:
            if not row or len(row) < 9:
                continue
            name = str(row[1] or "").strip()
            if not name:
                continue
            phone = str(row[4] or "").strip()
            email = str(row[5] or "").strip()
            region_province = str(row[6] or "").strip()
            district_info = str(row[7] or "").strip()
            election_position = str(row[8] or "").strip()
            doc_submitted = 1 if (len(row) > 9 and str(row[9] or "").strip() == "●") else 0
            interview_done = 1 if (len(row) > 10 and str(row[10] or "").strip() == "●") else 0
            status_note = str(row[11] or "").strip() if len(row) > 11 else ""
            conn.execute(
                """INSERT INTO party_applicants (name, phone, email, region_province, district_info, election_position, doc_submitted, interview_done, status_note)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (name, phone, email, region_province, district_info, election_position, doc_submitted, interview_done, status_note),
            )
            inserted += 1
        conn.commit()
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        logger.exception("엑셀 업로드 DB 저장 실패")
        raise HTTPException(status_code=500, detail=f"DB 저장 실패: {str(e)[:200]}")
    finally:
        conn.close()

    # 기존 가입자 전원 재검증
    try:
        reverified = reverify_all_users()
    except Exception:
        logger.exception("재검증 실패")
        reverified = 0
    return {"message": f"지원서 {inserted}건 저장, 기존 사용자 {reverified}명 재검증 완료"}


@app.post("/api/admin/applicants/reverify")
def api_admin_reverify(request: Request):
    """관리자가 수동으로 전원 재검증 실행."""
    _ = require_admin(request)
    _ensure_db_ready()
    from backend.applicant_verify import reverify_all_users
    n = reverify_all_users()
    return {"message": f"{n}명 재검증 완료"}


@app.get("/api/admin/applicants")
def api_admin_applicants(request: Request):
    """업로드된 지원서 목록 조회."""
    _ = require_admin(request)
    _ensure_db_ready()
    from backend.database import get_connection
    conn = get_connection()
    try:
        rows = conn.execute(
            "SELECT id, name, phone, email, region_province, district_info, election_position, doc_submitted, interview_done, status_note FROM party_applicants ORDER BY id"
        ).fetchall()
        return {"applicants": [dict(r) for r in rows], "total": len(rows)}
    finally:
        conn.close()


@app.get("/api/usage/summary")
def api_usage_summary(request: Request):
    u = require_user(request)
    from backend.database import get_connection
    from backend.quota_rate import is_unlimited_quota_user
    import time
    today = time.strftime("%Y-%m-%d")
    month = time.strftime("%Y-%m")
    conn = get_connection()
    try:
        cur = conn.execute(
            """
            SELECT
                COALESCE((SELECT SUM(COALESCE(token_in,0)+COALESCE(token_out,0)) FROM usage_logs WHERE user_id = ? AND date(created_at) = ? AND status_code >= 200 AND status_code < 300 AND endpoint != '/api/pledge/verify'), 0) AS daily_used,
                COALESCE((SELECT SUM(COALESCE(token_in,0)+COALESCE(token_out,0)) FROM usage_logs WHERE user_id = ? AND strftime('%Y-%m', created_at) = ? AND status_code >= 200 AND status_code < 300 AND endpoint != '/api/pledge/verify'), 0) AS monthly_used
            """,
            (u["id"], today, u["id"], month),
        )
        row = cur.fetchone()
        from backend.config import QUOTA_DAILY_TOKENS, QUOTA_MONTHLY_TOKENS
        daily_used = row["daily_used"] if row else 0
        monthly_used = row["monthly_used"] if row else 0
        if is_unlimited_quota_user(u):
            return {
                "daily_used": daily_used,
                "monthly_used": monthly_used,
                "daily_limit": None,
                "monthly_limit": None,
                "daily_remaining": None,
                "monthly_remaining": None,
                "unlimited": True,
            }
        return {
            "daily_used": daily_used,
            "monthly_used": monthly_used,
            "daily_limit": QUOTA_DAILY_TOKENS,
            "monthly_limit": QUOTA_MONTHLY_TOKENS,
            "daily_remaining": max(0, QUOTA_DAILY_TOKENS - daily_used),
            "monthly_remaining": max(0, QUOTA_MONTHLY_TOKENS - monthly_used),
            "unlimited": False,
        }
    finally:
        conn.close()


@app.get("/api/admin/usage/stats")
def api_admin_usage_stats(request: Request):
    user = require_admin(request)
    period = request.query_params.get("period", "7")
    days = min(90, max(1, int(period) if period.isdigit() else 7))
    from backend.database import get_connection
    conn = get_connection()
    try:
        cur = conn.execute(
            """
            SELECT user_id, COUNT(*) as cnt, SUM(COALESCE(cost_estimate, 0)) as cost
            FROM usage_logs
            WHERE datetime(created_at) >= datetime('now', ?)
            AND status_code >= 200 AND status_code < 300 AND action = 'analysis_run'
            GROUP BY user_id
            ORDER BY cnt DESC
            """,
            (f"-{days} days",),
        )
        rows = cur.fetchall()
        users = {r["user_id"]: r for r in rows}
        user_info = {}
        for uid in users:
            u = get_user(uid)
            user_info[uid] = {
                "email": u["email"] if u else str(uid),
                "name": (u.get("name") or "").strip() if u else "",
                "election_position": (u.get("election_position") or "").strip() if u else "",
                "region_name": (u.get("region_name") or "").strip() if u else "",
                "district_name": (u.get("district_name") or "").strip() if u else "",
            }
        return {
            "period_days": days,
            "by_user": [
                {
                    "user_id": r["user_id"],
                    "email": user_info.get(r["user_id"], {}).get("email", str(r["user_id"])),
                    "name": user_info.get(r["user_id"], {}).get("name", ""),
                    "election_position": user_info.get(r["user_id"], {}).get("election_position", ""),
                    "region_name": user_info.get(r["user_id"], {}).get("region_name", ""),
                    "district_name": user_info.get(r["user_id"], {}).get("district_name", ""),
                    "count": r["cnt"],
                    "cost_estimate": r["cost"] or 0,
                }
                for r in rows
            ],
        }
    finally:
        conn.close()


@app.get("/api/admin/ops/status")
def api_admin_ops_status(request: Request):
    """관리자 전용: 벡터스토어 ID·PDF 디렉터리 요약 등 운영 상태 (읽기 전용)."""
    require_admin(request)
    _ensure_startup()
    global _vector_store_id, _regional_vector_store_id, _winners2022_vector_store_id
    try:
        from backend.rag_registry import get_vector_store_ids
        policy_id, regional_id, winners2022_id = get_vector_store_ids()
    except Exception as e:
        logger.warning("rag_registry get_vector_store_ids failed: %s", e)
        policy_id = regional_id = winners2022_id = ""
    out = {
        "vector_stores": {
            "policy": _vector_store_id or policy_id or "",
            "regional": _regional_vector_store_id or regional_id or "",
            "winners2022": _winners2022_vector_store_id or winners2022_id or "",
        },
        "use_openai_vector_store": USE_OPENAI_VECTOR_STORE,
    }
    try:
        out["fs"] = _get_fs_debug()
    except Exception as e:
        out["fs_error"] = str(e)
    return out


@app.get("/map")
def map_page():
    """지역별 출마자 공약 지도 페이지."""
    res = _serve_html("map.html")
    if res is not None:
        return res
    raise HTTPException(status_code=404, detail="map.html not found")


@app.get("/api")
def api_info():
    return {"service": "개혁신당 공약 멘토링", "endpoint": "POST /check"}


@app.get("/test")
def test():
    """간단한 테스트 엔드포인트."""
    return {"status": "ok", "message": "서버 작동 중", "version": "1.3.0"}




def _get_fs_debug() -> dict:
    """PDF 디렉터리·폴더별 파일 수·샘플 파일명 (GET /api/debug/fs용)."""
    pdf_dir = Path(PDF_DIR).resolve()
    folders = [
        ("정강정책", pdf_dir / _nfc("정강정책")),
        ("공약", pdf_dir / _nfc("공약")),
        ("지역별 공약", pdf_dir / _nfc("지역별 공약")),
    ]
    by_folder = {}
    for name, dir_path in folders:
        exists = dir_path.exists()
        try:
            pdf_list = list(_iter_doc_files(dir_path)) if exists else []
        except Exception:
            pdf_list = []
        by_folder[name] = {
            "exists": exists,
            "doc_count": len(pdf_list),
            "sample_names": [p.name for p in sorted(pdf_list, key=lambda x: x.name)[:10]],
        }
    return {
        "pdf_dir": str(pdf_dir),
        "pdf_dir_exists": pdf_dir.exists(),
        "folders": by_folder,
    }


def _debug_endpoint(allowed: bool = True):
    """DEBUG_ENDPOINTS_ENABLED=0 시 404 반환."""
    if not allowed or not DEBUG_ENDPOINTS_ENABLED:
        raise HTTPException(status_code=404, detail="Debug endpoint disabled (DEBUG_ENDPOINTS_ENABLED=0)")


@app.get("/api/debug/admin-check")
def debug_admin_check(request: Request):
    """
    ADMIN_EMAILS 로드 여부·현재 로그인 사용자 포함 여부 확인.
    승인 우회가 안 될 때 점검용. 이메일 자체는 반환하지 않음.
    """
    _debug_endpoint()
    user = get_current_user(request)
    in_admin = user is not None and user.get("email") in ADMIN_EMAILS
    return {
        "admin_emails_count": len(ADMIN_EMAILS),
        "logged_in": user is not None,
        "user_email": user.get("email") if user else None,
        "user_in_admin_list": in_admin,
        "user_status": user.get("status") if user else None,
        "user_role": user.get("role") if user else None,
    }


@app.get("/api/debug/fs")
def debug_fs():
    """PDF 디렉터리 존재·폴더별 PDF 개수·샘플 파일명. AWS 배포 확인용."""
    _debug_endpoint()
    return _get_fs_debug()


@app.get("/api/debug/vectorstore")
def debug_vectorstore():
    """
    persist_path, collection_name, total_count, embedding_model, embedding_dim, sample_doc 반환.
    AWS 배포 시 벡터스토어 상태 확인용.
    """
    _debug_endpoint()
    global _indexes, _vector_store_id, _regional_vector_store_id, _winners2022_vector_store_id
    if USE_OPENAI_VECTOR_STORE:
        return {
            "mode": "openai_vector_store",
            "vector_store_id": _vector_store_id,
            "regional_vector_store_id": _regional_vector_store_id,
            "winners2022_vector_store_id": _winners2022_vector_store_id,
            "persist_path": "N/A (OpenAI 호스팅)",
            "collection_names": ["policy-rag-store"],
            "total_count": "N/A",
            "embedding_model_name": "OpenAI file_search",
            "embedding_dim": "N/A",
            "sample_doc": None,
        }
    if _indexes is None:
        raise HTTPException(status_code=503, detail="인덱스가 아직 초기화되지 않았습니다.")
    from backend.config import EMBEDDING_MODEL, EMBEDDING_DIMENSION
    cache_dir = Path(INDEX_CACHE_DIR).resolve()
    collections = ["platform", "pledge", "regional"]
    total_count = sum((_indexes.get(k).size() if _indexes.get(k) else 0) for k in collections)
    sample = None
    for name in collections:
        idx = _indexes.get(name)
        if idx and idx.chunks:
            c = idx.chunks[0]
            sample = {
                "collection": name,
                "doc_id": c.doc_id,
                "path": c.path,
                "text_length": len(c.text),
                "snippet": (c.text[:150] + "...") if len(c.text) > 150 else c.text,
            }
            break
    return {
        "persist_path": str(cache_dir),
        "collection_names": collections,
        "total_count": total_count,
        "embedding_model_name": EMBEDDING_MODEL,
        "embedding_dim": EMBEDDING_DIMENSION,
        "sample_doc": sample,
    }


@app.get("/api/debug/models")
def debug_models():
    """현재 서버에서 사용 중인 OpenAI 모델명을 반환 (AWS 등 배포 환경 확인용)."""
    _debug_endpoint()
    return {
        "openai_model": OPENAI_MODEL,
        "chat_model": CHAT_MODEL,
        "hint": "/check 는 OPENAI_MODEL, /api/pledge/verify·카드는 CHAT_MODEL 사용. 동일하게 쓰려면 .env에 둘 다 설정.",
    }


@app.get("/api/debug/context-summary")
def debug_context_summary():
    """
    폴더별 PDF 파일 수·추출 성공 수·총 문자 수. 로컬 vs AWS 비교용.
    수치가 AWS에서 현저히 작으면 PDF 추출이 다르게 되고 있는 것이므로 출력 차이 원인일 수 있음.
    """
    _debug_endpoint()
    from backend.config import PDF_EXTRACTOR
    try:
        summary = get_context_summary()
        return {
            "pdf_extractor": PDF_EXTRACTOR,
            "context": summary,
            "hint": "로컬과 AWS에서 이 수치를 비교하세요. total_chars 차이가 크면 추출이 다릅니다.",
        }
    except Exception as e:
        logger.exception("context-summary 실패")
        return {"error": str(e), "context": {}}


@app.get("/api/debug/index")
def debug_index():
    """인덱스 벡터 수를 반환하는 디버깅 엔드포인트."""
    _debug_endpoint()
    global _indexes, _vector_store_id, _regional_vector_store_id
    if USE_OPENAI_VECTOR_STORE:
        return {
            "mode": "openai_vector_store",
            "vector_store_id": _vector_store_id,
            "regional_vector_store_id": _regional_vector_store_id,
            "platform_vectors": 0,
            "pledge_vectors": 0,
            "regional_vectors": 0,
        }
    if _indexes is None:
        raise HTTPException(status_code=503, detail="인덱스가 아직 초기화되지 않았습니다.")

    try:
        platform_vectors = _indexes.get("platform").size() if _indexes.get("platform") else 0
        pledge_vectors = _indexes.get("pledge").size() if _indexes.get("pledge") else 0
        regional_vectors = _indexes.get("regional").size() if _indexes.get("regional") else 0
    except Exception as e:
        logger.error(f"인덱스 디버그 실패: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="인덱스 정보를 가져오는 중 오류 발생")

    return {
        "platform_vectors": platform_vectors,
        "pledge_vectors": pledge_vectors,
        "regional_vectors": regional_vectors,
    }


def _run_debug_search(source: Literal["platform", "pledge", "regional"], q: str, top_k: int):
    """source/q/top_k로 인덱스 검색 후 [{ path, chunk_id, score, snippet }] 반환."""
    _debug_endpoint()
    global _indexes
    if _indexes is None or not _indexes:
        raise HTTPException(status_code=503, detail="인덱스가 아직 초기화되지 않았습니다.")

    index = _indexes.get(source)
    if index is None:
        raise HTTPException(status_code=500, detail=f"{source} 인덱스가 없습니다.")

    from backend.embeddings import embed_texts
    from backend.report import exact_match_search, _merge_exact_and_embedding

    embeddings = embed_texts([q], batch_size=1)
    if not embeddings:
        raise HTTPException(status_code=500, detail="쿼리 임베딩 생성 실패")

    query_embedding = embeddings[0]
    exact_hits = exact_match_search(q, index, top_k_exact=min(5, top_k))
    emb_hits = index.search(query_embedding, k=top_k)
    merged = _merge_exact_and_embedding(exact_hits, emb_hits, top_k)

    return [
        {
            "path": chunk.path,
            "chunk_id": chunk.chunk_id,
            "score": round(score, 6),
            "snippet": (chunk.text[:200] + "...") if len(chunk.text) > 200 else chunk.text,
        }
        for chunk, score in merged
    ]


@app.get("/api/debug/search")
def debug_search_get(
    source: Literal["platform", "pledge", "regional"] = Query(..., description="platform | pledge | regional"),
    q: str = Query(..., min_length=1, description="검색 쿼리"),
    top_k: int = Query(10, ge=1, le=50, description="상위 결과 개수"),
):
    """
    특정 인덱스에서 검색 결과를 확인하는 디버깅 엔드포인트 (GET).
    응답: [{ "path": "...", "chunk_id": 0, "score": 0.123, "snippet": "..." }]
    """
    return _run_debug_search(source, q, top_k)


class DebugSearchBody(BaseModel):
    source: Literal["platform", "pledge", "regional"] = Field(..., description="platform | pledge | regional")
    q: str = Field(..., min_length=1, description="검색 쿼리")
    top_k: int = Field(10, ge=1, le=50, description="상위 결과 개수")


@app.post("/api/debug/search")
def debug_search_post(body: DebugSearchBody):
    """
    특정 인덱스에서 검색 (POST, JSON 바디).
    응답: [{ "path": "...", "chunk_id": 0, "score": 0.123, "snippet": "..." }]
    """
    return _run_debug_search(body.source, body.q, body.top_k)


@app.get("/api/debug/scan")
def debug_scan():
    """PDF 폴더 구조 및 파일 목록을 반환하는 디버깅 엔드포인트."""
    _debug_endpoint()
    base_dir = PDF_DIR

    def list_files(subdir_name: str):
        subdir = base_dir / subdir_name
        if not subdir.exists():
            return []
        return [str(p.relative_to(base_dir)) for p in _iter_doc_files(subdir)]

    return {
        "platform_files": list_files("정강정책"),
        "pledge_files": list_files("공약"),
        "regional_files": list_files("지역별 공약"),
    }


@app.get("/api/debug/test-public-apis")
def debug_test_public_apis():
    """공공데이터 API 키 유효성 테스트 (SEMAS/TAAS/KOSIS/Seoul)."""
    _debug_endpoint()
    from backend.public_data_api import test_all_apis
    return test_all_apis()


@app.get("/api/debug/test-public-data")
def debug_test_public_data(region: str = Query("서울특별시 강북구"), topic: str = Query("생활환경")):
    """공공데이터 통합 조회 테스트."""
    _debug_endpoint()
    from backend.public_data_api import query_public_data_context
    return query_public_data_context(region=region, topic=topic)


@app.post("/check", response_model=PledgeCheckResponse)
def check_pledge(body: PledgeCheckRequest, request: Request):
    """공약을 입력하면 중앙당의 정강정책·공약과의 적합도, 근거, 수정·보완 체크리스트를 반환한다. (승인 사용자 전용)"""
    import time
    t0 = time.perf_counter()
    logger.info("[check] started")
    try:
        _ensure_startup()  # 지연 초기화

        user = require_approved(request)
        ip = _client_ip(request)
        ok, msg = check_rate_limit_ip(ip)
        if not ok:
            raise HTTPException(status_code=429, detail=msg)
        ok, msg = check_rate_limit_user(user["id"])
        if not ok:
            raise HTTPException(status_code=429, detail=msg)

        from backend.analysis_service import run_check_analysis
        global _indexes, _vector_store_id, _regional_vector_store_id, _winners2022_vector_store_id
        vs_id = _vector_store_id if USE_OPENAI_VECTOR_STORE else None
        regional_vs_id = _regional_vector_store_id if USE_OPENAI_VECTOR_STORE else None
        winners2022_vs_id = _winners2022_vector_store_id  # VS 모드·FAISS 모드 공통 사용
        result, status_code, from_cache = run_check_analysis(
            user["id"],
            body.pledge or "",
            ip,
            vs_id,
            regional_vs_id,
            winners2022_vs_id,
            _indexes if not USE_OPENAI_VECTOR_STORE else None,
        )
        if status_code >= 400:
            raise HTTPException(status_code=status_code, detail=result)
        try:
            from backend.history import add_history

            add_history(
                user_id=user["id"],
                kind="check",
                input_text=body.pledge or "",
                result=result,
                status_code=status_code,
                from_cache=from_cache,
                options={"source": "check"},
            )
        except Exception:
            pass
        elapsed = time.perf_counter() - t0
        logger.info("[check] completed in %.1fs", elapsed)
        return PledgeCheckResponse(result=result)
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("check_pledge 오류 (after %.1fs)", time.perf_counter() - t0)
        raise HTTPException(status_code=500, detail="점검 처리 중 오류가 발생했습니다. 잠시 후 다시 시도해 주세요.")


@app.post("/check/stream")
def check_pledge_stream(body: PledgeCheckRequest, request: Request):
    """공약 당 부합 점검 - SSE 스트리밍 버전. GPT 토큰을 실시간으로 클라이언트에 전달."""
    import json as _json
    import time as _time

    _ensure_startup()

    user = require_approved(request)
    ip = _client_ip(request)
    ok, msg = check_rate_limit_ip(ip)
    if not ok:
        raise HTTPException(status_code=429, detail=msg)
    ok, msg = check_rate_limit_user(user["id"])
    if not ok:
        raise HTTPException(status_code=429, detail=msg)

    from backend.quota_rate import check_quota
    ok, msg = check_quota(user["id"])
    if not ok:
        raise HTTPException(status_code=429, detail=msg)

    pledge_text = (body.pledge or "").strip()
    if not pledge_text:
        raise HTTPException(status_code=400, detail="공약 내용이 비어 있습니다.")

    global _indexes, _vector_store_id, _regional_vector_store_id, _winners2022_vector_store_id
    vs_id = _vector_store_id if USE_OPENAI_VECTOR_STORE else None
    regional_vs_id = _regional_vector_store_id if USE_OPENAI_VECTOR_STORE else None
    winners2022_vs_id = _winners2022_vector_store_id

    t0 = _time.perf_counter()

    def generate():
        from backend.check_service import check_pledge_alignment_stream
        from backend.usage_logger import log_usage, _estimate_cost, parse_usage_marker

        accumulated: list[str] = []
        final_text = ""
        from_cache = False
        had_error = False

        try:
            gen = check_pledge_alignment_stream(
                pledge_text,
                vs_id,
                regional_vs_id,
                winners2022_vs_id,
                _indexes if not USE_OPENAI_VECTOR_STORE else None,
                user["id"],
            )
            actual_in = 0
            actual_out = 0
            for item in gen:
                if item == "[CACHED]":
                    from_cache = True
                    yield f"data: {_json.dumps({'type': 'cached'}, ensure_ascii=False)}\n\n"
                elif item.startswith("[USAGE]"):
                    usage = parse_usage_marker(item)
                    if usage:
                        actual_in = usage["token_in"]
                        actual_out = usage["token_out"]
                    continue
                elif item.startswith("[FINAL]"):
                    final_text = item[len("[FINAL]"):]
                    yield f"data: {_json.dumps({'type': 'final', 'text': final_text}, ensure_ascii=False)}\n\n"
                elif item.startswith("[ERROR]"):
                    had_error = True
                    yield f"data: {_json.dumps({'type': 'error', 'detail': item[7:]}, ensure_ascii=False)}\n\n"
                else:
                    accumulated.append(item)
                    yield f"data: {_json.dumps({'type': 'chunk', 'text': item}, ensure_ascii=False)}\n\n"

            yield f"data: {_json.dumps({'type': 'done'}, ensure_ascii=False)}\n\n"

        except Exception as e:
            had_error = True
            logger.exception("[check/stream] 오류")
            yield f"data: {_json.dumps({'type': 'error', 'detail': str(e)[:300]}, ensure_ascii=False)}\n\n"

        # 사용량 로깅
        elapsed_ms = int((_time.perf_counter() - t0) * 1000)
        status = 500 if had_error else 200
        out_chars = len(final_text) if final_text else len("".join(accumulated))
        token_in = actual_in if actual_in else len(pledge_text) // 2
        token_out = actual_out if actual_out else out_chars // 2
        cost = _estimate_cost(token_in, token_out, OPENAI_MODEL) if not had_error else None
        log_usage(
            user_id=user["id"],
            ip=ip,
            endpoint="/check/stream",
            action="cache_hit" if from_cache else "analysis_run",
            input_chars=len(pledge_text),
            output_chars=out_chars,
            model=OPENAI_MODEL,
            token_in=0 if from_cache else token_in,
            token_out=0 if from_cache else token_out,
            cost_estimate=0.0 if from_cache else cost,
            status_code=status,
            latency_ms=elapsed_ms,
        )
        # history 저장
        if not had_error and final_text:
            try:
                from backend.history import add_history
                add_history(
                    user_id=user["id"],
                    kind="check",
                    input_text=pledge_text,
                    result=final_text,
                    status_code=200,
                    from_cache=from_cache,
                    options={"source": "check/stream"},
                )
            except Exception:
                pass

    return StreamingResponse(
        generate(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


REGION_NAME_MAP = {
    "11": "서울특별시",
    "26": "부산광역시",
    "27": "대구광역시",
    "28": "인천광역시",
    "29": "광주광역시",
    "30": "대전광역시",
    "31": "울산광역시",
    "36": "세종특별자치시",
    "41": "경기도",
    "42": "강원특별자치도",
    "43": "충청북도",
    "44": "충청남도",
    "45": "전북특별자치도",
    "46": "전라남도",
    "47": "경상북도",
    "48": "경상남도",
    "50": "제주특별자치도",
}


class RegionResponse(BaseModel):
    region_code: str = Field(..., description="행정구역 코드")
    region_name: str = Field(..., description="행정구역 이름")
    candidate_count: int = Field(..., description="등록된 후보 수")


class CandidatePledgeResponse(BaseModel):
    id: Optional[int] = Field(default=None, description="공약 ID")
    title: str = Field(..., description="공약 제목")
    content: Optional[str] = Field(default=None, description="공약 세부내용")
    total_score: Optional[float] = Field(default=None, description="점검 종합점수")
    created_at: Optional[str] = Field(default=None, description="공약 등록일")


class CandidateExternalProfileResponse(BaseModel):
    source: str = Field(..., description="외부 프로필 소스 키")
    external_id: Optional[str] = Field(default=None, description="외부 서비스 후보 ID")
    profile_url: Optional[str] = Field(default=None, description="외부 후보 프로필 링크")
    photo_url: Optional[str] = Field(default=None, description="외부 후보 사진 링크")
    support_url: Optional[str] = Field(default=None, description="외부 후원 링크")
    bio: Optional[str] = Field(default=None, description="외부 프로필 소개")


class CandidateListItemResponse(BaseModel):
    candidate_id: int = Field(..., description="후보 ID")
    name: str = Field(..., description="후보명")
    district_name: Optional[str] = Field(default=None, description="선거구명")
    district_code: Optional[str] = Field(default=None, description="선거구 코드")
    region_code: str = Field(..., description="행정구역 코드")
    election_type: str = Field(..., description="선거 구분")
    election_level: Optional[str] = Field(default=None, description="선거 레벨(광역/기초 등)")
    pledges: list[CandidatePledgeResponse] = Field(default_factory=list, description="핵심 공약(최대 3개)")


class CandidateDetailResponse(BaseModel):
    candidate_id: int = Field(..., description="후보 ID")
    name: str = Field(..., description="후보명")
    district_name: Optional[str] = Field(default=None, description="선거구명")
    district_code: Optional[str] = Field(default=None, description="선거구 코드")
    region_code: str = Field(..., description="행정구역 코드")
    region_name: str = Field(..., description="행정구역 이름")
    election_type: str = Field(..., description="선거 구분")
    election_level: Optional[str] = Field(default=None, description="선거 레벨(광역/기초 등)")
    external_profile: Optional[CandidateExternalProfileResponse] = Field(default=None, description="외부 후보 프로필")
    pledges: list[CandidatePledgeResponse] = Field(default_factory=list, description="공약 전체")


class PledgeShareSummaryRequest(BaseModel):
    candidate_name: str = Field(..., description="후보명")
    election_label: Optional[str] = Field(default=None, description="출마 직위 라벨")
    region: Optional[str] = Field(default=None, description="지역/선거구")
    pledge_title: str = Field(..., description="공약 제목")
    pledge_content: str = Field(..., description="공약 본문")


class PledgeShareSummaryResponse(BaseModel):
    title: str = Field(..., description="카드뉴스용 정제 제목")
    headline: str = Field(..., description="핵심 한 줄")
    bullets: list[str] = Field(default_factory=list, description="실행 포인트")


class DistrictResponse(BaseModel):
    district_code: str = Field(..., description="선거구 코드")
    district_name: str = Field(..., description="선거구명")
    region_code: str = Field(..., description="행정구역 코드")
    candidate_count: int = Field(..., description="등록된 후보 수")


def _validate_region_code(region_code: Optional[str]) -> str:
    code = (region_code or "").strip()
    if not code:
        raise HTTPException(status_code=400, detail="region_code는 필수입니다.")
    if code in REGION_NAME_MAP:
        return code

    from backend.database import get_connection

    conn = get_connection()
    try:
        row = conn.execute("SELECT 1 FROM region_codes WHERE region_code = ? LIMIT 1", (code,)).fetchone()
    except Exception:
        row = None
    finally:
        conn.close()
    if row is None:
        raise HTTPException(status_code=400, detail=f"유효하지 않은 region_code: {code}")
    return code


def _fetch_candidate_pledges(
    candidate_id: int,
    limit: Optional[int] = None,
    public_only: bool = False,
) -> list[CandidatePledgeResponse]:
    from backend.database import get_connection

    conn = get_connection()
    try:
        sql = """
            SELECT id, title, content, total_score, created_at
            FROM candidate_pledges
            WHERE candidate_id = ?
        """
        if public_only:
            sql += """
                AND (
                    COALESCE(approval_status, 'PENDING') = 'APPROVED'
                    OR (
                        COALESCE(approval_status, 'PENDING') = 'PENDING'
                        AND (
                            NOT EXISTS (
                                SELECT 1
                                FROM candidate_pledge_review_history h0
                                WHERE h0.pledge_id = candidate_pledges.id
                            )
                            OR EXISTS (
                                SELECT 1
                                FROM candidate_pledge_review_history h
                                WHERE h.pledge_id = candidate_pledges.id
                                  AND h.approval_status = 'APPROVED'
                            )
                        )
                    )
                )
            """
        sql += """
            ORDER BY priority ASC, datetime(created_at) DESC, id DESC
        """
        params: tuple = (candidate_id,)
        if limit is not None:
            sql += " LIMIT ?"
            params = (candidate_id, limit)
        rows = conn.execute(sql, params).fetchall()
        return [
            CandidatePledgeResponse(
                id=r["id"],
                title=r["title"],
                content=r["content"],
                total_score=r["total_score"],
                created_at=r["created_at"],
            )
            for r in rows
        ]
    finally:
        conn.close()


def _fetch_candidate_external_profile(candidate_id: int) -> Optional[CandidateExternalProfileResponse]:
    from backend.database import get_connection

    conn = get_connection()
    try:
        row = conn.execute(
            """
            SELECT source_key, external_id, external_profile_url, external_photo_url,
                   external_support_url, external_bio
            FROM candidate_external_profiles
            WHERE candidate_id = ?
            ORDER BY datetime(last_synced_at) DESC, id DESC
            LIMIT 1
            """,
            (candidate_id,),
        ).fetchone()
    finally:
        conn.close()

    if row is None:
        return None

    return CandidateExternalProfileResponse(
        source=row["source_key"],
        external_id=row["external_id"],
        profile_url=row["external_profile_url"],
        photo_url=row["external_photo_url"],
        support_url=row["external_support_url"],
        bio=row["external_bio"],
    )


def _try_sync_candidate_external_profile(candidate_name: str) -> None:
    """후보 저장 직후 givemoney 프로필을 가볍게 재동기화한다."""
    safe_name = (candidate_name or "").strip()
    if not safe_name:
        return
    try:
        from scripts.sync_givemoney_profiles import sync_profiles

        result = sync_profiles(
            dry_run=False,
            search_term=safe_name,
            candidate_name=safe_name,
            limit=20,
            verbose=False,
        )
        logger.info(
            "givemoney profile sync attempted for %s: scanned=%s matched=%s upserted=%s",
            safe_name,
            result.get("scanned"),
            result.get("matched"),
            result.get("upserted"),
        )
    except Exception as e:
        logger.warning("givemoney profile sync failed for %s: %s", safe_name, e)


def _clean_pledge_share_title(value: str) -> str:
    text = (value or "").strip()
    text = re.sub(r"^\s*<[^>]+>\s*", "", text)
    text = re.sub(r"^\s*\[[^\]]+\]\s*", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text or "공약"


_PLEDGE_SHARE_SUMMARY_VERSION = "rule-v1"


def _normalize_share_sentence(text: str, *, limit: int) -> str:
    text = re.sub(r"<[^>]+>", " ", str(text or ""))
    text = re.sub(r"\[[^\]]+\]", " ", text)
    text = re.sub(r"\([^)]*\)", " ", text)
    text = re.sub(r"\s+", " ", text).strip(" -:;,.")
    if not text:
        return ""
    if re.search(r"(합니다|겠습니다|됩니다|입니다)$", text):
        text = f"{text}."
    elif re.search(r"(다|요|음|함|임|됨)$", text):
        text = f"{text}."
    elif not re.search(r"[.!?]$", text):
        if text.endswith(("개선", "확대", "지원", "추진", "도입", "구축", "정비", "강화", "완화")):
            text = f"{text}합니다."
        else:
            text = f"{text}입니다."
    return _truncate_complete(text, limit)


def _extract_share_summary_candidates(title: str, content: str) -> list[str]:
    raw_lines = [line.strip() for line in str(content or "").replace("\r", "").split("\n") if line.strip()]
    candidates: list[str] = []
    seen: set[str] = set()

    def add(text: str) -> None:
        normalized = _normalize_share_sentence(text, limit=120)
        if not normalized:
            return
        key = re.sub(r"\W+", "", normalized)
        if len(key) < 8 or key in seen:
            return
        seen.add(key)
        candidates.append(normalized)

    bullet_lines = []
    for line in raw_lines:
        cleaned = re.sub(r"^[-•·*\s]+|^\d+[.)]\s*|^[①-⑳]\s*", "", line).strip()
        if cleaned != line or re.match(r"^[-•·*]|^\d+[.)]|^[①-⑳]", line):
            bullet_lines.append(cleaned)
    for line in bullet_lines:
        add(line)

    merged = " ".join(raw_lines)
    for part in re.split(r"(?<=[.!?])\s+|\n+|(?<=다\.)\s+|(?<=요\.)\s+|(?<=니다\.)\s+", merged):
        add(part)

    add(title)
    return candidates


def _build_rule_based_pledge_share_summary(title: str, content: str) -> PledgeShareSummaryResponse:
    clean_title = _clean_pledge_share_title(title)
    candidates = _extract_share_summary_candidates(clean_title, content)
    headline = candidates[0] if candidates else "핵심 내용을 준비 중입니다."
    bullets = [item for item in candidates[1:] if item != headline][:2]
    return PledgeShareSummaryResponse(
        title=_truncate_complete(clean_title, 12),
        headline=_truncate_complete(headline or clean_title, 55),
        bullets=[_truncate_complete(item, 40) for item in bullets],
    )


def _fallback_pledge_share_summary(title: str, content: str) -> PledgeShareSummaryResponse:
    return _build_rule_based_pledge_share_summary(title, content)


def _load_stored_share_summary(row) -> Optional[PledgeShareSummaryResponse]:
    raw_bullets = row["share_summary_bullets"] if "share_summary_bullets" in row.keys() else None
    bullets: list[str] = []
    if raw_bullets:
        try:
            bullets = [str(item).strip() for item in json.loads(raw_bullets) if str(item).strip()][:2]
        except Exception:
            bullets = []
    title = (row["share_summary_title"] if "share_summary_title" in row.keys() else "") or ""
    headline = (row["share_summary_headline"] if "share_summary_headline" in row.keys() else "") or ""
    if not title or not headline:
        return None
    return PledgeShareSummaryResponse(title=title, headline=headline, bullets=bullets)


def _persist_pledge_share_summary(pledge_id: int, source_hash: str, summary: PledgeShareSummaryResponse) -> None:
    from backend.database import get_connection

    conn = get_connection()
    try:
        conn.execute(
            """
            UPDATE candidate_pledges
            SET share_summary_title = ?,
                share_summary_headline = ?,
                share_summary_bullets = ?,
                share_summary_version = ?,
                share_summary_source_hash = ?,
                share_summary_updated_at = datetime('now')
            WHERE id = ?
            """,
            (
                summary.title,
                summary.headline,
                json.dumps(summary.bullets, ensure_ascii=False),
                _PLEDGE_SHARE_SUMMARY_VERSION,
                source_hash,
                pledge_id,
            ),
        )
        conn.commit()
    finally:
        conn.close()


def _get_or_create_persisted_pledge_share_summary(payload: dict) -> PledgeShareSummaryResponse:
    import hashlib

    source_hash = hashlib.md5(f'{payload["title"]}|{payload["content"]}'.encode("utf-8")).hexdigest()
    stored = payload.get("share_summary")
    if (
        stored is not None
        and payload.get("share_summary_version") == _PLEDGE_SHARE_SUMMARY_VERSION
        and payload.get("share_summary_source_hash") == source_hash
    ):
        return stored

    summary = _build_rule_based_pledge_share_summary(payload["title"], payload["content"])
    _persist_pledge_share_summary(int(payload["pledge_id"]), source_hash, summary)
    return summary


def _fetch_candidate_pledges_current_public(
    candidate_id: int,
    limit: Optional[int] = None,
) -> list[CandidatePledgeResponse]:
    """공개용 공약 목록.

    - APPROVED 공약: 현재 내용을 그대로 반환.
    - PENDING 공약(수정 후 재승인 대기): review_history에서 마지막 APPROVED
      스냅샷의 내용을 대신 반환 → 관리자가 새 버전을 승인하기 전까지 이전
      승인 버전이 계속 공개됨.
    - REJECTED / 스냅샷 없는 PENDING: 반환하지 않음.
    """
    from backend.database import get_connection

    conn = get_connection()
    try:
        # 1) 현재 APPROVED인 공약
        sql_approved = """
            SELECT cp.id, cp.title, cp.content, cp.total_score, cp.created_at
            FROM candidate_pledges cp
            WHERE cp.candidate_id = ?
              AND (
                  COALESCE(cp.approval_status, 'PENDING') = 'APPROVED'
                  OR (
                      COALESCE(cp.approval_status, 'PENDING') = 'PENDING'
                      AND NOT EXISTS (
                          SELECT 1 FROM candidate_pledge_review_history h0
                          WHERE h0.pledge_id = cp.id
                      )
                  )
              )
            ORDER BY cp.priority ASC, cp.id DESC
        """
        params_approved: tuple = (candidate_id,)
        if limit is not None:
            sql_approved += " LIMIT ?"
            params_approved = (candidate_id, limit)
        approved_rows = conn.execute(sql_approved, params_approved).fetchall()
        approved_ids = {r["id"] for r in approved_rows}

        # 2) PENDING인 공약 중 이전에 APPROVED 스냅샷이 있는 것 → 이전 버전으로 공개
        sql_pending_fallback = """
            SELECT cp.id AS pledge_id, h.title, h.content, h.total_score, h.reviewed_at AS created_at
            FROM candidate_pledges cp
            JOIN (
                SELECT pledge_id, title, content, total_score, reviewed_at,
                       ROW_NUMBER() OVER (PARTITION BY pledge_id ORDER BY reviewed_at DESC) AS rn
                FROM candidate_pledge_review_history
                WHERE approval_status = 'APPROVED'
                  AND pledge_id IS NOT NULL
            ) h ON h.pledge_id = cp.id AND h.rn = 1
            WHERE cp.candidate_id = ?
              AND cp.approval_status = 'PENDING'
        """
        pending_rows = conn.execute(sql_pending_fallback, (candidate_id,)).fetchall()

        results: list[CandidatePledgeResponse] = []
        for r in approved_rows:
            results.append(CandidatePledgeResponse(
                id=r["id"],
                title=r["title"],
                content=r["content"],
                total_score=r["total_score"],
                created_at=r["created_at"],
            ))
        for r in pending_rows:
            if r["pledge_id"] not in approved_ids:
                results.append(CandidatePledgeResponse(
                    id=r["pledge_id"],
                    title=r["title"],
                    content=r["content"],
                    total_score=r["total_score"],
                    created_at=r["created_at"],
                ))

        if limit is not None:
            results = results[:limit]
        return results
    finally:
        conn.close()


def _analysis_title_from_text(text: Optional[str]) -> str:
    raw = (text or "").replace("\r\n", "\n")
    for line in raw.split("\n"):
        line = line.strip()
        if line:
            return line[:100]
    return "불러온 공약"


def _normalize_snapshot_title(title: str) -> str:
    value = (title or "").strip()
    return re.sub(r"^\[\s*핵심\s*공약\s*\d+\s*\]\s*", "", value)


def _as_of_kst_to_utc_string(as_of: str) -> str:
    import datetime as _dt

    raw = (as_of or "").strip()
    if not raw:
        return raw
    try:
        dt = _dt.datetime.fromisoformat(raw.replace(" ", "T"))
    except ValueError:
        return raw
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=_dt.timezone(_dt.timedelta(hours=9)))
    return dt.astimezone(_dt.timezone.utc).strftime("%Y-%m-%d %H:%M:%S")


def _weekly_snapshot_limit(candidate_id: int, as_of: str) -> Optional[int]:
    import datetime as _dt

    raw = (as_of or "").strip()
    if not raw:
        return None
    try:
        dt = _dt.datetime.fromisoformat(raw.replace(" ", "T"))
    except ValueError:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=_dt.timezone(_dt.timedelta(hours=9)))
    local_date = dt.astimezone(_dt.timezone(_dt.timedelta(hours=9))).date()
    week_start = (local_date - _dt.timedelta(days=local_date.weekday())).isoformat()

    from backend.database import get_connection

    conn = get_connection()
    try:
        row = conn.execute(
            """
            SELECT scored_pledge_count
            FROM weekly_champions
            WHERE candidate_id = ?
              AND week_start = ?
            LIMIT 1
            """,
            (candidate_id, week_start),
        ).fetchone()
        if row is None:
            return None
        count = int(row["scored_pledge_count"] or 0)
        return count if count > 0 else None
    finally:
        conn.close()


def _is_public_nomination_status(status_note: Optional[str], applicant_match_id: Optional[int] = None) -> bool:
    if applicant_match_id is not None:
        return True
    note = (status_note or "").strip()
    return not note or note == PUBLIC_NOMINATION_NOTE


def _public_candidate_sql_condition() -> str:
    return """
        c.approval_status IN ('APPROVED', 'MIXED')
        AND EXISTS (
            SELECT 1 FROM candidate_pledges cp
            WHERE cp.candidate_id = c.id
              AND (cp.approval_status = 'APPROVED'
                   OR (cp.approval_status = 'PENDING'
                       AND EXISTS (SELECT 1 FROM candidate_pledge_review_history h
                                   WHERE h.pledge_id = cp.id AND h.approval_status = 'APPROVED')))
        )
        AND (
            u.applicant_match_id IS NOT NULL
            OR TRIM(COALESCE(pa.status_note, '')) = ''
            OR EXISTS (
                SELECT 1
                FROM party_applicants pa_public
                WHERE TRIM(COALESCE(pa_public.status_note, '')) = '공천 확정'
                  AND (
                      (
                          lower(trim(COALESCE(pa_public.email, ''))) <> ''
                          AND lower(trim(COALESCE(pa_public.email, ''))) = lower(trim(COALESCE(u.email, '')))
                          AND lower(replace(trim(COALESCE(pa_public.name, '')), ' ', '')) = lower(replace(trim(COALESCE(u.name, '')), ' ', ''))
                      )
                      OR (
                          replace(replace(replace(replace(trim(COALESCE(pa_public.phone, '')), '-', ''), ' ', ''), '(', ''), ')', '') <> ''
                          AND replace(replace(replace(replace(trim(COALESCE(pa_public.phone, '')), '-', ''), ' ', ''), '(', ''), ')', '') = replace(replace(replace(replace(trim(COALESCE(u.phone, '')), '-', ''), ' ', ''), '(', ''), ')', '')
                          AND lower(replace(trim(COALESCE(pa_public.name, '')), ' ', '')) = lower(replace(trim(COALESCE(u.name, '')), ' ', ''))
                      )
                  )
            )
        )
    """


def _leaderboard_reset_monday():
    return datetime(2026, 4, 7).date()


def _fetch_candidate_analysis_snapshot(
    user_id: Optional[int],
    as_of: str,
    limit: Optional[int] = None,
) -> list[CandidatePledgeResponse]:
    if not user_id:
        return []
    from backend.database import get_connection

    conn = get_connection()
    try:
        as_of_utc = _as_of_kst_to_utc_string(as_of)
        rows = conn.execute(
            """
            SELECT id, input_text, total_score, created_at
            FROM analysis_history
            WHERE user_id = ?
              AND status_code = 200
              AND datetime(created_at) <= datetime(?)
            ORDER BY datetime(created_at) DESC, id DESC
            LIMIT 100
            """,
            (user_id, as_of_utc),
        ).fetchall()
        picked: list[dict] = []
        seen_titles: set[str] = set()
        for row in rows:
            title = _analysis_title_from_text(row["input_text"])
            normalized_title = _normalize_snapshot_title(title)
            if normalized_title in seen_titles:
                continue
            seen_titles.add(normalized_title)
            picked.append(
                {
                    "id": None,
                    "title": normalized_title or title,
                    "content": row["input_text"],
                    "total_score": row["total_score"],
                    "created_at": row["created_at"],
                }
            )
            if limit is not None and len(picked) >= limit:
                break
        picked.sort(key=lambda item: item["created_at"] or "")
        return [
            CandidatePledgeResponse(
                id=item["id"],
                title=item["title"],
                content=item["content"],
                total_score=item["total_score"],
                created_at=item["created_at"],
            )
            for item in picked
        ]
    finally:
        conn.close()


def _fetch_candidate_pledges_snapshot(
    candidate_id: int,
    as_of: str,
    user_id: Optional[int] = None,
) -> list[CandidatePledgeResponse]:
    from backend.database import get_connection

    conn = get_connection()
    try:
        as_of_utc = _as_of_kst_to_utc_string(as_of)
        snapshot = conn.execute(
            """
            SELECT snapshot_group
            FROM candidate_pledge_review_history
            WHERE candidate_id = ?
              AND approval_status = 'APPROVED'
              AND datetime(reviewed_at) <= datetime(?)
            ORDER BY datetime(reviewed_at) DESC, snapshot_group DESC, id DESC
            LIMIT 1
            """,
            (candidate_id, as_of_utc),
        ).fetchone()
        if snapshot is None:
            rows = conn.execute(
                """
                SELECT id, title, content, total_score, created_at
                FROM candidate_pledges
                WHERE candidate_id = ?
                  AND approval_status = 'APPROVED'
                  AND datetime(created_at) <= datetime(?)
                ORDER BY priority ASC, datetime(created_at) DESC, id DESC
                """,
                (candidate_id, as_of_utc),
            ).fetchall()
            fallback_rows = [
                CandidatePledgeResponse(
                    id=r["id"],
                    title=r["title"],
                    content=r["content"],
                    total_score=r["total_score"],
                    created_at=r["created_at"],
                )
                for r in rows
            ]
            if fallback_rows:
                return fallback_rows
            analysis_limit = _weekly_snapshot_limit(candidate_id, as_of)
            analysis_rows = _fetch_candidate_analysis_snapshot(user_id, as_of, limit=analysis_limit)
            return analysis_rows

        rows = conn.execute(
            """
            SELECT pledge_id, title, content, total_score, pledge_created_at
            FROM candidate_pledge_review_history
            WHERE candidate_id = ?
              AND snapshot_group = ?
              AND approval_status = 'APPROVED'
            ORDER BY priority ASC, id ASC
            """,
            (candidate_id, snapshot["snapshot_group"]),
        ).fetchall()
        return [
            CandidatePledgeResponse(
                id=r["pledge_id"],
                title=r["title"],
                content=r["content"],
                total_score=r["total_score"],
                created_at=r["pledge_created_at"],
            )
            for r in rows
        ]
    finally:
        conn.close()


def _recalculate_candidate_approval(conn, candidate_id: int):
    rows = conn.execute(
        """
        SELECT approval_status, rejection_reason
        FROM candidate_pledges
        WHERE candidate_id = ?
        ORDER BY priority ASC, id ASC
        """,
        (candidate_id,),
    ).fetchall()
    if not rows:
        conn.execute(
            "UPDATE candidates SET approval_status = 'PENDING', rejection_reason = NULL, updated_at = datetime('now') WHERE id = ?",
            (candidate_id,),
        )
        return "PENDING"

    statuses = [(row["approval_status"] or "PENDING").upper() for row in rows]
    unique_statuses = set(statuses)
    candidate_status = "PENDING"
    candidate_reason = None
    if unique_statuses == {"APPROVED"}:
        candidate_status = "APPROVED"
    elif unique_statuses == {"REJECTED"}:
        candidate_status = "REJECTED"
        reasons = [row["rejection_reason"] for row in rows if row["rejection_reason"]]
        candidate_reason = reasons[0] if reasons else None
    elif "APPROVED" in unique_statuses or "REJECTED" in unique_statuses:
        candidate_status = "MIXED"

    conn.execute(
        "UPDATE candidates SET approval_status = ?, rejection_reason = ?, updated_at = datetime('now') WHERE id = ?",
        (candidate_status, candidate_reason, candidate_id),
    )
    return candidate_status


def _resolve_region_name(code: str) -> str:
    default_name = REGION_NAME_MAP.get(code, code)
    from backend.database import get_connection

    conn = get_connection()
    try:
        row = conn.execute(
            "SELECT region_name FROM region_codes WHERE region_code = ? LIMIT 1",
            (code,),
        ).fetchone()
        if row and row["region_name"]:
            return str(row["region_name"])
        return default_name
    except Exception:
        return default_name
    finally:
        conn.close()


def _normalize_district_code(value: Optional[str]) -> Optional[str]:
    if value is None or not isinstance(value, str):
        return None
    code = value.strip()
    if not code:
        return None
    if not re.fullmatch(r"[A-Za-z0-9가-힣_:\-]{2,120}", code):
        raise HTTPException(status_code=400, detail="district_code 형식이 올바르지 않습니다.")
    return code


def _derive_district_code(region_code: str, district_code: Optional[str], district_name: Optional[str]) -> Optional[str]:
    if district_code:
        return district_code
    name = (district_name or "").strip()
    if not name:
        return None
    norm = re.sub(r"\s+", "", name)
    norm = re.sub(r"[^0-9A-Za-z가-힣_-]", "", norm)
    if not norm:
        return None
    return f"{region_code}:{norm}"


def _normalize_election_type(value: Optional[str]) -> Optional[str]:
    if value is None or not isinstance(value, str):
        return None
    text = value.strip()
    if not text:
        return None
    if not re.fullmatch(r"[A-Za-z0-9_-]{1,40}", text):
        raise HTTPException(status_code=400, detail="election_type 형식이 올바르지 않습니다.")
    return text


class AdminCandidatePledgeInput(BaseModel):
    title: str = Field(..., min_length=1, max_length=100, description="공약 제목")
    content: Optional[str] = Field(default=None, max_length=50000, description="공약 세부내용")
    priority: int = Field(default=100, ge=1, le=9999, description="정렬 우선순위(작을수록 상위)")


class AdminCandidateUpsertBody(BaseModel):
    name: str = Field(..., min_length=1, max_length=80, description="후보명")
    district_name: Optional[str] = Field(default=None, max_length=120, description="선거구명")
    district_code: Optional[str] = Field(default=None, max_length=64, description="선거구 코드")
    region_code: str = Field(..., description="행정구역 코드")
    election_type: str = Field(default="local", min_length=1, max_length=40, description="선거 구분")
    election_level: str = Field(default="regional", min_length=1, max_length=40, description="선거 레벨")
    pledges: list[AdminCandidatePledgeInput] = Field(default_factory=list, description="후보 공약 목록")


@app.get("/api/admin/candidates", tags=["admin", "candidates"])
def admin_list_candidates(
    request: Request,
    region_code: Optional[str] = Query(default=None, description="행정구역 코드"),
    election_type: Optional[str] = Query(default=None, description="선거 타입"),
):
    """관리자 전용 후보 목록 (user_id/등록자 정보 포함)."""
    _ensure_db_ready()
    _ = require_admin(request)
    code = (region_code or "").strip()
    sel_et = _normalize_election_type(election_type)
    from backend.database import get_connection

    conn = get_connection()
    try:
        sql = """
            SELECT c.id, c.name, c.district_name, c.district_code, c.region_code,
                   c.election_type, c.election_level, c.user_id, c.approval_status,
                   c.rejection_reason,
                   u.email AS user_email, u.name AS user_name
            FROM candidates c
            LEFT JOIN users u ON c.user_id = u.id
            WHERE 1=1
        """
        params: list[object] = []
        if code:
            sql += " AND c.region_code = ?"
            params.append(code)
        if sel_et:
            sql += " AND c.election_type = ?"
            params.append(sel_et)
        sql += """
            ORDER BY c.region_code,
                CASE c.election_type
                    WHEN 'metro_mayor' THEN 1
                    WHEN 'local_mayor' THEN 2
                    WHEN 'regional_council' THEN 3
                    WHEN 'local_council' THEN 4
                    ELSE 5
                END,
                COALESCE(c.district_name, '') ASC,
                c.name ASC
        """
        rows = conn.execute(sql, tuple(params)).fetchall()
    finally:
        conn.close()

    candidate_ids = [int(r["id"]) for r in rows]
    pledges_map: dict[int, list[dict]] = {cid: [] for cid in candidate_ids}
    if candidate_ids:
        placeholders = ",".join("?" * len(candidate_ids))
        conn2 = get_connection()
        try:
            p_rows = conn2.execute(
                f"""
                SELECT id, candidate_id, title, content, total_score, approval_status, rejection_reason, created_at
                FROM candidate_pledges
                WHERE candidate_id IN ({placeholders})
                ORDER BY priority ASC, id ASC
                """,
                tuple(candidate_ids),
            ).fetchall()
            for p in p_rows:
                pledges_map[int(p["candidate_id"])].append({
                    "id": p["id"],
                    "title": p["title"],
                    "content": p["content"],
                    "total_score": p["total_score"],
                    "approval_status": p["approval_status"] or "PENDING",
                    "rejection_reason": p["rejection_reason"] or "",
                    "created_at": p["created_at"],
                })
        finally:
            conn2.close()

    result = []
    for r in rows:
        cid = int(r["id"])
        rc = r["region_code"] or ""
        result.append({
            "candidate_id": cid,
            "name": r["name"],
            "district_name": r["district_name"],
            "district_code": _derive_district_code(rc, r["district_code"], r["district_name"]),
            "region_code": rc,
            "region_name": REGION_NAME_MAP.get(rc, rc),
            "election_type": r["election_type"],
            "election_level": r["election_level"],
            "user_id": r["user_id"],
            "registered_by": r["user_email"] or r["user_name"] if r["user_id"] else None,
            "approval_status": r["approval_status"] or "PENDING",
            "rejection_reason": r["rejection_reason"] or "",
            "pledges": pledges_map.get(cid, []),
        })
    return result


@app.delete("/api/admin/candidates/{candidate_id}", tags=["admin", "candidates"])
def admin_delete_candidate(candidate_id: int, request: Request):
    """관리자 전용 후보 삭제."""
    _ensure_db_ready()
    _ = require_admin(request)
    from backend.database import get_connection

    conn = get_connection()
    try:
        existing = conn.execute("SELECT id FROM candidates WHERE id = ?", (candidate_id,)).fetchone()
        if existing is None:
            raise HTTPException(status_code=404, detail="후보를 찾을 수 없습니다.")
        conn.execute("DELETE FROM candidate_pledges WHERE candidate_id = ?", (candidate_id,))
        conn.execute("DELETE FROM candidates WHERE id = ?", (candidate_id,))
        conn.commit()
    except HTTPException:
        conn.rollback()
        raise
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    return {"ok": True}


@app.post("/api/admin/candidates/{candidate_id}/approve", tags=["admin", "candidates"])
def admin_approve_candidate(candidate_id: int, request: Request):
    """관리자 전용 후보 승인."""
    _ensure_db_ready()
    _ = require_admin(request)
    from backend.database import get_connection

    conn = get_connection()
    try:
        existing = conn.execute(
            """SELECT c.id, c.approval_status, c.rejection_reason, c.name AS candidate_name,
                      u.email AS user_email, u.name AS user_name
               FROM candidates c
               LEFT JOIN users u ON u.id = c.user_id
               WHERE c.id = ?""",
            (candidate_id,),
        ).fetchone()
        if existing is None:
            raise HTTPException(status_code=404, detail="후보를 찾을 수 없습니다.")
        _snapshot_candidate_pledges(
            conn,
            candidate_id,
            approval_status="APPROVED",
            rejection_reason=None,
            source_action="APPROVE",
        )
        conn.execute(
            "UPDATE candidate_pledges SET approval_status = 'APPROVED', rejection_reason = NULL WHERE candidate_id = ?",
            (candidate_id,),
        )
        _recalculate_candidate_approval(conn, candidate_id)
        conn.commit()
    except HTTPException:
        conn.rollback()
        raise
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    # 이메일 알림
    if existing and existing["user_email"]:
        try:
            from backend.email_sender import send_candidate_approval_status_email
            send_candidate_approval_status_email(
                to_email=existing["user_email"],
                status="APPROVED",
                name=existing["user_name"] or "",
                candidate_name=existing["candidate_name"] or "",
            )
        except Exception as e:
            logger.warning("공약 승인 알림 메일 발송 실패 (무시): %s", e)

    return {"ok": True, "approval_status": "APPROVED"}


class RejectBody(BaseModel):
    reason: Optional[str] = Field(default=None, max_length=500, description="거절 사유")


@app.post("/api/admin/candidates/{candidate_id}/reject", tags=["admin", "candidates"])
def admin_reject_candidate(candidate_id: int, request: Request, body: Optional[RejectBody] = None):
    """관리자 전용 후보 거절."""
    _ensure_db_ready()
    _ = require_admin(request)
    from backend.database import get_connection

    rejection_reason = (body.reason or "").strip() if body else ""

    conn = get_connection()
    try:
        existing = conn.execute(
            """SELECT c.id, c.approval_status, c.name AS candidate_name,
                      u.email AS user_email, u.name AS user_name
               FROM candidates c
               LEFT JOIN users u ON u.id = c.user_id
               WHERE c.id = ?""",
            (candidate_id,),
        ).fetchone()
        if existing is None:
            raise HTTPException(status_code=404, detail="후보를 찾을 수 없습니다.")
        _snapshot_candidate_pledges(
            conn,
            candidate_id,
            approval_status="REJECTED",
            rejection_reason=rejection_reason,
            source_action="REJECT",
        )
        conn.execute(
            "UPDATE candidate_pledges SET approval_status = 'REJECTED', rejection_reason = ? WHERE candidate_id = ?",
            (rejection_reason or None, candidate_id),
        )
        _recalculate_candidate_approval(conn, candidate_id)
        conn.commit()
    except HTTPException:
        conn.rollback()
        raise
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    # 이메일 알림
    if existing and existing["user_email"]:
        try:
            from backend.email_sender import send_candidate_approval_status_email
            send_candidate_approval_status_email(
                to_email=existing["user_email"],
                status="REJECTED",
                name=existing["user_name"] or "",
                candidate_name=existing["candidate_name"] or "",
                rejection_reason=rejection_reason,
            )
        except Exception as e:
            logger.warning("공약 거절 알림 메일 발송 실패 (무시): %s", e)

    return {"ok": True, "approval_status": "REJECTED"}


# 새 엔진 배포일 — 이 날짜 이전 분석 결과는 재분석 필요
ANALYSIS_ENGINE_CUTOFF = "2026-04-13"

@app.post("/api/admin/pledges/{pledge_id}/approve", tags=["admin", "candidates"])
def admin_approve_pledge(pledge_id: int, request: Request, force: bool = False):
    _ensure_db_ready()
    _ = require_admin(request)
    from backend.database import get_connection

    conn = get_connection()
    try:
        row = conn.execute(
            """
            SELECT cp.id, cp.candidate_id, cp.analyzed_at, cp.title
            FROM candidate_pledges cp
            WHERE cp.id = ?
            """,
            (pledge_id,),
        ).fetchone()
        if row is None:
            raise HTTPException(status_code=404, detail="공약을 찾을 수 없습니다.")

        # 구버전 엔진 분석 결과 차단 (force=true 로 어드민이 명시적으로 우회 가능)
        analyzed_at = row["analyzed_at"] or ""
        if not force and analyzed_at and analyzed_at[:10] < ANALYSIS_ENGINE_CUTOFF:
            raise HTTPException(
                status_code=409,
                detail=f"이 공약은 구버전 엔진({analyzed_at[:10]})으로 분석되었습니다. "
                       f"새 엔진({ANALYSIS_ENGINE_CUTOFF} 이후) 재분석 후 승인해 주세요. "
                       f"강제 승인은 ?force=true 파라미터를 사용하세요."
            )

        candidate_id = int(row["candidate_id"])
        _snapshot_candidate_pledges(conn, candidate_id, approval_status="APPROVED", source_action="PLEDGE_APPROVE")
        conn.execute(
            "UPDATE candidate_pledges SET approval_status = 'APPROVED', rejection_reason = NULL WHERE id = ?",
            (pledge_id,),
        )
        candidate_status = _recalculate_candidate_approval(conn, candidate_id)
        conn.commit()
        return {"ok": True, "approval_status": "APPROVED", "candidate_status": candidate_status}
    except HTTPException:
        conn.rollback()
        raise
    except Exception:
        conn.rollback()
        logger.exception("공약 승인 처리 중 오류")
        raise HTTPException(status_code=500, detail="공약 승인 처리 중 오류가 발생했습니다.")
    finally:
        conn.close()


@app.post("/api/admin/pledges/{pledge_id}/reject", tags=["admin", "candidates"])
def admin_reject_pledge(pledge_id: int, request: Request, body: Optional[RejectBody] = None):
    _ensure_db_ready()
    _ = require_admin(request)
    rejection_reason = (body.reason or "").strip() if body else ""
    from backend.database import get_connection

    conn = get_connection()
    try:
        row = conn.execute(
            """
            SELECT cp.id, cp.candidate_id
            FROM candidate_pledges cp
            WHERE cp.id = ?
            """,
            (pledge_id,),
        ).fetchone()
        if row is None:
            raise HTTPException(status_code=404, detail="공약을 찾을 수 없습니다.")
        candidate_id = int(row["candidate_id"])
        _snapshot_candidate_pledges(
            conn,
            candidate_id,
            approval_status="REJECTED",
            rejection_reason=rejection_reason,
            source_action="PLEDGE_REJECT",
        )
        conn.execute(
            "UPDATE candidate_pledges SET approval_status = 'REJECTED', rejection_reason = ? WHERE id = ?",
            (rejection_reason or None, pledge_id),
        )
        candidate_status = _recalculate_candidate_approval(conn, candidate_id)
        conn.commit()
        return {"ok": True, "approval_status": "REJECTED", "candidate_status": candidate_status}
    except HTTPException:
        conn.rollback()
        raise
    except Exception:
        conn.rollback()
        logger.exception("공약 거절 처리 중 오류")
        raise HTTPException(status_code=500, detail="공약 거절 처리 중 오류가 발생했습니다.")
    finally:
        conn.close()


@app.post("/api/admin/candidates", response_model=CandidateDetailResponse, tags=["admin", "candidates"])
def admin_create_candidate(body: AdminCandidateUpsertBody, request: Request):
    """관리자 전용 후보 등록 API. region_code 검증을 강제한다."""
    _ensure_db_ready()
    user = require_admin(request)

    code = _validate_region_code(body.region_code)
    district_code = _normalize_district_code(body.district_code)
    election_type = _normalize_election_type(body.election_type) or "local"
    resolved_district_code = _derive_district_code(code, district_code, body.district_name)
    district_name_clean = (body.district_name or "").strip()
    from backend.database import get_connection

    conn = get_connection()
    try:
        cur = conn.execute(
            """
            INSERT INTO candidates (name, district_name, district_code, region_code, election_type, election_level, approval_status, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, 'PENDING', datetime('now'))
            """,
            (
                body.name.strip(),
                district_name_clean or None,
                resolved_district_code,
                code,
                election_type,
                (body.election_level or "regional").strip(),
            ),
        )
        if resolved_district_code and district_name_clean:
            conn.execute(
                "INSERT INTO region_codes (region_code, region_name, aliases_json, updated_at) VALUES (?, ?, '[]', datetime('now')) ON CONFLICT(region_code) DO NOTHING",
                (code, REGION_NAME_MAP.get(code, code)),
            )
            conn.execute(
                """
                INSERT INTO district_codes (district_code, district_name, region_code, election_type, aliases_json, updated_at)
                VALUES (?, ?, ?, ?, '[]', datetime('now'))
                ON CONFLICT(district_code) DO UPDATE SET
                    district_name = excluded.district_name,
                    region_code = excluded.region_code,
                    election_type = excluded.election_type,
                    updated_at = datetime('now')
                """,
                (resolved_district_code, district_name_clean, code, election_type),
            )
        candidate_id = int(cur.lastrowid)
        for idx, pledge in enumerate(body.pledges):
            conn.execute(
                """
                INSERT INTO candidate_pledges (candidate_id, title, content, priority)
                VALUES (?, ?, ?, ?)
                """,
                (
                    candidate_id,
                    pledge.title.strip(),
                    (pledge.content or "").strip() or None,
                    pledge.priority if pledge.priority else (idx + 1),
                ),
            )
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    return _get_candidate_detail_any_status(candidate_id)


@app.put("/api/admin/candidates/{candidate_id}", response_model=CandidateDetailResponse, tags=["admin", "candidates"])
def admin_update_candidate(candidate_id: int, body: AdminCandidateUpsertBody, request: Request):
    """관리자 전용 후보 수정 API. region_code 검증을 강제한다."""
    _ensure_db_ready()
    user = require_admin(request)

    code = _validate_region_code(body.region_code)
    district_code = _normalize_district_code(body.district_code)
    election_type = _normalize_election_type(body.election_type) or "local"
    resolved_district_code = _derive_district_code(code, district_code, body.district_name)
    district_name_clean = (body.district_name or "").strip()
    from backend.database import get_connection

    conn = get_connection()
    try:
        existing = conn.execute("SELECT id FROM candidates WHERE id = ?", (candidate_id,)).fetchone()
        if existing is None:
            raise HTTPException(status_code=404, detail=f"candidate_id={candidate_id} 후보를 찾을 수 없습니다.")

        conn.execute(
            """
            UPDATE candidates
            SET name = ?, district_name = ?, district_code = ?, region_code = ?, election_type = ?, election_level = ?, approval_status = 'PENDING', updated_at = datetime('now')
            WHERE id = ?
            """,
            (
                body.name.strip(),
                district_name_clean or None,
                resolved_district_code,
                code,
                election_type,
                (body.election_level or "regional").strip(),
                candidate_id,
            ),
        )
        if resolved_district_code and district_name_clean:
            conn.execute(
                "INSERT INTO region_codes (region_code, region_name, aliases_json, updated_at) VALUES (?, ?, '[]', datetime('now')) ON CONFLICT(region_code) DO NOTHING",
                (code, REGION_NAME_MAP.get(code, code)),
            )
            conn.execute(
                """
                INSERT INTO district_codes (district_code, district_name, region_code, election_type, aliases_json, updated_at)
                VALUES (?, ?, ?, ?, '[]', datetime('now'))
                ON CONFLICT(district_code) DO UPDATE SET
                    district_name = excluded.district_name,
                    region_code = excluded.region_code,
                    election_type = excluded.election_type,
                    updated_at = datetime('now')
                """,
                (resolved_district_code, district_name_clean, code, election_type),
            )
        conn.execute("DELETE FROM candidate_pledges WHERE candidate_id = ?", (candidate_id,))
        for idx, pledge in enumerate(body.pledges):
            conn.execute(
                """
                INSERT INTO candidate_pledges (candidate_id, title, content, priority)
                VALUES (?, ?, ?, ?)
                """,
                (
                    candidate_id,
                    pledge.title.strip(),
                    (pledge.content or "").strip() or None,
                    pledge.priority if pledge.priority else (idx + 1),
                ),
            )
        conn.commit()
    except HTTPException:
        conn.rollback()
        raise
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    return _get_candidate_detail_any_status(candidate_id)


@app.get("/api/regions", response_model=list[RegionResponse], tags=["candidates"])
def get_regions():
    """지역 코드 테이블 기준으로 후보 수를 집계해 반환한다."""
    _ensure_db_ready()
    from backend.database import get_connection

    conn = get_connection()
    try:
        count_rows = conn.execute(
            """
            SELECT c.region_code, COUNT(*) AS candidate_count
            FROM candidates c
            LEFT JOIN users u ON u.id = c.user_id
            LEFT JOIN party_applicants pa ON pa.id = u.applicant_match_id
            WHERE c.approval_status IN ('APPROVED', 'MIXED')
              AND EXISTS (
                  SELECT 1 FROM candidate_pledges cp
                  WHERE cp.candidate_id = c.id
                    AND (cp.approval_status = 'APPROVED'
                         OR (cp.approval_status = 'PENDING'
                             AND EXISTS (SELECT 1 FROM candidate_pledge_review_history h
                                         WHERE h.pledge_id = cp.id AND h.approval_status = 'APPROVED')))
              )
              AND (
                  u.applicant_match_id IS NOT NULL
                  OR TRIM(COALESCE(pa.status_note, '')) = ''
                  OR EXISTS (
                      SELECT 1
                      FROM party_applicants pa_public
                      WHERE TRIM(COALESCE(pa_public.status_note, '')) = '공천 확정'
                        AND (
                            (
                                lower(trim(COALESCE(pa_public.email, ''))) <> ''
                                AND lower(trim(COALESCE(pa_public.email, ''))) = lower(trim(COALESCE(u.email, '')))
                                AND lower(replace(trim(COALESCE(pa_public.name, '')), ' ', '')) = lower(replace(trim(COALESCE(u.name, '')), ' ', ''))
                            )
                            OR (
                                replace(replace(replace(replace(trim(COALESCE(pa_public.phone, '')), '-', ''), ' ', ''), '(', ''), ')', '') <> ''
                                AND replace(replace(replace(replace(trim(COALESCE(pa_public.phone, '')), '-', ''), ' ', ''), '(', ''), ')', '') = replace(replace(replace(replace(trim(COALESCE(u.phone, '')), '-', ''), ' ', ''), '(', ''), ')', '')
                                AND lower(replace(trim(COALESCE(pa_public.name, '')), ' ', '')) = lower(replace(trim(COALESCE(u.name, '')), ' ', ''))
                            )
                        )
                  )
              )
            GROUP BY c.region_code
            """
        ).fetchall()
        count_map = {r["region_code"]: int(r["candidate_count"]) for r in count_rows}
    finally:
        conn.close()

    return [
        RegionResponse(
            region_code=code,
            region_name=name,
            candidate_count=count_map.get(code, 0),
        )
        for code, name in REGION_NAME_MAP.items()
    ]



@app.get("/api/stats/election-types", tags=["candidates"])
def get_election_type_counts(region_code: Optional[str] = Query(default=None)):
    """선거 타입별 승인된 후보 수를 반환한다 (지도 페이지 선거 타입 셀렉트 카운팅용).

    region_code 를 넘기면 해당 지역 한정으로 집계한다.
    """
    _ensure_db_ready()
    from backend.database import get_connection

    conn = get_connection()
    try:
        if region_code:
            rows = conn.execute(
                """
                SELECT c.election_type, COUNT(*) AS n
                FROM candidates c
                LEFT JOIN users u ON u.id = c.user_id
                LEFT JOIN party_applicants pa ON pa.id = u.applicant_match_id
                WHERE c.approval_status IN ('APPROVED', 'MIXED')
              AND EXISTS (
                  SELECT 1 FROM candidate_pledges cp
                  WHERE cp.candidate_id = c.id
                    AND (cp.approval_status = 'APPROVED'
                         OR (cp.approval_status = 'PENDING'
                             AND EXISTS (SELECT 1 FROM candidate_pledge_review_history h
                                         WHERE h.pledge_id = cp.id AND h.approval_status = 'APPROVED')))
              )
              AND (
                  u.applicant_match_id IS NOT NULL
                  OR TRIM(COALESCE(pa.status_note, '')) = ''
                  OR EXISTS (
                      SELECT 1
                      FROM party_applicants pa_public
                      WHERE TRIM(COALESCE(pa_public.status_note, '')) = '공천 확정'
                        AND (
                            (
                                lower(trim(COALESCE(pa_public.email, ''))) <> ''
                                AND lower(trim(COALESCE(pa_public.email, ''))) = lower(trim(COALESCE(u.email, '')))
                                AND lower(replace(trim(COALESCE(pa_public.name, '')), ' ', '')) = lower(replace(trim(COALESCE(u.name, '')), ' ', ''))
                            )
                            OR (
                                replace(replace(replace(replace(trim(COALESCE(pa_public.phone, '')), '-', ''), ' ', ''), '(', ''), ')', '') <> ''
                                AND replace(replace(replace(replace(trim(COALESCE(pa_public.phone, '')), '-', ''), ' ', ''), '(', ''), ')', '') = replace(replace(replace(replace(trim(COALESCE(u.phone, '')), '-', ''), ' ', ''), '(', ''), ')', '')
                                AND lower(replace(trim(COALESCE(pa_public.name, '')), ' ', '')) = lower(replace(trim(COALESCE(u.name, '')), ' ', ''))
                            )
                        )
                  )
              )
                  AND c.region_code = ?
                GROUP BY c.election_type
                """,
                (region_code,),
            ).fetchall()
        else:
            rows = conn.execute(
                """
                SELECT c.election_type, COUNT(*) AS n
                FROM candidates c
                LEFT JOIN users u ON u.id = c.user_id
                LEFT JOIN party_applicants pa ON pa.id = u.applicant_match_id
                WHERE c.approval_status IN ('APPROVED', 'MIXED')
              AND EXISTS (
                  SELECT 1 FROM candidate_pledges cp
                  WHERE cp.candidate_id = c.id
                    AND (cp.approval_status = 'APPROVED'
                         OR (cp.approval_status = 'PENDING'
                             AND EXISTS (SELECT 1 FROM candidate_pledge_review_history h
                                         WHERE h.pledge_id = cp.id AND h.approval_status = 'APPROVED')))
              )
              AND (
                  u.applicant_match_id IS NOT NULL
                  OR TRIM(COALESCE(pa.status_note, '')) = ''
                  OR EXISTS (
                      SELECT 1
                      FROM party_applicants pa_public
                      WHERE TRIM(COALESCE(pa_public.status_note, '')) = '공천 확정'
                        AND (
                            (
                                lower(trim(COALESCE(pa_public.email, ''))) <> ''
                                AND lower(trim(COALESCE(pa_public.email, ''))) = lower(trim(COALESCE(u.email, '')))
                                AND lower(replace(trim(COALESCE(pa_public.name, '')), ' ', '')) = lower(replace(trim(COALESCE(u.name, '')), ' ', ''))
                            )
                            OR (
                                replace(replace(replace(replace(trim(COALESCE(pa_public.phone, '')), '-', ''), ' ', ''), '(', ''), ')', '') <> ''
                                AND replace(replace(replace(replace(trim(COALESCE(pa_public.phone, '')), '-', ''), ' ', ''), '(', ''), ')', '') = replace(replace(replace(replace(trim(COALESCE(u.phone, '')), '-', ''), ' ', ''), '(', ''), ')', '')
                                AND lower(replace(trim(COALESCE(pa_public.name, '')), ' ', '')) = lower(replace(trim(COALESCE(u.name, '')), ' ', ''))
                            )
                        )
                  )
              )
                GROUP BY c.election_type
                """
            ).fetchall()
        return {r["election_type"]: int(r["n"]) for r in rows}
    finally:
        conn.close()


@app.get("/api/districts", response_model=list[DistrictResponse], tags=["candidates"])
def get_districts(
    region_code: Optional[str] = Query(default=None, description="행정구역 코드"),
    election_type: Optional[str] = Query(default=None, description="선거 타입(local, mayor, etc)"),
):
    """선택한 시/도(region_code)의 선거구 목록과 후보 수를 반환한다."""
    _ensure_db_ready()
    code = _validate_region_code(region_code)
    selected_election_type = _normalize_election_type(election_type)
    from backend.database import get_connection

    conn = get_connection()
    try:
        candidate_sql = """
            SELECT c.district_name, c.district_code, c.election_type,
                   pa.status_note AS applicant_status_note,
                   u.applicant_match_id AS applicant_match_id
            FROM candidates c
            LEFT JOIN users u ON u.id = c.user_id
            LEFT JOIN party_applicants pa ON pa.id = u.applicant_match_id
            WHERE c.region_code = ?
              AND c.approval_status IN ('APPROVED', 'MIXED')
              AND EXISTS (
                  SELECT 1 FROM candidate_pledges cp
                  WHERE cp.candidate_id = c.id
                    AND (cp.approval_status = 'APPROVED'
                         OR (cp.approval_status = 'PENDING'
                             AND EXISTS (SELECT 1 FROM candidate_pledge_review_history h
                                         WHERE h.pledge_id = cp.id AND h.approval_status = 'APPROVED')))
              )
              AND (
                  u.applicant_match_id IS NOT NULL
                  OR TRIM(COALESCE(pa.status_note, '')) = ''
                  OR EXISTS (
                      SELECT 1
                      FROM party_applicants pa_public
                      WHERE TRIM(COALESCE(pa_public.status_note, '')) = '공천 확정'
                        AND (
                            (
                                lower(trim(COALESCE(pa_public.email, ''))) <> ''
                                AND lower(trim(COALESCE(pa_public.email, ''))) = lower(trim(COALESCE(u.email, '')))
                                AND lower(replace(trim(COALESCE(pa_public.name, '')), ' ', '')) = lower(replace(trim(COALESCE(u.name, '')), ' ', ''))
                            )
                            OR (
                                replace(replace(replace(replace(trim(COALESCE(pa_public.phone, '')), '-', ''), ' ', ''), '(', ''), ')', '') <> ''
                                AND replace(replace(replace(replace(trim(COALESCE(pa_public.phone, '')), '-', ''), ' ', ''), '(', ''), ')', '') = replace(replace(replace(replace(trim(COALESCE(u.phone, '')), '-', ''), ' ', ''), '(', ''), ')', '')
                                AND lower(replace(trim(COALESCE(pa_public.name, '')), ' ', '')) = lower(replace(trim(COALESCE(u.name, '')), ' ', ''))
                            )
                        )
                  )
              )
              AND c.district_name IS NOT NULL
              AND TRIM(c.district_name) <> ''
        """
        params: list[object] = [code]
        if selected_election_type:
            candidate_sql += " AND election_type = ?"
            params.append(selected_election_type)
        candidate_rows = conn.execute(candidate_sql, tuple(params)).fetchall()

        district_rows = conn.execute(
            """
            SELECT district_code, district_name
            FROM district_codes
            WHERE region_code = ?
              AND (? IS NULL OR election_type = ?)
            """,
            (code, selected_election_type, selected_election_type),
        ).fetchall()
    finally:
        conn.close()

    count_map: dict[str, dict[str, object]] = {}
    for r in district_rows:
        d_code = (r["district_code"] or "").strip()
        d_name = (r["district_name"] or "").strip() or d_code
        if d_code:
            count_map[d_code] = {"district_name": d_name, "candidate_count": 0}

    for r in candidate_rows:
        district_name = (r["district_name"] or "").strip()
        district_code = _derive_district_code(code, r["district_code"], district_name)
        if not district_code:
            continue
        if district_code not in count_map:
            count_map[district_code] = {
                "district_name": district_name or district_code,
                "candidate_count": 0,
            }
        count_map[district_code]["candidate_count"] = int(count_map[district_code]["candidate_count"]) + 1

    return [
        DistrictResponse(
            district_code=dcode,
            district_name=str(meta["district_name"]),
            region_code=code,
            candidate_count=int(meta["candidate_count"]),
        )
        for dcode, meta in sorted(count_map.items(), key=lambda x: (-int(x[1]["candidate_count"]), str(x[1]["district_name"])))
    ]


@app.get("/api/candidates", response_model=list[CandidateListItemResponse], tags=["candidates"])
def get_candidates(
    region_code: Optional[str] = Query(default=None, description="행정구역 코드"),
    district_code: Optional[str] = Query(default=None, description="선거구 코드"),
    election_type: Optional[str] = Query(default=None, description="선거 타입(local, mayor, etc)"),
):
    """지역별 후보 목록 + 핵심 공약(최대 3개)을 반환한다."""
    _ensure_db_ready()
    code = _validate_region_code(region_code)
    selected_district_code = _normalize_district_code(district_code)
    selected_election_type = _normalize_election_type(election_type)
    from backend.database import get_connection

    conn = get_connection()
    try:
        sql = """
            SELECT c.id, c.name,
                   COALESCE(u.district_name, c.district_name) AS district_name,
                   c.district_code, c.region_code, c.election_type, c.election_level,
                   pa.status_note AS applicant_status_note,
                   u.applicant_match_id AS applicant_match_id
            FROM candidates c
            LEFT JOIN users u ON u.id = c.user_id
            LEFT JOIN party_applicants pa ON pa.id = u.applicant_match_id
            WHERE c.region_code = ?
              AND c.approval_status IN ('APPROVED', 'MIXED')
              AND EXISTS (
                  SELECT 1 FROM candidate_pledges cp
                  WHERE cp.candidate_id = c.id
                    AND (cp.approval_status = 'APPROVED'
                         OR (cp.approval_status = 'PENDING'
                             AND EXISTS (SELECT 1 FROM candidate_pledge_review_history h
                                         WHERE h.pledge_id = cp.id AND h.approval_status = 'APPROVED')))
              )
              AND (
                  u.applicant_match_id IS NOT NULL
                  OR TRIM(COALESCE(pa.status_note, '')) = ''
                  OR EXISTS (
                      SELECT 1
                      FROM party_applicants pa_public
                      WHERE TRIM(COALESCE(pa_public.status_note, '')) = '공천 확정'
                        AND (
                            (
                                lower(trim(COALESCE(pa_public.email, ''))) <> ''
                                AND lower(trim(COALESCE(pa_public.email, ''))) = lower(trim(COALESCE(u.email, '')))
                                AND lower(replace(trim(COALESCE(pa_public.name, '')), ' ', '')) = lower(replace(trim(COALESCE(u.name, '')), ' ', ''))
                            )
                            OR (
                                replace(replace(replace(replace(trim(COALESCE(pa_public.phone, '')), '-', ''), ' ', ''), '(', ''), ')', '') <> ''
                                AND replace(replace(replace(replace(trim(COALESCE(pa_public.phone, '')), '-', ''), ' ', ''), '(', ''), ')', '') = replace(replace(replace(replace(trim(COALESCE(u.phone, '')), '-', ''), ' ', ''), '(', ''), ')', '')
                                AND lower(replace(trim(COALESCE(pa_public.name, '')), ' ', '')) = lower(replace(trim(COALESCE(u.name, '')), ' ', ''))
                            )
                        )
                  )
              )
        """
        params: list[object] = [code]
        if selected_election_type:
            sql += " AND c.election_type = ?"
            params.append(selected_election_type)
        sql += """
            ORDER BY
                CASE c.election_type
                    WHEN 'metro_mayor' THEN 1
                    WHEN 'local_mayor' THEN 2
                    WHEN 'regional_council' THEN 3
                    WHEN 'local_council' THEN 4
                    ELSE 5
                END,
                COALESCE(c.district_name, '') ASC,
                c.name ASC
        """
        rows = conn.execute(sql, tuple(params)).fetchall()
    finally:
        conn.close()

    result: list[CandidateListItemResponse] = []
    for r in rows:
        if not _is_public_nomination_status(r["applicant_status_note"], r["applicant_match_id"]):
            continue
        candidate_id = int(r["id"])
        resolved_district_code = _derive_district_code(code, r["district_code"], r["district_name"])
        if selected_district_code and resolved_district_code != selected_district_code:
            continue
        result.append(
            CandidateListItemResponse(
                candidate_id=candidate_id,
                name=r["name"],
                district_name=r["district_name"],
                district_code=resolved_district_code,
                region_code=r["region_code"],
                election_type=r["election_type"],
                election_level=r["election_level"],
                pledges=_fetch_candidate_pledges(candidate_id, limit=3, public_only=True),
            )
        )
    return result


@app.get("/api/candidates/{candidate_id}", response_model=CandidateDetailResponse, tags=["candidates"])
def get_candidate_detail(candidate_id: int, as_of: Optional[str] = Query(default=None)):
    """후보 상세 정보와 공약 전체를 반환한다."""
    _ensure_db_ready()
    from backend.database import get_connection

    conn = get_connection()
    try:
        row = conn.execute(
            """
            SELECT c.id, c.name,
                   COALESCE(u.district_name, c.district_name) AS district_name,
                   c.district_code, c.region_code, c.election_type, c.election_level, c.approval_status, c.user_id,
                   pa.status_note AS applicant_status_note,
                   u.applicant_match_id AS applicant_match_id
            FROM candidates c
            LEFT JOIN users u ON u.id = c.user_id
            LEFT JOIN party_applicants pa ON pa.id = u.applicant_match_id
            WHERE c.id = ?
              AND c.approval_status IN ('APPROVED', 'MIXED')
              AND EXISTS (
                  SELECT 1 FROM candidate_pledges cp
                  WHERE cp.candidate_id = c.id
                    AND (cp.approval_status = 'APPROVED'
                         OR (cp.approval_status = 'PENDING'
                             AND EXISTS (SELECT 1 FROM candidate_pledge_review_history h
                                         WHERE h.pledge_id = cp.id AND h.approval_status = 'APPROVED')))
              )
            """,
            (candidate_id,),
        ).fetchone()
    finally:
        conn.close()

    if row is None or not _is_public_nomination_status(row["applicant_status_note"], row["applicant_match_id"]):
        raise HTTPException(status_code=404, detail=f"candidate_id={candidate_id} 후보를 찾을 수 없습니다.")

    code = row["region_code"]
    return CandidateDetailResponse(
        candidate_id=int(row["id"]),
        name=row["name"],
        district_name=row["district_name"],
        district_code=_derive_district_code(code, row["district_code"], row["district_name"]),
        region_code=code,
        region_name=_resolve_region_name(code),
        election_type=row["election_type"],
        election_level=row["election_level"],
        external_profile=_fetch_candidate_external_profile(int(row["id"])),
        pledges=_fetch_candidate_pledges_snapshot(int(row["id"]), as_of, int(row["user_id"]) if row["user_id"] is not None else None) if as_of else _fetch_candidate_pledges_current_public(int(row["id"]), limit=None),
    )


@app.get("/api/proxy-image", tags=["utility"])
def proxy_image(url: str = Query(..., description="외부 이미지 URL")):
    """카드뉴스 렌더링용 외부 이미지를 same-origin으로 중계한다."""
    from urllib.parse import urlparse
    from urllib.request import Request, urlopen

    parsed = urlparse(url)
    allowed_hosts = {"prod-api.givemoney.kr", "givemoney.kr"}
    if parsed.scheme not in {"http", "https"} or parsed.netloc not in allowed_hosts:
        raise HTTPException(status_code=400, detail="허용되지 않은 이미지 URL입니다.")
    try:
        req = Request(url, headers={"User-Agent": "PolicyMentoring/1.0"})
        with urlopen(req, timeout=15) as resp:
            content = resp.read()
            media_type = resp.headers.get_content_type() or "image/jpeg"
            return Response(content=content, media_type=media_type)
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(status_code=502, detail="이미지를 불러오지 못했습니다.")


def _truncate_complete(text: str, limit: int) -> str:
    """limit 이내로 자르되 문장이 잘리면 '…'을 붙인다."""
    text = _clean_pledge_share_title(text)
    if len(text) <= limit:
        return text
    # 마침표·종결어미 등으로 끝나면 거기서 자름
    cut = text[:limit]
    for end in ("다.", "요.", "음.", "함.", "임.", "됨."):
        idx = cut.rfind(end)
        if idx > 0:
            return cut[: idx + len(end)]
    # 공백 기준으로 단어가 끊기지 않게
    space = cut.rfind(" ")
    if space > limit // 2:
        return cut[:space] + "…"
    return cut.rstrip() + "…"


_share_summary_cache: dict[str, PledgeShareSummaryResponse] = {}

@app.post("/api/pledge/share-summary", response_model=PledgeShareSummaryResponse, tags=["candidates"])
def create_pledge_share_summary(body: PledgeShareSummaryRequest):
    """공약 카드뉴스용 짧은 요약을 규칙 기반으로 생성한다."""
    import hashlib
    cache_key = hashlib.md5(f'{body.pledge_title or ""}|{body.pledge_content or ""}'.encode("utf-8")).hexdigest()
    if cache_key in _share_summary_cache:
        return _share_summary_cache[cache_key]

    result = _build_rule_based_pledge_share_summary(body.pledge_title, body.pledge_content)
    _share_summary_cache[cache_key] = result
    return result


def _public_site_origin(request: Request) -> str:
    explicit = (os.getenv("PUBLIC_SITE_URL") or "").strip().rstrip("/")
    if explicit:
        return explicit
    return str(request.base_url).rstrip("/")


def _fetch_public_pledge_share_payload(pledge_id: int):
    from backend.database import get_connection

    conn = get_connection()
    try:
        row = conn.execute(
            """
            SELECT cp.id AS pledge_id, cp.title, cp.content, cp.total_score, cp.created_at,
                   cp.share_summary_title, cp.share_summary_headline, cp.share_summary_bullets,
                   cp.share_summary_version, cp.share_summary_source_hash,
                   c.id AS candidate_id, c.name AS candidate_name, c.election_type, c.election_level,
                   c.region_code, COALESCE(u.region_name, rc.region_name, c.region_code) AS region_name,
                   COALESCE(u.district_name, c.district_name) AS district_name,
                   pa.status_note AS applicant_status_note,
                   u.applicant_match_id AS applicant_match_id
            FROM candidate_pledges cp
            JOIN candidates c ON c.id = cp.candidate_id
            LEFT JOIN users u ON u.id = c.user_id
            LEFT JOIN party_applicants pa ON pa.id = u.applicant_match_id
            LEFT JOIN region_codes rc ON rc.region_code = c.region_code
            WHERE cp.id = ?
              AND cp.approval_status = 'APPROVED'
              AND (
                  u.applicant_match_id IS NOT NULL
                  OR TRIM(COALESCE(pa.status_note, '')) = ''
                  OR EXISTS (
                      SELECT 1
                      FROM party_applicants pa_public
                      WHERE TRIM(COALESCE(pa_public.status_note, '')) = '공천 확정'
                        AND (
                            (
                                lower(trim(COALESCE(pa_public.email, ''))) <> ''
                                AND lower(trim(COALESCE(pa_public.email, ''))) = lower(trim(COALESCE(u.email, '')))
                                AND lower(replace(trim(COALESCE(pa_public.name, '')), ' ', '')) = lower(replace(trim(COALESCE(u.name, '')), ' ', ''))
                            )
                            OR (
                                replace(replace(replace(replace(trim(COALESCE(pa_public.phone, '')), '-', ''), ' ', ''), '(', ''), ')', '') <> ''
                                AND replace(replace(replace(replace(trim(COALESCE(pa_public.phone, '')), '-', ''), ' ', ''), '(', ''), ')', '') = replace(replace(replace(replace(trim(COALESCE(u.phone, '')), '-', ''), ' ', ''), '(', ''), ')', '')
                                AND lower(replace(trim(COALESCE(pa_public.name, '')), ' ', '')) = lower(replace(trim(COALESCE(u.name, '')), ' ', ''))
                            )
                        )
                  )
              )
              AND c.approval_status IN ('APPROVED', 'MIXED')
            LIMIT 1
            """,
            (pledge_id,),
        ).fetchone()
    finally:
        conn.close()

    if row is None or not _is_public_nomination_status(row["applicant_status_note"], row["applicant_match_id"]):
        return None

    return {
        "pledge_id": int(row["pledge_id"]),
        "title": row["title"] or "",
        "content": row["content"] or "",
        "total_score": row["total_score"],
        "created_at": row["created_at"] or "",
        "candidate_id": int(row["candidate_id"]),
        "candidate_name": row["candidate_name"] or "",
        "election_type": row["election_type"] or "",
        "election_level": row["election_level"] or "",
        "region_code": row["region_code"] or "",
        "region_name": row["region_name"] or "",
        "district_name": row["district_name"] or "",
        "share_summary": _load_stored_share_summary(row),
        "share_summary_version": row["share_summary_version"] or "",
        "share_summary_source_hash": row["share_summary_source_hash"] or "",
    }


@app.get("/share/pledges/{pledge_id}", tags=["candidates"])
def share_pledge_page(pledge_id: int, request: Request):
    payload = _fetch_public_pledge_share_payload(pledge_id)
    if payload is None:
        raise HTTPException(status_code=404, detail="공유할 공약을 찾을 수 없습니다.")

    profile = _fetch_candidate_external_profile(payload["candidate_id"])
    election_label_map = {
        "metro_mayor": "광역단체장",
        "local_mayor": "기초단체장",
        "regional_council": "광역의원",
        "local_council": "기초의원",
    }
    election_label = election_label_map.get(payload["election_type"], payload["election_type"] or "")
    region_label = (
        f'{payload["region_name"]} {payload["district_name"]}'.strip()
        if payload["district_name"]
        else payload["region_name"]
    ).strip()
    summary = _get_or_create_persisted_pledge_share_summary(payload)

    origin = _public_site_origin(request)
    share_url = f"{origin}/share/pledges/{pledge_id}"
    target_url = f"{origin}/map?candidate={payload['candidate_id']}&pledge={pledge_id}"
    og_image = (profile.photo_url if profile and profile.photo_url else f"{origin}/og.png")

    import html as _html

    title_text = f"{payload['candidate_name']} 공약 | {summary.title}"
    description_text = " · ".join(
        part for part in [
            region_label,
            election_label,
            summary.headline or _truncate_complete(payload["content"], 80),
        ] if part
    )
    candidate_meta = " · ".join(part for part in [region_label, election_label] if part)
    bio_text = profile.bio.strip() if profile and profile.bio else ""

    html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{_html.escape(title_text)}</title>
  <meta name="description" content="{_html.escape(description_text)}">
  <meta property="og:type" content="website">
<meta property="og:site_name" content="개혁신당 공약 멘토링">
  <meta property="og:title" content="{_html.escape(title_text)}">
  <meta property="og:description" content="{_html.escape(description_text)}">
  <meta property="og:url" content="{_html.escape(share_url)}">
  <meta property="og:image" content="{_html.escape(og_image)}">
  <meta name="twitter:card" content="summary_large_image">
  <meta name="twitter:title" content="{_html.escape(title_text)}">
  <meta name="twitter:description" content="{_html.escape(description_text)}">
  <meta name="twitter:image" content="{_html.escape(og_image)}">
  <style>
    body {{
      margin: 0;
      min-height: 100vh;
      font-family: Pretendard, Apple SD Gothic Neo, sans-serif;
      background: radial-gradient(circle at top, #153a72 0%, #071b3a 65%, #041227 100%);
      color: #f8fafc;
      display: flex;
      align-items: center;
      justify-content: center;
      padding: 24px;
    }}
    .card {{
      width: min(720px, 100%);
      background: rgba(9, 23, 49, 0.92);
      border: 1px solid rgba(148, 163, 184, 0.18);
      border-radius: 28px;
      padding: 28px;
      box-shadow: 0 28px 60px rgba(2, 6, 23, 0.38);
    }}
    .eyebrow {{
      color: #fbbf24;
      font-weight: 800;
      font-size: 0.95rem;
      margin-bottom: 12px;
    }}
    h1 {{
      margin: 0 0 10px;
      font-size: clamp(1.7rem, 5vw, 2.6rem);
      line-height: 1.12;
    }}
    .meta {{
      color: #bfdbfe;
      font-size: 0.98rem;
      margin-bottom: 18px;
    }}
    .headline {{
      color: #e2e8f0;
      font-size: 1.05rem;
      line-height: 1.7;
      margin-bottom: 16px;
    }}
    .bio {{
      color: #cbd5e1;
      font-size: 0.95rem;
      line-height: 1.7;
      margin-bottom: 18px;
    }}
    .actions {{
      display: flex;
      gap: 10px;
      flex-wrap: wrap;
      margin-top: 20px;
    }}
    .button {{
      display: inline-flex;
      align-items: center;
      justify-content: center;
      min-height: 46px;
      padding: 0 18px;
      border-radius: 999px;
      text-decoration: none;
      font-weight: 800;
      font-size: 0.98rem;
      border: 1px solid rgba(191, 219, 254, 0.22);
      color: #f8fafc;
      background: rgba(30, 64, 175, 0.22);
    }}
    .button.primary {{
      background: linear-gradient(135deg, #f59e0b, #fb7185);
      color: #081120;
      border: 0;
    }}
  </style>
</head>
<body>
  <main class="card">
    <div class="eyebrow">개혁신당 공약 공유</div>
    <h1>{_html.escape(payload["title"])}</h1>
    <div class="meta">{_html.escape(candidate_meta or payload["candidate_name"])}</div>
    <div class="headline">{_html.escape(summary.headline or payload["content"])}</div>
    {f'<div class="bio">{_html.escape(bio_text)}</div>' if bio_text else ''}
    <div class="actions">
      <a class="button primary" href="{_html.escape(target_url)}">공약 보러가기</a>
      {f'<a class="button" href="{_html.escape(profile.support_url)}" target="_blank" rel="noopener noreferrer">후원하기</a>' if profile and profile.support_url else ''}
    </div>
  </main>
</body>
</html>"""
    return Response(content=html, media_type="text/html; charset=utf-8")


def _get_candidate_detail_any_status(candidate_id: int) -> CandidateDetailResponse:
    """승인 상태 무관하게 후보 상세를 반환 (관리자·내부용)."""
    from backend.database import get_connection

    conn = get_connection()
    try:
        row = conn.execute(
            """SELECT c.id, c.name,
                      COALESCE(u.district_name, c.district_name) AS district_name,
                      c.district_code, c.region_code, c.election_type, c.election_level
               FROM candidates c
               LEFT JOIN users u ON u.id = c.user_id
               WHERE c.id = ?""",
            (candidate_id,),
        ).fetchone()
    finally:
        conn.close()

    if row is None:
        raise HTTPException(status_code=404, detail=f"candidate_id={candidate_id} 후보를 찾을 수 없습니다.")

    code = row["region_code"]
    return CandidateDetailResponse(
        candidate_id=int(row["id"]),
        name=row["name"],
        district_name=row["district_name"],
        district_code=_derive_district_code(code, row["district_code"], row["district_name"]),
        region_code=code,
        region_name=_resolve_region_name(code),
        election_type=row["election_type"],
        election_level=row["election_level"],
        external_profile=_fetch_candidate_external_profile(int(row["id"])),
        pledges=_fetch_candidate_pledges(int(row["id"]), limit=None, public_only=True),
    )


def _snapshot_candidate_pledges(
    conn,
    candidate_id: int,
    approval_status: str,
    rejection_reason: Optional[str] = None,
    source_action: str = "REVIEW",
    reviewed_at: Optional[str] = None,
):
    rows = conn.execute(
        """
        SELECT id, title, content, priority, total_score, analysis_result, analyzed_at, created_at, approval_status
        FROM candidate_pledges
        WHERE candidate_id = ?
        ORDER BY priority ASC, id ASC
        """,
        (candidate_id,),
    ).fetchall()
    if not rows:
        return None

    snapshot_group = uuid.uuid4().hex
    reviewed_value = reviewed_at or datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for row in rows:
        # Use each pledge's own approval_status, not the candidate-level one
        pledge_status = row["approval_status"] or approval_status or "PENDING"
        conn.execute(
            """
            INSERT INTO candidate_pledge_review_history (
                candidate_id, snapshot_group, source_action, approval_status, rejection_reason,
                reviewed_at, pledge_id, title, content, priority, total_score, analysis_result,
                analyzed_at, pledge_created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                candidate_id,
                snapshot_group,
                source_action,
                pledge_status.upper(),
                rejection_reason or None,
                reviewed_value,
                row["id"],
                row["title"],
                row["content"],
                row["priority"],
                row["total_score"],
                row["analysis_result"],
                row["analyzed_at"],
                row["created_at"],
            ),
        )
    return snapshot_group


def _fetch_candidate_review_history(conn, candidate_id: int) -> list[dict]:
    rows = conn.execute(
        """
        SELECT snapshot_group, source_action, approval_status, rejection_reason, reviewed_at,
               title, content, priority, total_score, analyzed_at, pledge_created_at
        FROM candidate_pledge_review_history
        WHERE candidate_id = ?
        ORDER BY datetime(reviewed_at) DESC, snapshot_group DESC, priority ASC, id ASC
        """,
        (candidate_id,),
    ).fetchall()
    groups: list[dict] = []
    current_group: Optional[str] = None
    current_item: Optional[dict] = None
    for row in rows:
        group_id = row["snapshot_group"]
        if group_id != current_group:
            current_group = group_id
            current_item = {
                "snapshot_group": group_id,
                "source_action": row["source_action"] or "REVIEW",
                "approval_status": row["approval_status"] or "PENDING",
                "rejection_reason": row["rejection_reason"] or "",
                "reviewed_at": row["reviewed_at"],
                "pledges": [],
            }
            groups.append(current_item)
        current_item["pledges"].append(
            {
                "title": row["title"],
                "content": row["content"],
                "priority": row["priority"],
                "total_score": row["total_score"],
                "analyzed_at": row["analyzed_at"],
                "created_at": row["pledge_created_at"],
            }
        )
    return groups


PLEDGE_TEMPLATES_PATH = ROOT_DIR / "data" / "pledge_templates.json"


@app.get("/api/pledge/templates")
def api_pledge_templates(
    request: Request,
    election_position: str = Query(default="", description="(관리자 전용) 선거유형 오버라이드"),
    region_name: str = Query(default="", description="(관리자 전용) 출마지역(광역) 오버라이드"),
    district_name: str = Query(default="", description="(관리자 전용) 지역구/선거구 오버라이드"),
):
    """
    로그인 사용자의 선거유형·지역에 맞는 공약 초안 템플릿을 반환한다.
    회원가입 시 저장된 election_position, region_name, district_name을 자동 반영.
    관리자는 쿼리 파라미터로 선거유형·지역을 지정해 테스트할 수 있다.
    """
    user = require_user(request)
    full = get_user(user["id"])
    is_admin = user["role"] == ROLE_ADMIN or user["email"] in ADMIN_EMAILS

    ep = (full.get("election_position") or "").strip().lower() if full else ""
    region_name_val = (full.get("region_name") or "").strip() if full else ""
    district_name_val = (full.get("district_name") or "").strip() if full else ""

    if is_admin:
        if (election_position or "").strip():
            ep = election_position.strip().lower()
        if (region_name or "").strip():
            region_name_val = region_name.strip()
        if (district_name or "").strip():
            district_name_val = district_name.strip()

    if not PLEDGE_TEMPLATES_PATH.exists():
        out = {"label": "", "sections": [], "election_position": ep, "region_name": region_name_val, "district_name": district_name_val}
        if is_admin:
            out["template_options"] = []
        return out

    try:
        raw = PLEDGE_TEMPLATES_PATH.read_text(encoding="utf-8")
        data = json.loads(raw)
    except Exception as e:
        logger.warning("pledge_templates.json load failed: %s", e)
        out = {"label": "", "sections": [], "election_position": ep, "region_name": region_name_val, "district_name": district_name_val}
        if is_admin:
            out["template_options"] = []
        return out

    if is_admin:
        template_options = [{"key": k, "label": (v.get("label") or k)} for k, v in data.items() if isinstance(v, dict) and v.get("label") and not v.get("not_available")]
    else:
        template_options = []

    def substitute(s: str) -> str:
        return (s or "").replace("{{region_name}}", region_name_val).replace("{{district_name}}", district_name_val)

    template = data.get(ep) if ep else None
    if not template:
        out = {"label": "", "election_position": ep, "region_name": region_name_val, "district_name": district_name_val}
        if template_options:
            out["template_options"] = template_options
        return out

    if template.get("not_available"):
        out = {
            "label": substitute(template.get("label", "")),
            "not_available": True,
            "message": substitute(template.get("message", "해당 선거유형은 공약 초안 도우미 대상이 아닙니다.")),
            "election_position": ep,
            "region_name": region_name_val,
            "district_name": district_name_val,
        }
        if template_options:
            out["template_options"] = template_options
        return out

    label = substitute(template.get("label", ""))
    structure_raw = template.get("structure") or {}
    structure = {
        "background": substitute(structure_raw.get("background", "")),
        "action": substitute(structure_raw.get("action", "")),
        "effect": substitute(structure_raw.get("effect", "")),
    }
    checklist = [substitute(x) for x in (template.get("checklist") or [])]
    do_dont = [substitute(x) for x in (template.get("do_dont") or [])]
    standard_intro = (data.get("standard_intro") or "").strip()
    out = {
        "label": label,
        "standard_intro": standard_intro,
        "guide_title": substitute(template.get("guide_title", "공약 한 편 잘 쓰기")),
        "guide_intro": substitute(template.get("guide_intro", "")),
        "structure": structure,
        "checklist": checklist,
        "do_dont": do_dont,
        "example": substitute(template.get("example", "")),
        "notice_one_line": substitute(template.get("notice_one_line", "")),
        "election_position": ep,
        "region_name": region_name_val,
        "district_name": district_name_val,
    }
    if template_options:
        out["template_options"] = template_options
    return out


class PledgeVerifyRequest(BaseModel):
    text: str = Field(..., description="검증할 출마자 공약 텍스트")
    top_k_platform: int = Field(default=6, description="정강정책 검색 개수")
    top_k_pledge: int = Field(default=6, description="공약 검색 개수")
    top_k_regional: int = Field(default=8, description="지역별 공약 검색 개수")
    phase: str = Field(default="full", description="quick=1차 빠른 판정(결과 3개, 속도 우선), full=2차 상세 근거·상충 분석(6개)")
    judge: bool = Field(default=False, description="true=strict judge 모드 (evidence, specificity cap, QUERY/VERIFY)")


@app.post("/api/pledge/verify")
def verify_pledge(body: PledgeVerifyRequest, request: Request):
    """
    벡터 검색 기반 공약 검증 리포트를 생성한다. (승인 사용자 전용)
    """
    import time
    t0 = time.perf_counter()
    logger.info("[verify] started")
    _ensure_startup()  # 지연 초기화

    user = require_approved(request)
    ip = _client_ip(request)
    ok, msg = check_rate_limit_ip(ip)
    if not ok:
        raise HTTPException(status_code=429, detail=msg)
    ok, msg = check_rate_limit_user(user["id"])
    if not ok:
        raise HTTPException(status_code=429, detail=msg)

    global _indexes, _vector_store_id, _regional_vector_store_id
    if USE_OPENAI_VECTOR_STORE and not _vector_store_id:
        raise HTTPException(status_code=503, detail="Vector Store가 준비되지 않았습니다.")
    if not USE_OPENAI_VECTOR_STORE and (not _indexes or not _indexes.get("pledge")):
        raise HTTPException(status_code=503, detail="인덱스가 준비되지 않았습니다.")

    from backend.analysis_service import run_verify_analysis
    options = {
        "top_k_platform": body.top_k_platform,
        "top_k_pledge": body.top_k_pledge,
        "top_k_regional": body.top_k_regional,
        "phase": body.phase or "full",
        "judge": body.judge,
    }
    result, status_code, from_cache = run_verify_analysis(
        user["id"],
        body.text or "",
        ip,
        options,
        _vector_store_id if USE_OPENAI_VECTOR_STORE else None,
        _regional_vector_store_id if USE_OPENAI_VECTOR_STORE else None,
        _indexes if not USE_OPENAI_VECTOR_STORE else None,
    )
    if status_code >= 400:
        raise HTTPException(status_code=status_code, detail=result.get("detail", result) if isinstance(result, dict) else result)
    # 기록에는 당 부합 점검(텍스트)만 저장. verify(JSON)는 저장하지 않음.
    logger.info("[verify] completed in %.1fs", time.perf_counter() - t0)
    return result


@app.get("/api/history")
def api_history(request: Request, limit: int = Query(default=20, ge=1, le=100)):
    user = require_approved(request)
    from backend.history import list_history

    return {"items": list_history(user["id"], limit=limit)}


@app.get("/api/history/{history_id}")
def api_history_item(history_id: int, request: Request):
    user = require_approved(request)
    from backend.history import get_history_item

    item = get_history_item(user["id"], history_id)
    if not item:
        raise HTTPException(status_code=404, detail="not found")
    return item


@app.post("/api/history/{history_id}/delete")
def api_history_delete(history_id: int, request: Request):
    user = require_approved(request)
    from backend.history import delete_history_item

    ok = delete_history_item(user["id"], history_id)
    if not ok:
        raise HTTPException(status_code=404, detail="not found")
    return {"ok": True}


@app.post("/api/history/clear")
def api_history_clear(request: Request):
    user = require_approved(request)
    from backend.history import clear_history

    deleted = clear_history(user["id"])
    return {"ok": True, "deleted": deleted}


# ── 사용자 본인 공약 등록/관리 ──────────────────────────────

ELECTION_POSITION_TO_TYPE = {
    "metro_mayor": "metro_mayor",
    "regional_council": "regional_council",
    "local_mayor": "local_mayor",
    "local_council": "local_council",
    "party_official": "party_official",
}

ELECTION_POSITION_TO_LEVEL = {
    "metro_mayor": "metro",
    "regional_council": "metro",
    "local_mayor": "local",
    "local_council": "local",
    "party_official": "none",
}


@app.api_route("/my-pledges", methods=["GET", "HEAD"])
def my_pledges_page(request: Request):
    """사용자 본인 공약 등록/관리 페이지."""
    user = get_current_user(request)
    if not user:
        return _login_redirect(request.url.path)
    if (
        user["status"] != STATUS_APPROVED
        and user["email"] not in ADMIN_EMAILS
        and user["role"] != ROLE_ADMIN
    ):
        return RedirectResponse(url="/pending", status_code=302)
    res = _serve_html("my-pledges.html")
    if res is not None:
        return res
    raise HTTPException(status_code=404, detail="my-pledges.html not found")


@app.api_route("/my-pledges-status", methods=["GET", "HEAD"])
def my_pledges_status_page(request: Request):
    """사용자 본인 공약 상태 확인 페이지."""
    user = get_current_user(request)
    if not user:
        return _login_redirect(request.url.path)
    if (
        user["status"] != STATUS_APPROVED
        and user["email"] not in ADMIN_EMAILS
        and user["role"] != ROLE_ADMIN
    ):
        return RedirectResponse(url="/pending", status_code=302)
    res = _serve_html("my-pledges-status.html")
    if res is not None:
        return res
    raise HTTPException(status_code=404, detail="my-pledges-status.html not found")


@app.get("/api/my/candidate", tags=["my-candidate"])
def api_my_candidate_get(request: Request):
    """로그인 사용자의 후보 프로필 + 공약 목록 반환. 미등록이면 user 정보만 반환."""
    _ensure_db_ready()
    user = require_approved(request)
    uid = user["id"]

    election_position = user.get("election_position") or ""
    region_code = user.get("region_code") or ""
    region_name = user.get("region_name") or REGION_NAME_MAP.get(region_code, "")
    district_code = user.get("district_code") or ""
    district_name = user.get("district_name") or ""
    election_type = ELECTION_POSITION_TO_TYPE.get(election_position, election_position)
    election_level = ELECTION_POSITION_TO_LEVEL.get(election_position, "regional")

    from backend.database import get_connection
    conn = get_connection()
    try:
        row = conn.execute(
            "SELECT id, name, district_name, district_code, region_code, election_type, election_level, approval_status, rejection_reason, created_at, updated_at FROM candidates WHERE user_id = ? LIMIT 1",
            (uid,),
        ).fetchone()

        if row is None:
            return {
                "candidate": None,
                "pledges": [],
                "user_info": {
                    "name": user.get("name") or user.get("email", ""),
                    "election_position": election_position,
                    "election_type": election_type,
                    "election_level": election_level,
                    "region_code": region_code,
                    "region_name": region_name,
                    "district_code": district_code,
                    "district_name": district_name,
                },
            }

        candidate_id = int(row["id"])
        pledges = conn.execute(
            """
            SELECT id, title, content, priority, total_score, analysis_result, analyzed_at, created_at,
                   approval_status, rejection_reason
            FROM candidate_pledges
            WHERE candidate_id = ?
            ORDER BY priority ASC, id ASC
            """,
            (candidate_id,),
        ).fetchall()
        review_history = _fetch_candidate_review_history(conn, candidate_id)
        if not review_history and (row["approval_status"] or "PENDING").upper() == "REJECTED" and pledges:
            review_history = [
                {
                    "snapshot_group": f"synthetic-{candidate_id}",
                    "source_action": "REJECT",
                    "approval_status": "REJECTED",
                    "rejection_reason": row["rejection_reason"] or "",
                    "reviewed_at": row["updated_at"] if "updated_at" in row.keys() else row["created_at"],
                    "pledges": [
                        {
                            "title": p["title"],
                            "content": p["content"],
                            "priority": p["priority"],
                            "total_score": p["total_score"],
                            "analyzed_at": p["analyzed_at"],
                            "created_at": p["created_at"],
                        }
                        for p in pledges
                    ],
                }
            ]

        return {
            "candidate": {
                "candidate_id": candidate_id,
                "name": row["name"],
                "district_name": row["district_name"],
                "district_code": row["district_code"],
                "region_code": row["region_code"],
                "region_name": region_name,
                "election_type": row["election_type"],
                "election_level": row["election_level"],
                "approval_status": row["approval_status"] or "PENDING",
                "rejection_reason": row["rejection_reason"] or "",
            },
            "pledges": [
                {"id": p["id"], "title": p["title"], "content": p["content"], "priority": p["priority"],
                 "total_score": p["total_score"], "analysis_result": p["analysis_result"], "analyzed_at": p["analyzed_at"], "created_at": p["created_at"],
                 "approval_status": p["approval_status"] or "PENDING", "rejection_reason": p["rejection_reason"] or ""}
                for p in pledges
            ],
            "review_history": review_history,
            "user_info": {
                "name": user.get("name") or user.get("email", ""),
                "election_position": election_position,
                "election_type": election_type,
                "election_level": election_level,
                "region_code": region_code,
                "region_name": region_name,
                "district_code": district_code,
                "district_name": district_name,
            },
        }
    finally:
        conn.close()


class MyPledgeInput(BaseModel):
    title: str = Field(..., min_length=1, max_length=100, description="공약 제목")
    content: Optional[str] = Field(default=None, max_length=50000, description="공약 세부내용")
    priority: int = Field(default=100, ge=1, le=9999, description="정렬 우선순위")
    pledge_id: Optional[int] = Field(default=None, description="기존 공약 ID (삭제-전용 판별용)")
    imported_score: Optional[float] = Field(default=None, description="불러오기한 점수 (analysis_history에서)")
    imported_result: Optional[str] = Field(default=None, max_length=80000, description="불러오기한 분석 결과")
    imported_analyzed_at: Optional[str] = Field(default=None, description="불러오기한 분석 일시")


class MyPledgesBody(BaseModel):
    pledges: list[MyPledgeInput] = Field(..., min_length=1, max_length=30, description="공약 목록 (1~30개)")


@app.post("/api/my/candidate", tags=["my-candidate"])
def api_my_candidate_save(body: MyPledgesBody, request: Request):
    """
    사용자 본인의 후보 등록 + 공약 저장.
    candidates 행이 없으면 INSERT, 있으면 공약만 교체(UPSERT).
    region_code/election_type 등은 회원가입 시 입력한 정보에서 자동 결정된다.
    """
    _ensure_db_ready()
    user = require_approved(request)
    uid = user["id"]

    election_position = (user.get("election_position") or "").strip()
    if not election_position:
        raise HTTPException(status_code=400, detail="회원가입 시 출마 유형을 선택하지 않아 공약을 등록할 수 없습니다.")
    region_code = (user.get("region_code") or "").strip()
    if not region_code:
        raise HTTPException(status_code=400, detail="회원가입 시 지역을 선택하지 않아 공약을 등록할 수 없습니다.")

    # 분석 안 된 공약 차단: 모든 공약에 imported_score 필수
    unanalyzed = [p for p in body.pledges if p.imported_score is None]
    if unanalyzed:
        raise HTTPException(
            status_code=400,
            detail=f"분석되지 않은 공약이 {len(unanalyzed)}개 있습니다. 모든 공약은 분석 기록에서 불러와야 합니다.",
        )

    election_type = ELECTION_POSITION_TO_TYPE.get(election_position, election_position)
    election_level = ELECTION_POSITION_TO_LEVEL.get(election_position, "regional")
    region_name = user.get("region_name") or REGION_NAME_MAP.get(region_code, "")
    district_code_raw = (user.get("district_code") or "").strip()
    district_name = (user.get("district_name") or "").strip()
    resolved_district_code = district_code_raw or _derive_district_code(region_code, None, district_name)
    candidate_name = (user.get("name") or user.get("email", "")).strip()

    from backend.database import get_connection
    conn = get_connection()
    try:
        row = conn.execute(
            "SELECT id, approval_status, rejection_reason FROM candidates WHERE user_id = ? LIMIT 1",
            (uid,),
        ).fetchone()

        if row is None:
            cur = conn.execute(
                """
                INSERT INTO candidates (name, district_name, district_code, region_code, election_type, election_level, user_id, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, datetime('now'))
                """,
                (candidate_name, district_name or None, resolved_district_code, region_code, election_type, election_level, uid),
            )
            candidate_id = int(cur.lastrowid)
            if resolved_district_code and district_name:
                conn.execute(
                    """
                    INSERT INTO region_codes (region_code, region_name, aliases_json, updated_at)
                    VALUES (?, ?, '[]', datetime('now'))
                    ON CONFLICT(region_code) DO NOTHING
                    """,
                    (region_code, region_name or REGION_NAME_MAP.get(region_code, region_code)),
                )
                conn.execute(
                    """
                    INSERT INTO district_codes (district_code, district_name, region_code, election_type, aliases_json, updated_at)
                    VALUES (?, ?, ?, ?, '[]', datetime('now'))
                    ON CONFLICT(district_code) DO UPDATE SET
                        district_name = excluded.district_name,
                        region_code = excluded.region_code,
                        election_type = excluded.election_type,
                        updated_at = datetime('now')
                    """,
                    (resolved_district_code, district_name, region_code, election_type),
                )
            for idx, p in enumerate(body.pledges):
                # 분석 결과 없이 점수만 있는 경우 점수 무시 (불러오기 통해서만 점수 허용)
                imported_result = (p.imported_result or "").strip() or None
                valid_score = p.imported_score if (p.imported_score is not None and imported_result) else None
                valid_analyzed_at = p.imported_analyzed_at if valid_score is not None else None
                conn.execute(
                    """INSERT INTO candidate_pledges
                       (candidate_id, title, content, priority, total_score, analysis_result, analyzed_at)
                       VALUES (?, ?, ?, ?, ?, ?, ?)""",
                    (
                        candidate_id,
                        p.title.strip(),
                        (p.content or "").strip() or None,
                        p.priority if p.priority else (idx + 1),
                        valid_score,
                        imported_result,
                        valid_analyzed_at,
                    ),
                )
            _recalculate_candidate_approval(conn, candidate_id)
            conn.commit()
            _try_sync_candidate_external_profile(candidate_name)
        else:
            candidate_id = int(row["id"])
            previous_status = (row["approval_status"] or "PENDING").upper()
            previous_reason = row["rejection_reason"] or ""
            existing_rows = conn.execute(
                """
                SELECT id, title, content, priority, total_score, analysis_result, analyzed_at,
                       approval_status, rejection_reason
                FROM candidate_pledges
                WHERE candidate_id = ?
                ORDER BY priority ASC, id ASC
                """,
                (candidate_id,),
            ).fetchall()
            existing_by_id = {int(r["id"]): r for r in existing_rows}
            existing_ids = set(existing_by_id.keys())
            submitted_existing_ids = {int(p.pledge_id) for p in body.pledges if p.pledge_id is not None and int(p.pledge_id) in existing_ids}
            reorder_all = bool(existing_ids) and existing_ids.issubset(submitted_existing_ids)
            next_priority = max((int(r["priority"] or 0) for r in existing_rows), default=0)
            has_changes = False

            for idx, p in enumerate(body.pledges):
                title = p.title.strip()
                content = (p.content or "").strip() or None
                imported_result = (p.imported_result or "").strip() or None
                valid_score = p.imported_score if (p.imported_score is not None and imported_result) else None
                valid_analyzed_at = p.imported_analyzed_at if valid_score is not None else None
                existing_id = int(p.pledge_id) if p.pledge_id is not None and int(p.pledge_id) in existing_ids else None

                if existing_id is None:
                    next_priority += 1
                    conn.execute(
                        """INSERT INTO candidate_pledges
                           (candidate_id, title, content, priority, total_score, analysis_result, analyzed_at)
                           VALUES (?, ?, ?, ?, ?, ?, ?)""",
                        (
                            candidate_id,
                            title,
                            content,
                            (idx + 1) if reorder_all else next_priority,
                            valid_score,
                            imported_result,
                            valid_analyzed_at,
                        ),
                    )
                    has_changes = True
                    continue

                prev = existing_by_id[existing_id]
                prev_title = prev["title"] or ""
                prev_content = (prev["content"] or "").strip() or None
                prev_priority = int(prev["priority"] or 0)
                prev_score = prev["total_score"]
                prev_result = (prev["analysis_result"] or "").strip() or None
                prev_analyzed_at = prev["analyzed_at"] or None
                prev_status = (prev["approval_status"] or "PENDING").upper()
                prev_rejection_reason = prev["rejection_reason"] or None

                new_priority = (idx + 1) if reorder_all else (p.priority if p.priority else prev_priority)
                content_changed = (prev_title != title) or (prev_content != content)
                analysis_changed = (
                    imported_result is not None and (
                        prev_score != valid_score
                        or prev_result != imported_result
                        or prev_analyzed_at != valid_analyzed_at
                    )
                )
                needs_reapproval = content_changed or analysis_changed
                row_changed = needs_reapproval or (prev_priority != new_priority)

                if not row_changed:
                    continue

                conn.execute(
                    """
                    UPDATE candidate_pledges
                    SET title = ?, content = ?, priority = ?, total_score = ?, analysis_result = ?, analyzed_at = ?,
                        approval_status = ?, rejection_reason = ?
                    WHERE id = ?
                    """,
                    (
                        title,
                        content,
                        new_priority,
                        valid_score if analysis_changed else prev_score,
                        imported_result if analysis_changed else prev_result,
                        valid_analyzed_at if analysis_changed else prev_analyzed_at,
                        "PENDING" if needs_reapproval else prev_status,
                        None if needs_reapproval else prev_rejection_reason,
                        existing_id,
                    ),
                )
                has_changes = True

            conn.execute(
                "UPDATE candidates SET name = ?, rejection_reason = NULL, updated_at = datetime('now') WHERE id = ?",
                (candidate_name, candidate_id),
            )
            if existing_rows and previous_status in {"APPROVED", "REJECTED", "MIXED"} and has_changes:
                _snapshot_candidate_pledges(
                    conn,
                    candidate_id,
                    approval_status=previous_status,
                    rejection_reason=previous_reason,
                    source_action="RESUBMIT",
                )
            if has_changes:
                _recalculate_candidate_approval(conn, candidate_id)
            conn.commit()
            _try_sync_candidate_external_profile(candidate_name)
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        logger.exception("내 공약 저장 실패: %s", e)
        raise HTTPException(status_code=500, detail=f"저장 중 오류: {str(e)[:200]}")
    finally:
        conn.close()

    # 관리자에게 공약 등록 알림 메일 발송
    try:
        from backend.email_sender import send_pledge_registration_notification
        pledges_summary_lines = []
        for i, p in enumerate(body.pledges):
            score_str = f" ({p.imported_score}점)" if p.imported_score is not None else ""
            pledges_summary_lines.append(f"{i+1}. {p.title.strip()}{score_str}")
        send_pledge_registration_notification(
            user_email=user.get("email", ""),
            name=user.get("name", ""),
            candidate_name=candidate_name,
            election_position=election_position,
            region_name=region_name,
            district_name=district_name,
            pledge_count=len(body.pledges),
            pledges_summary="\n".join(pledges_summary_lines),
        )
    except Exception as e:
        logger.warning("공약 등록 알림 메일 발송 실패 (무시): %s", e)

    return {"ok": True, "candidate_id": candidate_id, "pledge_count": len(body.pledges)}


class NecFormRequest(BaseModel):
    pledge_text: str = Field(..., description="공약 원문 텍스트")
    candidate_name: str = Field(default="", description="후보자명")
    election_type: str = Field(default="", description="선거유형 (예: metro_mayor)")
    region_name: str = Field(default="", description="시·도명")
    district_name: str = Field(default="", description="선거구명")
    result_text: str = Field(default="", description="AI 분석 결과 텍스트 (generate 모드 컨텍스트)")
    mode: str = Field(default="extract", description="'extract' | 'generate'")


@app.post("/api/documents/nec-form", tags=["documents"])
def api_generate_nec_form(body: NecFormRequest, request: Request):
    """공약 텍스트를 선관위 제출용 선거공약서 구조(JSON)로 변환."""
    _ = require_user(request)

    pledge_text = (body.pledge_text or "").strip()
    if not pledge_text:
        raise HTTPException(status_code=400, detail="공약 내용이 비어 있습니다.")

    mode = (body.mode or "extract").strip().lower()
    result_text = (body.result_text or "").strip()

    from openai import OpenAI
    from backend.config import OPENAI_API_KEY, CHAT_MODEL

    json_schema = (
        "[\n"
        "  {\n"
        "    \"순위\": 1,\n"
        "    \"제목\": \"공약 제목\",\n"
        "    \"내용\": [\"핵심 공약 내용 또는 세부 추진 항목 1\", \"세부 항목 2\"],\n"
        "    \"목표\": [\"목표 내용\"],\n"
        "    \"이행방법\": [\"방법 1\", \"방법 2\"],\n"
        "    \"이행기간\": [\"예: 취임 후 1년 이내\"],\n"
        "    \"재원조달방안\": [\"시비, 국비보조 등\"]\n"
        "  }\n"
        "]"
    )

    if mode == "generate" and result_text:
        system_prompt = (
            "당신은 선거공약서 작성 전문가입니다. "
            "AI 분석이 제안한 수정·보완 사항을 실제로 공약에 반영하여, "
            "공직선거법 제66조에 따른 완성된 선거공약서를 작성하세요. "
            "반드시 JSON 배열만 출력하고 다른 텍스트는 포함하지 마세요."
        )
        user_prompt = (
            "아래 공약 원문과 AI 분석 결과를 읽으세요.\n"
            "분석 결과의 '수정·보완 제안' 내용을 실제로 공약에 반영하여, "
            "선관위 제출용 선거공약서 최종본을 완성된 문장으로 작성해주세요.\n\n"
            "작성 규칙:\n"
            "- 각 공약을 하나의 JSON 객체로 표현하고, 아래 형식의 JSON 배열로만 응답하세요.\n"
            "- 보완 제안이 있는 항목은 반드시 반영하여 내용을 강화하세요.\n"
            "- 모든 항목(목표, 이행방법, 이행기간, 재원조달방안)을 완성된 문장으로 작성하세요.\n"
            "- 이행기간, 재원조달방안이 원문에 없으면 공약 성격에 맞게 합리적으로 작성하세요.\n"
            "- 기호는 빈 문자열로 두세요.\n\n"
            f"{json_schema}\n\n"
            f"공약 원문:\n{pledge_text}\n\n"
            f"AI 분석 결과 (수정·보완 제안 포함):\n{result_text}"
        )
    else:
        system_prompt = (
            "당신은 선거공약서 작성 전문가입니다. "
            "입력된 공약 텍스트를 분석하여, 각 공약 항목을 공직선거법 제66조에 따른 선거공약서 양식에 맞게 구조화하세요. "
            "반드시 JSON 배열만 출력하고 다른 텍스트는 포함하지 마세요."
        )
        user_prompt = (
            "다음 공약 텍스트를 선관위 제출용 선거공약서 형식으로 구조화해주세요.\n"
            "각 공약 항목을 분리하고, 아래 형식의 JSON 배열로만 응답하세요.\n\n"
            "각 항목(목표, 이행방법, 이행기간, 재원조달방안)은 공약 내용에 자연스럽게 존재하는 것만 채우세요.\n"
            "명시되지 않은 항목은 빈 배열([])로 두세요. 억지로 추론하거나 없는 내용을 만들어내지 마세요.\n"
            "세부 항목들은 배열로 표현하세요.\n\n"
            f"{json_schema}\n\n"
            f"공약 텍스트:\n{pledge_text}"
        )

    try:
        client = OpenAI(api_key=OPENAI_API_KEY)
        response = client.chat.completions.create(
            model=CHAT_MODEL,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
            temperature=0.2,
        )
        raw = response.choices[0].message.content or ""
        # JSON 블록 추출 (```json ... ``` 감싸인 경우 대비)
        raw = raw.strip()
        if raw.startswith("```"):
            raw = re.sub(r"^```[a-z]*\n?", "", raw)
            raw = re.sub(r"\n?```$", "", raw)
            raw = raw.strip()
        items = json.loads(raw)
        if not isinstance(items, list):
            items = [items]
    except json.JSONDecodeError as e:
        logger.warning("NEC form JSON 파싱 실패: %s | raw=%s", e, raw[:200])
        raise HTTPException(status_code=502, detail="GPT 응답 파싱 실패. 다시 시도해 주세요.")
    except Exception as e:
        logger.exception("NEC form 생성 오류: %s", e)
        raise HTTPException(status_code=500, detail="선거공약서 생성 중 오류가 발생했습니다.")

    return {"items": items}


@app.delete("/api/my/candidate", tags=["my-candidate"])
def api_my_candidate_delete(request: Request):
    """사용자 본인의 후보 프로필 + 공약 전체 삭제."""
    _ensure_db_ready()
    user = require_approved(request)
    uid = user["id"]

    from backend.database import get_connection
    conn = get_connection()
    try:
        row = conn.execute("SELECT id FROM candidates WHERE user_id = ? LIMIT 1", (uid,)).fetchone()
        if row is None:
            raise HTTPException(status_code=404, detail="등록된 후보 정보가 없습니다.")
        candidate_id = int(row["id"])
        conn.execute("DELETE FROM candidate_pledges WHERE candidate_id = ?", (candidate_id,))
        conn.execute("DELETE FROM candidates WHERE id = ?", (candidate_id,))
        conn.commit()
    except HTTPException:
        conn.rollback()
        raise
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    return {"ok": True}


@app.delete("/api/my/pledges/{pledge_id}", tags=["my-candidate"])
def api_my_pledge_delete(pledge_id: int, request: Request):
    """사용자 본인의 개별 공약 삭제."""
    _ensure_db_ready()
    user = require_approved(request)
    uid = user["id"]

    from backend.database import get_connection
    conn = get_connection()
    try:
        row = conn.execute(
            """
            SELECT cp.id, cp.candidate_id, c.approval_status, c.rejection_reason
            FROM candidate_pledges cp
            JOIN candidates c ON c.id = cp.candidate_id
            WHERE cp.id = ? AND c.user_id = ?
            """,
            (pledge_id, uid),
        ).fetchone()
        if row is None:
            raise HTTPException(status_code=404, detail="해당 공약을 찾을 수 없습니다.")

        candidate_id = int(row["candidate_id"])
        previous_status = (row["approval_status"] or "PENDING").upper()
        previous_reason = row["rejection_reason"] or ""
        if previous_status in {"APPROVED", "REJECTED", "MIXED"}:
            _snapshot_candidate_pledges(
                conn,
                candidate_id,
                approval_status=previous_status,
                rejection_reason=previous_reason,
                source_action="DELETE_PLEDGE",
            )

        conn.execute("DELETE FROM candidate_pledges WHERE id = ?", (pledge_id,))
        remaining = conn.execute(
            "SELECT COUNT(*) AS n FROM candidate_pledges WHERE candidate_id = ?",
            (candidate_id,),
        ).fetchone()
        if int(remaining["n"] or 0) == 0:
            conn.execute(
                "UPDATE candidates SET approval_status = 'PENDING', rejection_reason = NULL, updated_at = datetime('now') WHERE id = ?",
                (candidate_id,),
            )
            candidate_status = "PENDING"
        else:
            candidate_status = _recalculate_candidate_approval(conn, candidate_id)
        conn.commit()
    except HTTPException:
        conn.rollback()
        raise
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    return {"ok": True, "candidate_status": candidate_status}


# ─────────────────────── 개별 공약 분석 ───────────────────────

@app.post("/api/my/pledges/{pledge_id}/analyze", tags=["my-candidate"])
def api_analyze_pledge(pledge_id: int, request: Request):
    """등록된 개별 공약을 AI로 분석하여 점수를 매긴다."""
    import time
    _ensure_startup()
    user = require_approved(request)
    uid = user["id"]
    ip = _client_ip(request)

    ok, msg = check_rate_limit_ip(ip)
    if not ok:
        raise HTTPException(status_code=429, detail=msg)
    ok, msg = check_rate_limit_user(uid)
    if not ok:
        raise HTTPException(status_code=429, detail=msg)

    from backend.database import get_connection

    conn = get_connection()
    try:
        row = conn.execute(
            """
            SELECT cp.id, cp.title, cp.content, cp.candidate_id, c.user_id
            FROM candidate_pledges cp
            JOIN candidates c ON c.id = cp.candidate_id
            WHERE cp.id = ? AND c.user_id = ?
            """,
            (pledge_id, uid),
        ).fetchone()
        if row is None:
            raise HTTPException(status_code=404, detail="해당 공약을 찾을 수 없습니다.")
    finally:
        conn.close()

    pledge_title = row["title"] or ""
    pledge_content = row["content"] or ""
    text = f"{pledge_title}\n\n{pledge_content}".strip()
    if not text:
        raise HTTPException(status_code=400, detail="공약 내용이 비어 있습니다.")

    from backend.analysis_service import run_check_analysis
    global _indexes, _vector_store_id, _regional_vector_store_id, _winners2022_vector_store_id
    vs_id = _vector_store_id if USE_OPENAI_VECTOR_STORE else None
    regional_vs_id = _regional_vector_store_id if USE_OPENAI_VECTOR_STORE else None
    winners2022_vs_id = _winners2022_vector_store_id

    t0 = time.perf_counter()
    result, status_code, from_cache = run_check_analysis(
        uid, text, ip, vs_id, regional_vs_id, winners2022_vs_id,
        _indexes if not USE_OPENAI_VECTOR_STORE else None,
    )
    elapsed = time.perf_counter() - t0
    logger.info("[analyze_pledge] pledge_id=%d completed in %.1fs", pledge_id, elapsed)

    if status_code >= 400:
        raise HTTPException(status_code=status_code, detail=result)

    from backend.score_parser import parse_total_score
    score = parse_total_score(str(result))

    # candidate_pledges 업데이트
    conn2 = get_connection()
    try:
        result_text = str(result) if not isinstance(result, str) else result
        conn2.execute(
            "UPDATE candidate_pledges SET total_score = ?, analysis_result = ?, analyzed_at = datetime('now') WHERE id = ?",
            (score, result_text[:60000], pledge_id),
        )
        conn2.commit()
    finally:
        conn2.close()

    # analysis_history에도 저장
    try:
        from backend.history import add_history
        add_history(user_id=uid, kind="check", input_text=text, result=result,
                    status_code=status_code, from_cache=from_cache, options={"source": "pledge_analyze", "pledge_id": pledge_id})
    except Exception:
        pass

    return {"score": score, "result_text": str(result), "pledge_id": pledge_id}


# ─────────────────────── 리더보드 ───────────────────────

def _week_label(monday_date):
    """월요일 날짜로부터 'N월 M째주' 라벨 생성."""
    import datetime as _dt
    m = monday_date.month
    first_of_month = monday_date.replace(day=1)
    first_monday = first_of_month + _dt.timedelta(days=(7 - first_of_month.weekday()) % 7)
    if first_monday > monday_date:
        wn = 1
    else:
        wn = ((monday_date - first_monday).days // 7) + 1
        if first_of_month.weekday() != 0:
            wn += 1
    return f"{m}월 {wn}째주"


@app.get("/api/leaderboard", tags=["leaderboard"])
def api_leaderboard(
    region_code: Optional[str] = Query(default=None),
    election_type: Optional[str] = Query(default=None),
    week_start: Optional[str] = Query(default=None, description="조회할 주의 월요일 ISO날짜 (예: 2026-02-16). 미지정 시 이번 주"),
):
    """공약 평균 점수 기준 후보자 랭킹 (공개 API). 주간 챔피언 포함."""
    _ensure_db_ready()
    import datetime as _dt

    from backend.database import get_connection
    conn = get_connection()
    try:
        approved_review_subquery = """
            SELECT pledge_id, MAX(reviewed_at) AS approved_reviewed_at
            FROM candidate_pledge_review_history
            WHERE approval_status = 'APPROVED'
              AND pledge_id IS NOT NULL
            GROUP BY pledge_id
        """
        today = _dt.date.today()
        current_monday = today - _dt.timedelta(days=today.weekday())
        reset_monday = _leaderboard_reset_monday()

        # 조회 대상 주 결정
        if week_start:
            try:
                req_date = _dt.date.fromisoformat(week_start)
                target_monday = req_date - _dt.timedelta(days=req_date.weekday())
            except ValueError:
                target_monday = current_monday
        else:
            target_monday = current_monday

        if target_monday < reset_monday:
            target_monday = reset_monday

        # 미래 주는 허용하지 않음
        if target_monday > current_monday and today >= reset_monday:
            target_monday = current_monday

        target_sunday = target_monday + _dt.timedelta(days=6)
        is_current_week = (target_monday == current_monday)
        ranking_end = today if is_current_week else target_sunday

        if today < reset_monday:
            return {
                "week_label": _week_label(reset_monday) + " 랭킹",
                "week_start": reset_monday.isoformat(),
                "is_current_week": True,
                "prev_week": None,
                "next_week": None,
                "champions": [],
                "ranking": [],
                "total_count": 0,
            }

        # ── 주간 챔피언 lazy snapshot (이번 주 조회 시만) ──
        if is_current_week:
            last_monday = current_monday - _dt.timedelta(days=7)
            last_monday_str = last_monday.isoformat()
            existing = conn.execute(
                "SELECT id FROM weekly_champions WHERE week_start = ?", (last_monday_str,)
            ).fetchone()
            if existing is None and last_monday >= reset_monday:
                last_sunday_str = (current_monday - _dt.timedelta(days=1)).isoformat()
                champ = conn.execute(
                    f"""
                    SELECT c.id AS candidate_id, c.name,
                           COALESCE(u.region_name, rc.region_name, c.region_code) AS region_name,
                           COALESCE(u.district_name, c.district_name) AS district_name,
                           c.election_type,
                           pa.status_note AS applicant_status_note,
                           ROUND(AVG(cp.total_score), 1) AS avg_score,
                           COUNT(cp.id) AS cnt
                    FROM candidates c
                    LEFT JOIN users u ON u.id = c.user_id
                    LEFT JOIN party_applicants pa ON pa.id = u.applicant_match_id
                    LEFT JOIN region_codes rc ON rc.region_code = c.region_code
                    JOIN candidate_pledges cp ON cp.candidate_id = c.id
                    LEFT JOIN ({approved_review_subquery}) prh ON prh.pledge_id = cp.id
                    WHERE cp.total_score IS NOT NULL
                      AND cp.approval_status = 'APPROVED'
                      AND c.approval_status = 'APPROVED'
        AND (
            u.applicant_match_id IS NOT NULL
            OR TRIM(COALESCE(pa.status_note, '')) = ''
            OR EXISTS (
                SELECT 1
                FROM party_applicants pa_public
                WHERE TRIM(COALESCE(pa_public.status_note, '')) = '공천 확정'
                  AND (
                      (
                          lower(trim(COALESCE(pa_public.email, ''))) <> ''
                          AND lower(trim(COALESCE(pa_public.email, ''))) = lower(trim(COALESCE(u.email, '')))
                          AND lower(replace(trim(COALESCE(pa_public.name, '')), ' ', '')) = lower(replace(trim(COALESCE(u.name, '')), ' ', ''))
                      )
                      OR (
                          replace(replace(replace(replace(trim(COALESCE(pa_public.phone, '')), '-', ''), ' ', ''), '(', ''), ')', '') <> ''
                          AND replace(replace(replace(replace(trim(COALESCE(pa_public.phone, '')), '-', ''), ' ', ''), '(', ''), ')', '') = replace(replace(replace(replace(trim(COALESCE(u.phone, '')), '-', ''), ' ', ''), '(', ''), ')', '')
                          AND lower(replace(trim(COALESCE(pa_public.name, '')), ' ', '')) = lower(replace(trim(COALESCE(u.name, '')), ' ', ''))
                      )
                  )
            )
        )
                      AND date(COALESCE(prh.approved_reviewed_at, cp.analyzed_at, cp.created_at)) >= ?
                      AND date(COALESCE(prh.approved_reviewed_at, cp.analyzed_at, cp.created_at)) <= ?
                    GROUP BY c.id
                    HAVING cnt > 0
                    ORDER BY avg_score DESC, cnt DESC
                    LIMIT 1
                    """,
                    (max(last_monday_str, RANKING_SCORE_START_DATE.isoformat()), last_sunday_str),
                ).fetchone()
                if champ and champ["avg_score"] and champ["avg_score"] > 0:
                    try:
                        conn.execute(
                            """INSERT OR IGNORE INTO weekly_champions
                               (week_start, candidate_id, candidate_name, region_name, district_name, election_type, avg_score, scored_pledge_count)
                               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                            (last_monday_str, champ["candidate_id"], champ["name"], champ["region_name"],
                             champ["district_name"], champ["election_type"], champ["avg_score"], champ["cnt"]),
                        )
                        conn.commit()
                    except Exception:
                        pass

        # ── 챔피언 목록 (최신 district_name 반영) ──
        champions = [
            dict(r) for r in conn.execute(
                """SELECT wc.week_start, wc.candidate_id, wc.candidate_name,
                          wc.region_name,
                          COALESCE(u.district_name, wc.district_name) AS district_name,
                          wc.election_type, wc.avg_score, wc.scored_pledge_count,
                          pa.status_note AS applicant_status_note,
                          u.applicant_match_id AS applicant_match_id
                   FROM weekly_champions wc
                   LEFT JOIN candidates c ON c.id = wc.candidate_id
                   LEFT JOIN users u ON u.id = c.user_id
                   LEFT JOIN party_applicants pa ON pa.id = u.applicant_match_id
                   WHERE wc.week_start >= ?
                   ORDER BY wc.week_start DESC LIMIT 20""",
                (reset_monday.isoformat(),)
            ).fetchall()
        ]
        champions = [champ for champ in champions if _is_public_nomination_status(champ.get("applicant_status_note"), champ.get("applicant_match_id"))]
        for champ in champions:
            try:
                champ_week = _dt.date.fromisoformat(champ["week_start"])
                champ["detail_as_of"] = f"{(champ_week + _dt.timedelta(days=6)).isoformat()} 23:59:59"
            except Exception:
                champ["detail_as_of"] = None

        # ── 랭킹 쿼리 ──
        sql = f"""
            SELECT c.id AS candidate_id, c.name,
                   COALESCE(u.region_name, rc.region_name, c.region_code) AS region_name,
                   COALESCE(u.district_name, c.district_name) AS district_name,
                   c.election_type,
                   pa.status_note AS applicant_status_note,
                   ROUND(AVG(cp.total_score), 1) AS avg_score,
                   COUNT(cp.id) AS scored_pledge_count,
                   c.updated_at
            FROM candidates c
            LEFT JOIN users u ON u.id = c.user_id
            LEFT JOIN party_applicants pa ON pa.id = u.applicant_match_id
            LEFT JOIN region_codes rc ON rc.region_code = c.region_code
            JOIN candidate_pledges cp ON cp.candidate_id = c.id
            LEFT JOIN ({approved_review_subquery}) prh ON prh.pledge_id = cp.id
            WHERE cp.total_score IS NOT NULL
              AND cp.approval_status = 'APPROVED'
              AND c.approval_status = 'APPROVED'
        AND (
            u.applicant_match_id IS NOT NULL
            OR TRIM(COALESCE(pa.status_note, '')) = ''
            OR EXISTS (
                SELECT 1
                FROM party_applicants pa_public
                WHERE TRIM(COALESCE(pa_public.status_note, '')) = '공천 확정'
                  AND (
                      (
                          lower(trim(COALESCE(pa_public.email, ''))) <> ''
                          AND lower(trim(COALESCE(pa_public.email, ''))) = lower(trim(COALESCE(u.email, '')))
                          AND lower(replace(trim(COALESCE(pa_public.name, '')), ' ', '')) = lower(replace(trim(COALESCE(u.name, '')), ' ', ''))
                      )
                      OR (
                          replace(replace(replace(replace(trim(COALESCE(pa_public.phone, '')), '-', ''), ' ', ''), '(', ''), ')', '') <> ''
                          AND replace(replace(replace(replace(trim(COALESCE(pa_public.phone, '')), '-', ''), ' ', ''), '(', ''), ')', '') = replace(replace(replace(replace(trim(COALESCE(u.phone, '')), '-', ''), ' ', ''), '(', ''), ')', '')
                          AND lower(replace(trim(COALESCE(pa_public.name, '')), ' ', '')) = lower(replace(trim(COALESCE(u.name, '')), ' ', ''))
                      )
                  )
            )
        )
        """
        params: list = []

        if is_current_week:
            sql += f"""
              AND EXISTS (
                  SELECT 1
                  FROM candidate_pledges cpw
                  LEFT JOIN ({approved_review_subquery}) prhw ON prhw.pledge_id = cpw.id
                  WHERE cpw.candidate_id = c.id
                    AND cpw.total_score IS NOT NULL
                    AND cpw.approval_status = 'APPROVED'
                    AND date(COALESCE(prhw.approved_reviewed_at, cpw.analyzed_at, cpw.created_at)) >= ?
                    AND date(COALESCE(prhw.approved_reviewed_at, cpw.analyzed_at, cpw.created_at)) <= ?
              )
            """
            params.extend([max(target_monday, RANKING_SCORE_START_DATE).isoformat(), ranking_end.isoformat()])
        else:
            sql += " AND date(COALESCE(prh.approved_reviewed_at, cp.analyzed_at, cp.created_at)) >= ? AND date(COALESCE(prh.approved_reviewed_at, cp.analyzed_at, cp.created_at)) <= ?"
            params.extend([max(target_monday, RANKING_SCORE_START_DATE).isoformat(), ranking_end.isoformat()])

        if is_current_week:
            # 이번 주: 챔피언 제외 (단, 이번 주에 공약을 업데이트한 챔피언은 다시 포함)
            champion_ids = [
                r["candidate_id"]
                for r in conn.execute(
                    "SELECT candidate_id FROM weekly_champions WHERE week_start >= ?",
                    (reset_monday.isoformat(),),
                ).fetchall()
            ]
            if champion_ids:
                updated_champ_ids = {
                    r["candidate_id"] for r in conn.execute(
                        f"SELECT id AS candidate_id FROM candidates WHERE id IN ({','.join('?' for _ in champion_ids)}) AND date(updated_at) >= ?",
                        (*champion_ids, target_monday.isoformat()),
                    ).fetchall()
                }
                exclude_ids = [cid for cid in champion_ids if cid not in updated_champ_ids]
                if exclude_ids:
                    placeholders = ",".join("?" for _ in exclude_ids)
                    sql += f" AND c.id NOT IN ({placeholders})"
                    params.extend(exclude_ids)

        if region_code:
            sql += " AND u.region_code = ?"
            params.append(region_code.strip())
        if election_type:
            sql += " AND c.election_type = ?"
            params.append(election_type.strip())

        sql += """
            GROUP BY c.id
            HAVING scored_pledge_count > 0
            ORDER BY avg_score DESC, scored_pledge_count DESC
            LIMIT 50
        """
        rows = conn.execute(sql, tuple(params)).fetchall()

        ranking = []
        detail_as_of = None if is_current_week else f"{target_sunday.isoformat()} 23:59:59"
        for rank, r in enumerate(rows, 1):
            ranking.append({
                "rank": rank,
                "candidate_id": r["candidate_id"],
                "name": r["name"],
                "region_name": r["region_name"] or "",
                "district_name": r["district_name"] or "",
                "election_type": r["election_type"] or "",
                "avg_score": float(r["avg_score"]),
                "scored_pledge_count": int(r["scored_pledge_count"]),
                "updated_at": r["updated_at"] or "",
                "detail_as_of": detail_as_of,
            })

        # ── 주차 라벨 + 네비게이션 ──
        label = _week_label(target_monday) + " 랭킹"
        prev_monday = target_monday - _dt.timedelta(days=7)
        next_monday = target_monday + _dt.timedelta(days=7)

        return {
            "week_label": label,
            "week_start": target_monday.isoformat(),
            "is_current_week": is_current_week,
            "prev_week": prev_monday.isoformat(),
            "next_week": next_monday.isoformat() if next_monday <= current_monday else None,
            "champions": champions,
            "ranking": ranking,
            "total_count": len(ranking),
        }
    finally:
        conn.close()


@app.get("/api/candidates/{candidate_id}/rank", tags=["candidates"])
def api_candidate_rank(candidate_id: int):
    """후보자의 현재 주 랭킹 순위 및 주간 챔피언 이력 반환 (공개 API)."""
    _ensure_db_ready()
    import datetime as _dt
    from backend.database import get_connection

    conn = get_connection()
    try:
        reset_monday = _leaderboard_reset_monday()
        today = _dt.date.today()
        current_monday = today - _dt.timedelta(days=today.weekday())

        # 점수 데이터가 있는지 확인
        has_score = conn.execute(
            "SELECT 1 FROM candidate_pledges WHERE candidate_id = ? AND total_score IS NOT NULL AND approval_status = 'APPROVED' LIMIT 1",
            (candidate_id,),
        ).fetchone() is not None

        if not has_score or today < reset_monday:
            return {
                "candidate_id": candidate_id,
                "current_week_rank": None,
                "current_week_label": _week_label(current_monday),
                "champion_weeks": [],
                "has_score": has_score,
            }

        # 챔피언 이력 조회 (최근 4주)
        four_weeks_ago = (current_monday - _dt.timedelta(days=28)).isoformat()
        champion_rows = conn.execute(
            """SELECT week_start FROM weekly_champions
               WHERE candidate_id = ? AND week_start >= ? AND week_start >= ?
               ORDER BY week_start DESC LIMIT 4""",
            (candidate_id, four_weeks_ago, reset_monday.isoformat()),
        ).fetchall()
        champion_weeks = [
            {
                "week_label": _week_label(_dt.date.fromisoformat(r["week_start"])),
                "week_start": r["week_start"],
            }
            for r in champion_rows
        ]

        # 이번 주 랭킹 계산 (leaderboard 쿼리 재활용)
        ranking_end = today
        approved_review_subquery = """
            SELECT pledge_id, MAX(reviewed_at) AS approved_reviewed_at
            FROM candidate_pledge_review_history
            WHERE approval_status = 'APPROVED'
              AND pledge_id IS NOT NULL
            GROUP BY pledge_id
        """
        sql = f"""
            SELECT c.id AS candidate_id,
                   ROUND(AVG(cp.total_score), 1) AS avg_score,
                   COUNT(cp.id) AS scored_pledge_count
            FROM candidates c
            LEFT JOIN users u ON u.id = c.user_id
            LEFT JOIN party_applicants pa ON pa.id = u.applicant_match_id
            JOIN candidate_pledges cp ON cp.candidate_id = c.id
            LEFT JOIN ({approved_review_subquery}) prh ON prh.pledge_id = cp.id
            WHERE cp.total_score IS NOT NULL
              AND cp.approval_status = 'APPROVED'
              AND c.approval_status = 'APPROVED'
              AND EXISTS (
                  SELECT 1
                  FROM candidate_pledges cpw
                  LEFT JOIN ({approved_review_subquery}) prhw ON prhw.pledge_id = cpw.id
                  WHERE cpw.candidate_id = c.id
                    AND cpw.total_score IS NOT NULL
                    AND cpw.approval_status = 'APPROVED'
                    AND date(COALESCE(prhw.approved_reviewed_at, cpw.analyzed_at, cpw.created_at)) >= ?
                    AND date(COALESCE(prhw.approved_reviewed_at, cpw.analyzed_at, cpw.created_at)) <= ?
              )
            GROUP BY c.id
            HAVING scored_pledge_count > 0
            ORDER BY avg_score DESC, scored_pledge_count DESC
            LIMIT 50
        """
        params = [max(current_monday, RANKING_SCORE_START_DATE).isoformat(), ranking_end.isoformat()]
        rows = conn.execute(sql, tuple(params)).fetchall()

        current_week_rank = None
        for rank, r in enumerate(rows, 1):
            if r["candidate_id"] == candidate_id:
                current_week_rank = rank
                break

        return {
            "candidate_id": candidate_id,
            "current_week_rank": current_week_rank,
            "current_week_label": _week_label(current_monday),
            "champion_weeks": champion_weeks,
            "has_score": has_score,
        }
    finally:
        conn.close()


@app.post("/api/admin/repair-pledge-scores", tags=["admin"])
def api_admin_repair_pledge_scores(request: Request):
    """analysis_result가 있지만 total_score가 NULL인 공약의 점수를 재파싱하여 복구한다."""
    require_admin(request)
    _ensure_db_ready()
    from backend.database import get_connection
    from backend.score_parser import parse_total_score

    conn = get_connection()
    try:
        rows = conn.execute(
            "SELECT id, analysis_result FROM candidate_pledges WHERE total_score IS NULL AND analysis_result IS NOT NULL AND analysis_result != ''"
        ).fetchall()
        repaired = 0
        details = []
        for r in rows:
            score = parse_total_score(r["analysis_result"])
            if score is not None:
                conn.execute(
                    "UPDATE candidate_pledges SET total_score = ? WHERE id = ?",
                    (score, r["id"]),
                )
                repaired += 1
                details.append({"pledge_id": r["id"], "score": score})
        conn.commit()
        return {"total_null": len(rows), "repaired": repaired, "details": details}
    finally:
        conn.close()

from backend.policy_admin_routes import register_policy_routes
register_policy_routes(app, require_admin, _ensure_db_ready, _serve_html)

from backend.tools_routes import register_tools_routes
register_tools_routes(app, require_approved, _client_ip)






class ContactRequest(BaseModel):
    name: str = Field('', max_length=100)
    email: str = Field(..., min_length=1, max_length=200)
    message: str = Field(..., min_length=1, max_length=5000)

@app.post('/api/contact')
def api_contact(body: ContactRequest, request: Request):
    import re as _re
    email = body.email.strip()
    if not _re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', email):
        raise HTTPException(status_code=400, detail='올바른 이메일 주소를 입력해 주세요.')
    ip = _client_ip(request)
    from backend.quota_rate import check_rate_limit_ip
    ok, msg = check_rate_limit_ip(ip)
    if not ok:
        raise HTTPException(status_code=429, detail='잠시 후 다시 시도해 주세요.')
    from backend.email_sender import send_contact_email
    success = send_contact_email(
        sender_name=body.name.strip(),
        sender_email=email,
        message=body.message.strip(),
    )
    if not success:
        raise HTTPException(status_code=500, detail='메일 발송에 실패했습니다. 잠시 후 다시 시도해 주세요.')
    return {'ok': True}

@app.get("/hub-briefing", include_in_schema=False)
def hub_briefing_page():
    from fastapi.responses import RedirectResponse
    return RedirectResponse("/hub", status_code=301)

@app.get("/hub/archive", include_in_schema=False)
def hub_archive_redirect():
    from fastapi.responses import RedirectResponse
    return RedirectResponse("/hub", status_code=301)

def _candidate_public_rows_sql() -> str:
    return f"""
        FROM candidates c
        LEFT JOIN users u ON u.id = c.user_id
        LEFT JOIN party_applicants pa ON pa.id = u.applicant_match_id
        WHERE c.approval_status IN ('APPROVED', 'MIXED')
          AND {_sql_public_nomination_condition('u')}
    """


def _legacy_public_nomination_status(status_note: Optional[str], applicant_match_id: Optional[int] = None) -> bool:
    if applicant_match_id is not None:
        return True
    note = (status_note or "").strip()
    return not note or note == PUBLIC_NOMINATION_NOTE


def get_regions():
    _ensure_db_ready()
    from backend.database import get_connection

    conn = get_connection()
    try:
        count_rows = conn.execute(
            """
            SELECT c.region_code, COUNT(*) AS candidate_count
            """
            + _candidate_public_rows_sql()
            + """
            GROUP BY c.region_code
            """
        ).fetchall()
        count_map = {r["region_code"]: int(r["candidate_count"]) for r in count_rows}
    finally:
        conn.close()

    return [
        RegionResponse(
            region_code=code,
            region_name=name,
            candidate_count=count_map.get(code, 0),
        )
        for code, name in REGION_NAME_MAP.items()
    ]


def get_election_type_counts(region_code: Optional[str] = Query(default=None)):
    _ensure_db_ready()
    from backend.database import get_connection

    conn = get_connection()
    try:
        sql = "SELECT c.election_type, COUNT(*) AS n " + _candidate_public_rows_sql()
        params: list[object] = []
        if region_code:
            sql += " AND c.region_code = ?"
            params.append(region_code)
        sql += " GROUP BY c.election_type"
        rows = conn.execute(sql, tuple(params)).fetchall()
        return {r["election_type"]: int(r["n"]) for r in rows}
    finally:
        conn.close()


def get_districts(
    region_code: Optional[str] = Query(default=None, description="?됱젙援ъ뿭 肄붾뱶"),
    election_type: Optional[str] = Query(default=None, description="?좉굅 ???local, mayor, etc)"),
):
    _ensure_db_ready()
    code = _validate_region_code(region_code)
    selected_election_type = _normalize_election_type(election_type)
    from backend.database import get_connection

    conn = get_connection()
    try:
        candidate_sql = """
            SELECT c.district_name, c.district_code, c.election_type,
                   pa.status_note AS applicant_status_note,
                   u.applicant_match_id AS applicant_match_id
            """
        candidate_sql += _candidate_public_rows_sql()
        candidate_sql += """
              AND c.region_code = ?
              AND c.district_name IS NOT NULL
              AND TRIM(c.district_name) <> ''
        """
        params: list[object] = [code]
        if selected_election_type:
            candidate_sql += " AND c.election_type = ?"
            params.append(selected_election_type)
        candidate_rows = conn.execute(candidate_sql, tuple(params)).fetchall()

        district_rows = conn.execute(
            """
            SELECT district_code, district_name
            FROM district_codes
            WHERE region_code = ?
              AND (? IS NULL OR election_type = ?)
            """,
            (code, selected_election_type, selected_election_type),
        ).fetchall()
    finally:
        conn.close()

    count_map: dict[str, dict[str, object]] = {}
    for r in district_rows:
        d_code = (r["district_code"] or "").strip()
        d_name = (r["district_name"] or "").strip() or d_code
        if d_code:
            count_map[d_code] = {"district_name": d_name, "candidate_count": 0}

    for r in candidate_rows:
        if not _legacy_public_nomination_status(r["applicant_status_note"], r["applicant_match_id"]):
            continue
        district_name = (r["district_name"] or "").strip()
        district_code = _derive_district_code(code, r["district_code"], district_name)
        if not district_code:
            continue
        if district_code not in count_map:
            count_map[district_code] = {
                "district_name": district_name or district_code,
                "candidate_count": 0,
            }
        count_map[district_code]["candidate_count"] = int(count_map[district_code]["candidate_count"]) + 1

    return [
        DistrictResponse(
            district_code=dcode,
            district_name=str(meta["district_name"]),
            region_code=code,
            candidate_count=int(meta["candidate_count"]),
        )
        for dcode, meta in sorted(count_map.items(), key=lambda x: (-int(x[1]["candidate_count"]), str(x[1]["district_name"])))
    ]


def get_candidates(
    region_code: Optional[str] = Query(default=None, description="?됱젙援ъ뿭 肄붾뱶"),
    district_code: Optional[str] = Query(default=None, description="?좉굅援?肄붾뱶"),
    election_type: Optional[str] = Query(default=None, description="?좉굅 ???local, mayor, etc)"),
):
    _ensure_db_ready()
    code = _validate_region_code(region_code)
    selected_district_code = _normalize_district_code(district_code)
    selected_election_type = _normalize_election_type(election_type)
    from backend.database import get_connection

    conn = get_connection()
    try:
        sql = """
            SELECT c.id, c.name,
                   COALESCE(u.district_name, c.district_name) AS district_name,
                   c.district_code, c.region_code, c.election_type, c.election_level,
                   pa.status_note AS applicant_status_note,
                   u.applicant_match_id AS applicant_match_id
            """
        sql += _candidate_public_rows_sql()
        sql += " AND c.region_code = ?"
        params: list[object] = [code]
        if selected_election_type:
            sql += " AND c.election_type = ?"
            params.append(selected_election_type)
        sql += """
            ORDER BY
                CASE c.election_type
                    WHEN 'metro_mayor' THEN 1
                    WHEN 'local_mayor' THEN 2
                    WHEN 'regional_council' THEN 3
                    WHEN 'local_council' THEN 4
                    ELSE 5
                END,
                COALESCE(c.district_name, '') ASC,
                c.name ASC
        """
        rows = conn.execute(sql, tuple(params)).fetchall()
    finally:
        conn.close()

    result: list[CandidateListItemResponse] = []
    for r in rows:
        if not _legacy_public_nomination_status(r["applicant_status_note"], r["applicant_match_id"]):
            continue
        candidate_id = int(r["id"])
        resolved_district_code = _derive_district_code(code, r["district_code"], r["district_name"])
        if selected_district_code and resolved_district_code != selected_district_code:
            continue
        result.append(
            CandidateListItemResponse(
                candidate_id=candidate_id,
                name=r["name"],
                district_name=r["district_name"],
                district_code=resolved_district_code,
                region_code=r["region_code"],
                election_type=r["election_type"],
                election_level=r["election_level"],
                pledges=_fetch_candidate_pledges(candidate_id, limit=3, public_only=True),
            )
        )
    return result


def get_candidate_detail(candidate_id: int, as_of: Optional[str] = Query(default=None)):
    _ensure_db_ready()
    resolved_as_of = as_of if isinstance(as_of, str) else None
    from backend.database import get_connection

    conn = get_connection()
    try:
        row = conn.execute(
            """
            SELECT c.id, c.name,
                   COALESCE(u.district_name, c.district_name) AS district_name,
                   c.district_code, c.region_code, c.election_type, c.election_level, c.approval_status, c.user_id,
                   pa.status_note AS applicant_status_note,
                   u.applicant_match_id AS applicant_match_id
            """
            + _candidate_public_rows_sql()
            + """
              AND c.id = ?
            """,
            (candidate_id,),
        ).fetchone()
    finally:
        conn.close()

    if row is None or not _legacy_public_nomination_status(row["applicant_status_note"], row["applicant_match_id"]):
        raise HTTPException(status_code=404, detail=f"candidate_id={candidate_id} not found.")

    code = row["region_code"]
    return CandidateDetailResponse(
        candidate_id=int(row["id"]),
        name=row["name"],
        district_name=row["district_name"],
        district_code=_derive_district_code(code, row["district_code"], row["district_name"]),
        region_code=code,
        region_name=_resolve_region_name(code),
        election_type=row["election_type"],
        election_level=row["election_level"],
        external_profile=_fetch_candidate_external_profile(int(row["id"])),
        pledges=_fetch_candidate_pledges_snapshot(int(row["id"]), resolved_as_of, int(row["user_id"]) if row["user_id"] is not None else None) if resolved_as_of else _fetch_candidate_pledges_current_public(int(row["id"]), limit=None),
    )


def _get_candidate_detail_any_status(candidate_id: int) -> CandidateDetailResponse:
    from backend.database import get_connection

    conn = get_connection()
    try:
        row = conn.execute(
            """SELECT c.id, c.name,
                      COALESCE(u.district_name, c.district_name) AS district_name,
                      c.district_code, c.region_code, c.election_type, c.election_level
               FROM candidates c
               LEFT JOIN users u ON u.id = c.user_id
               WHERE c.id = ?""",
            (candidate_id,),
        ).fetchone()
    finally:
        conn.close()

    if row is None:
        raise HTTPException(status_code=404, detail=f"candidate_id={candidate_id} not found.")

    code = row["region_code"]
    return CandidateDetailResponse(
        candidate_id=int(row["id"]),
        name=row["name"],
        district_name=row["district_name"],
        district_code=_derive_district_code(code, row["district_code"], row["district_name"]),
        region_code=code,
        region_name=_resolve_region_name(code),
        election_type=row["election_type"],
        election_level=row["election_level"],
        external_profile=_fetch_candidate_external_profile(int(row["id"])),
        pledges=_fetch_candidate_pledges(int(row["id"]), limit=None, public_only=False),
    )
