#!/usr/bin/env python3
"""
district_dong_map.json 완전 재생성 스크립트

데이터 소스:
1. 공공데이터포털 제8회 전국동시지방선거 개표결과 엑셀 (886/671개)
2. NEC 선거통계시스템 제9회 BIGI05 스크래핑 (무투표 선거구 보완)
3. 공공데이터포털 API getCommonSggCodeList (마스터 검증)

핵심 원리:
- 8회와 9회의 sggName(선거구명)은 동일
- sdName(시도명)/wiwName(구시군명)만 변경 (강원도->강원특별자치도, 부천시->부천시원미구 등)
- 8회/9회 API를 비교해 정확한 매핑 테이블 구축

사용법:
  python scripts/rebuild_district_dong_map.py
"""
from __future__ import annotations

import io
import json
import os
import re
import sys
import time
from collections import OrderedDict, defaultdict
from pathlib import Path
from urllib.parse import urlencode
from urllib.request import Request, urlopen

import requests
from bs4 import BeautifulSoup

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
XLSX_PATH = DATA_DIR / "temp_vote_result.xlsx"
OUTPUT_PATH = DATA_DIR / "district_dong_map.json"

NEC_BASE = "http://info.nec.go.kr"
NEC_EID = "0020260603"  # 9회 (BIGI05 접근 가능)
SKIP_DONG = {"합계", "거소투표", "관외사전투표", "관내사전투표", "", "소계"}

SIDO_ORDER = [
    "서울특별시", "부산광역시", "대구광역시", "인천광역시",
    "광주광역시", "대전광역시", "울산광역시",
    "세종특별자치시", "경기도", "강원도",
    "충청북도", "충청남도", "전라북도",
    "전라남도", "경상북도", "경상남도", "제주특별자치도",
]

# NEC cityCode -> 9회 시도명
CITY_CODE_TO_SIDO_9 = {
    "1100": "서울특별시", "2600": "부산광역시", "2700": "대구광역시",
    "2800": "인천광역시", "2900": "광주광역시", "3000": "대전광역시",
    "3100": "울산광역시", "5100": "세종특별자치시",
    "4100": "경기도", "5200": "강원특별자치도",
    "4300": "충청북도", "4400": "충청남도", "5300": "전북특별자치도",
    "4600": "전라남도", "4700": "경상북도", "4800": "경상남도",
    "4900": "제주특별자치도",
}


def load_env_key() -> str:
    env_path = ROOT / ".env"
    if not env_path.exists():
        return ""
    for line in env_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        if k.strip() == "DATA_GO_KR_API_KEY":
            return v.strip().strip('"').strip("'").replace("\r", "").replace("\n", "")
    return ""


def sido_sort_key(n):
    try: return (SIDO_ORDER.index(n),)
    except ValueError: return (99,)


def district_sort_key(n):
    ga = list("가나다라마바사아자차카타파하")
    if n and n[0] in ga and "선거구" in n:
        return (0, ga.index(n[0]), n)
    m = re.match(r"제(\d+)", n)
    if m: return (1, int(m.group(1)), n)
    return (2, 0, n)


def strip_prefix(sgg, wiw):
    if sgg.startswith(wiw): return sgg[len(wiw):]
    i = sgg.find(wiw)
    if i >= 0: return sgg[i + len(wiw):]
    return sgg


# ──────────────────────────────────────────
# API helpers
# ──────────────────────────────────────────

def fetch_api(api_key, sg_typecode, sg_id):
    items = []
    page = 1
    while True:
        rest = urlencode({"sgId": sg_id, "sgTypecode": sg_typecode, "pageNo": page, "numOfRows": 500, "resultType": "json"})
        url = f"https://apis.data.go.kr/9760000/CommonCodeService/getCommonSggCodeList?ServiceKey={api_key}&{rest}"
        req = Request(url, headers={"User-Agent": "Mozilla/5.0", "Accept": "application/json"})
        with urlopen(req, timeout=30) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        body = data["response"]["body"]
        batch = body["items"]["item"]
        if isinstance(batch, dict): batch = [batch]
        items.extend(batch)
        if len(items) >= body["totalCount"]: break
        page += 1
        time.sleep(0.3)
    return items


def build_9to8_mapping(api_key, sg_typecode):
    """Build (9th_sdName, sggName) -> (8th_sdName, 8th_wiwName) mapping.
    Since sggName can repeat across different 시도 (e.g. 남구가선거구),
    we need to match 8th and 9th items carefully.
    Returns also the 8th master dict: sggName_full -> (sdName, wiwName) keyed by (sdName, sggName).
    """
    items_8 = fetch_api(api_key, sg_typecode, "20220601")
    time.sleep(0.5)
    items_9 = fetch_api(api_key, sg_typecode, "20260603")

    # Group by sggName
    by_sgg_8 = defaultdict(list)
    by_sgg_9 = defaultdict(list)
    for it in items_8: by_sgg_8[it["sggName"]].append(it)
    for it in items_9: by_sgg_9[it["sggName"]].append(it)

    # mapping: (9th_sdName, sggName) -> (8th_sdName, 8th_wiwName)
    mapping = {}
    for sgg_name in by_sgg_9:
        list_9 = by_sgg_9[sgg_name]
        list_8 = by_sgg_8.get(sgg_name, [])
        if not list_8:
            continue  # 9th-only district
        if len(list_9) == len(list_8):
            s9 = sorted(list_9, key=lambda x: int(x.get("sOrder", 0)))
            s8 = sorted(list_8, key=lambda x: int(x.get("sOrder", 0)))
            for i9, i8 in zip(s9, s8):
                mapping[(i9["sdName"], sgg_name)] = (i8["sdName"], i8["wiwName"])
        else:
            # Fallback: match by position
            for i9, i8 in zip(list_9, list_8):
                mapping[(i9["sdName"], sgg_name)] = (i8["sdName"], i8["wiwName"])

    # Also build 8th master: (8th_sdName, sggName) -> (8th_sdName, 8th_wiwName)
    master_8 = {}
    for it in items_8:
        master_8[(it["sdName"], it["sggName"])] = (it["sdName"], it["wiwName"])

    return mapping, master_8, len(items_8)


# ──────────────────────────────────────────
# Excel extraction
# ──────────────────────────────────────────

def download_xlsx():
    if XLSX_PATH.exists() and XLSX_PATH.stat().st_size > 100000:
        print(f"엑셀 존재: {XLSX_PATH.stat().st_size:,} bytes")
        return True
    print("엑셀 다운로드 중...")
    try:
        r = requests.get(
            "https://www.data.go.kr/cmm/cmm/fileDownload.do",
            params={"atchFileId": "FILE_000000003157459", "fileDetailSn": "1", "insertDataPrcus": "N"},
            headers={"User-Agent": "Mozilla/5.0", "Referer": "https://www.data.go.kr/data/15101509/fileData.do"},
            timeout=120, stream=True,
        )
        r.raise_for_status()
        with open(XLSX_PATH, "wb") as f:
            for chunk in r.iter_content(8192): f.write(chunk)
        print(f"다운로드 완료: {XLSX_PATH.stat().st_size:,} bytes")
        return True
    except Exception as e:
        print(f"실패: {e}", file=sys.stderr)
        return False


def extract_excel(master_8_local, master_8_regional):
    """Extract from Excel using 8th master for correct sido/wiw names."""
    import openpyxl
    print("\n[1단계] 엑셀 추출")
    wb = openpyxl.load_workbook(str(XLSX_PATH), read_only=True)

    local = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    for row in wb["구·시·군의회의원"].iter_rows(min_row=3, values_only=True):
        sido, gugun, sgg, dong, gubun = ((row[i] or "").strip() for i in range(5))
        if not all([sido, gugun, sgg, dong]) or dong in SKIP_DONG or gubun != "소계":
            continue
        key = (sido, sgg)
        if key in master_8_local:
            sd8, wiw8 = master_8_local[key]
            short = strip_prefix(sgg, wiw8)
            if dong not in local[sd8][wiw8][short]:
                local[sd8][wiw8][short].append(dong)

    regional = defaultdict(lambda: defaultdict(list))
    for row in wb["시·도의회의원"].iter_rows(min_row=3, values_only=True):
        sido, gugun, sgg, dong, gubun = ((row[i] or "").strip() for i in range(5))
        if not all([sido, sgg, dong]) or dong in SKIP_DONG or gubun != "소계":
            continue
        key = (sido, sgg)
        if key in master_8_regional:
            sd8, _ = master_8_regional[key]
            if dong not in regional[sd8][sgg]:
                regional[sd8][sgg].append(dong)

    wb.close()
    lc = sum(len(d) for g in local.values() for d in g.values())
    rc = sum(len(d) for d in regional.values())
    print(f"  기초: {lc}개, 광역: {rc}개")
    return dict(local), dict(regional)


# ──────────────────────────────────────────
# NEC scraping
# ──────────────────────────────────────────

def get_city_town_codes():
    h = {"User-Agent": "Mozilla/5.0"}
    r = requests.get(f"{NEC_BASE}/main/showDocument.xhtml",
                     params={"electionId": NEC_EID, "topMenuId": "BI", "secondMenuId": "BIGI05"}, headers=h, timeout=15)
    soup = BeautifulSoup(r.text, "html.parser")
    sel = soup.find("select", {"name": "cityCode"})
    cities = [(o.get("value"), o.text.strip()) for o in sel.find_all("option") if o.get("value") and o.get("value") != "-1"]
    result = []
    for cc, cn in cities:
        r2 = requests.get(f"{NEC_BASE}/bizcommon/selectbox/selectbox_townCodeJson.json",
                          params={"electionId": NEC_EID, "cityCode": cc}, headers=h, timeout=10)
        towns = [(t["CODE"], t["NAME"]) for t in r2.json()["jsonResult"]["body"]]
        result.append((cc, cn, towns))
        time.sleep(0.1)
    return result


def scrape_page(city_code, town_code, ec):
    url = f"{NEC_BASE}/electioninfo/electionInfo_report.xhtml"
    data = {
        "electionId": NEC_EID, "requestURI": f"/electioninfo/{NEC_EID}/bi/bigi05.jsp",
        "topMenuId": "BI", "secondMenuId": "BIGI05", "menuId": "BIGI05",
        "statementId": "BIGI05_2", "electionCode": ec,
        "cityCode": city_code, "townCode": town_code,
    }
    h = {"User-Agent": "Mozilla/5.0", "Content-Type": "application/x-www-form-urlencoded",
         "Referer": f"{NEC_BASE}/main/showDocument.xhtml?electionId={NEC_EID}&topMenuId=BI&secondMenuId=BIGI05"}
    try:
        r = requests.post(url, data=data, headers=h, timeout=15)
        if r.status_code != 200: return []
    except Exception: return []
    soup = BeautifulSoup(r.text, "html.parser")
    tables = soup.find_all("table")
    if not tables: return []
    out = []
    for row in tables[0].find_all("tr"):
        cells = [td.get_text(strip=True) for td in row.find_all(["td", "th"])]
        if len(cells) >= 4 and cells[0] != "구시군명":
            out.append((cells[0], cells[1], cells[2], cells[3]))
    return out


def scrape_nec(map_9to8_local, map_9to8_regional, master_8_local, master_8_regional):
    """Scrape NEC 9th BIGI05, map to 8th district names."""
    print("\n[2단계] NEC 스크래핑")
    cts = get_city_town_codes()

    local = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    regional = defaultdict(lambda: defaultdict(list))
    stats = {"lc_ok": 0, "lc_skip": 0, "rc_ok": 0, "rc_skip": 0}

    for city_code, city_name, towns in cts:
        sido_9 = CITY_CODE_TO_SIDO_9.get(city_code, city_name)
        targets = towns if towns else [("", city_name)]
        print(f"  {city_name} ({city_code}): {len(targets)}개", end="", flush=True)

        for tc, tn in targets:
            # 기초의원
            for _, sgg_nec, _, dong_csv in scrape_page(city_code, tc, "6"):
                dongs = [d.strip() for d in dong_csv.split(",") if d.strip()]
                if not dongs: continue
                key9 = (sido_9, sgg_nec)
                if key9 in map_9to8_local:
                    sd8, wiw8 = map_9to8_local[key9]
                    short = strip_prefix(sgg_nec, wiw8)
                    for dong in dongs:
                        if dong not in local[sd8][wiw8][short]:
                            local[sd8][wiw8][short].append(dong)
                    stats["lc_ok"] += 1
                else:
                    stats["lc_skip"] += 1
            time.sleep(0.12)

            # 광역의원
            for _, sgg_nec, _, dong_csv in scrape_page(city_code, tc, "5"):
                dongs = [d.strip() for d in dong_csv.split(",") if d.strip()]
                if not dongs: continue
                key9 = (sido_9, sgg_nec)
                if key9 in map_9to8_regional:
                    sd8, _ = map_9to8_regional[key9]
                    if dong not in regional[sd8][sgg_nec]:  # Bug: need to add all dongs
                        pass
                    # Add all dongs properly
                    for dong in dongs:
                        if dong not in regional[sd8][sgg_nec]:
                            regional[sd8][sgg_nec].append(dong)
                    stats["rc_ok"] += 1
                else:
                    stats["rc_skip"] += 1
            time.sleep(0.12)

        print(" ok")

    lc = sum(len(d) for g in local.values() for d in g.values())
    rc = sum(len(d) for d in regional.values())
    print(f"  기초: {lc}개 (ok={stats['lc_ok']}, skip={stats['lc_skip']})")
    print(f"  광역: {rc}개 (ok={stats['rc_ok']}, skip={stats['rc_skip']})")
    return dict(local), dict(regional)


# ──────────────────────────────────────────
# Merge & output
# ──────────────────────────────────────────

def merge(el, er, nl, nr):
    out = {"local_council": OrderedDict(), "regional_council": OrderedDict()}
    s = {"e": 0, "n": 0, "b": 0}
    for sido in sorted(set(list(el) + list(nl)), key=sido_sort_key):
        out["local_council"][sido] = OrderedDict()
        eg, ng = el.get(sido, {}), nl.get(sido, {})
        for gg in sorted(set(list(eg) + list(ng))):
            out["local_council"][sido][gg] = OrderedDict()
            ed, nd = eg.get(gg, {}), ng.get(gg, {})
            for d in sorted(set(list(ed) + list(nd)), key=district_sort_key):
                if d in ed:
                    out["local_council"][sido][gg][d] = ed[d]
                    s["b" if d in nd else "e"] += 1
                else:
                    out["local_council"][sido][gg][d] = nd[d]
                    s["n"] += 1
    print(f"\n  기초 병합: 양쪽={s['b']}, 엑셀만={s['e']}, NEC만={s['n']}")

    s2 = {"e": 0, "n": 0, "b": 0}
    for sido in sorted(set(list(er) + list(nr)), key=sido_sort_key):
        out["regional_council"][sido] = OrderedDict()
        ed, nd = er.get(sido, {}), nr.get(sido, {})
        for d in sorted(set(list(ed) + list(nd))):
            if d in ed:
                out["regional_council"][sido][d] = ed[d]
                s2["b" if d in nd else "e"] += 1
            else:
                out["regional_council"][sido][d] = nd[d]
                s2["n"] += 1
    print(f"  광역 병합: 양쪽={s2['b']}, 엑셀만={s2['e']}, NEC만={s2['n']}")
    return out


def validate(data, master_8_local, master_8_regional):
    print("\n[검증]")
    our_lc = set()
    for sido, gg in data["local_council"].items():
        for gugun, dists in gg.items():
            for d in dists:
                our_lc.add(f"{sido}|{gugun}|{d}")
    master_lc = set()
    for (sd, sgg), (sd8, wiw8) in master_8_local.items():
        short = strip_prefix(sgg, wiw8)
        master_lc.add(f"{sd8}|{wiw8}|{short}")
    miss_lc = master_lc - our_lc
    print(f"  기초: 마스터 {len(master_lc)}, 생성 {len(our_lc)}, 누락 {len(miss_lc)}")
    for m in sorted(miss_lc)[:10]: print(f"    {m}")

    our_rc = set()
    for sido, dists in data["regional_council"].items():
        for d in dists: our_rc.add(f"{sido}|{d}")
    master_rc = set()
    for (sd, sgg), (sd8, _) in master_8_regional.items():
        master_rc.add(f"{sd8}|{sgg}")
    miss_rc = master_rc - our_rc
    print(f"  광역: 마스터 {len(master_rc)}, 생성 {len(our_rc)}, 누락 {len(miss_rc)}")
    for m in sorted(miss_rc)[:10]: print(f"    {m}")

    # Empty dong lists
    issues = 0
    for sido, gg in data["local_council"].items():
        for gugun, dists in gg.items():
            for d, dl in dists.items():
                if not dl:
                    print(f"  [!] 빈 동: {sido} {gugun} {d}")
                    issues += 1
    for sido, dists in data["regional_council"].items():
        for d, dl in dists.items():
            if not dl:
                print(f"  [!] 빈 동: {sido} {d}")
                issues += 1
    if issues: print(f"  빈 동 목록: {issues}건")


def compare_old(data):
    if not OUTPUT_PATH.exists(): return
    try:
        old = json.loads(OUTPUT_PATH.read_text(encoding="utf-8"))
    except Exception: return
    print("\n" + "=" * 60)
    print("기존 대비 변경사항")
    for sec, lbl in [("local_council", "기초"), ("regional_council", "광역")]:
        od, nd = set(), set()
        if sec == "local_council":
            for s, gg in old.get(sec, {}).items():
                for g, ds in gg.items():
                    for d in ds: od.add(f"{s}|{g}|{d}")
            for s, gg in data.get(sec, {}).items():
                for g, ds in gg.items():
                    for d in ds: nd.add(f"{s}|{g}|{d}")
        else:
            for s, ds in old.get(sec, {}).items():
                for d in ds: od.add(f"{s}|{d}")
            for s, ds in data.get(sec, {}).items():
                for d in ds: nd.add(f"{s}|{d}")
        added, removed = nd - od, od - nd
        print(f"\n[{lbl}] {len(od)} -> {len(nd)} (+{len(added)} -{len(removed)})")
        if added:
            for a in sorted(added)[:20]: print(f"  + {a}")
            if len(added) > 20: print(f"  ... +{len(added)-20}개")


def main():
    api_key = os.environ.get("DATA_GO_KR_API_KEY") or load_env_key()
    if not api_key:
        print("DATA_GO_KR_API_KEY 필요", file=sys.stderr)
        return 1

    print("=" * 60)
    print("district_dong_map.json 재생성")
    print("=" * 60)

    # Build 9th->8th mapping for both local and regional
    print("\n[마스터] 8회/9회 API 매핑 구축")
    print("  기초의원...")
    map_9to8_lc, master_8_lc, n_lc = build_9to8_mapping(api_key, 6)
    print(f"  기초: {n_lc}개 (매핑 {len(map_9to8_lc)}개)")
    time.sleep(0.5)
    print("  광역의원...")
    map_9to8_rc, master_8_rc, n_rc = build_9to8_mapping(api_key, 5)
    print(f"  광역: {n_rc}개 (매핑 {len(map_9to8_rc)}개)")

    # Save old for comparison
    compare_old_data = None
    if OUTPUT_PATH.exists():
        compare_old_data = True

    # Excel
    if download_xlsx():
        el, er = extract_excel(master_8_lc, master_8_rc)
    else:
        el, er = {}, {}

    # NEC scraping
    nl, nr = scrape_nec(map_9to8_lc, map_9to8_rc, master_8_lc, master_8_rc)

    # Merge
    print("\n[3단계] 병합")
    result = merge(el, er, nl, nr)

    lc_t = sum(len(d) for g in result["local_council"].values() for d in g.values())
    rc_t = sum(len(d) for d in result["regional_council"].values())
    lc_d = sum(len(dl) for g in result["local_council"].values() for d in g.values() for dl in d.values())
    rc_d = sum(len(dl) for d in result["regional_council"].values() for dl in d.values())
    print(f"\n  기초: {lc_t}개 선거구, {lc_d}개 읍면동")
    print(f"  광역: {rc_t}개 선거구, {rc_d}개 읍면동")

    validate(result, master_8_lc, master_8_rc)
    if compare_old_data:
        compare_old(result)

    # Save
    OUTPUT_PATH.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n저장: {OUTPUT_PATH} ({OUTPUT_PATH.stat().st_size:,} bytes)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
